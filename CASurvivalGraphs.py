import pandas as pd
import matplotlib.pyplot as plt
import pandas as pd
import gc
import json
# import config
from openpyxl import load_workbook
from redcap import Project, RedcapError
from datetime import datetime
from dateutil import relativedelta
from datetime import timedelta
from smtplib import SMTP
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.drawing.image import Image
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles import  Alignment,Font
from openpyxl.worksheet.pagebreak import Break
import statistics

URL = 'https://redcap.smh.ca/redcap/api/'

API_KEY = '' #CANROCPROJECT
project = Project(URL, API_KEY)

def CreateSiteLevelCharts(site,ax):

    # len(df.index) == 0
    start_date = '2018-SEP-01'
    end_date = '2019-SEP-30'
    # start_date =  input ("Enter start date as yyyy-mmm/dd:")
    # end_date = input ("Enter end date : as yyyy-mmm/dd:")
    allfields = ['cr_surv','cr_scause','cr_tx','cr_frhyem','cr_epdt','cr_witbys', 'cr_cpratt', 'cr_rig1tm','cr_ptmrcv']
    allrhythmsList = []
    asystoleList = []
    notShockableList = []
    peaList = []
    vfvtShockableList = []
    can_vfvtShockableList = []

    allrhythmsTable = []
    asystoleTable = []
    notShockableTable = []
    peaTable = []
    vfvtShockableTable = []
    can_vfvtShockableTable = []

    df = project.export_records(format='df', fields=allfields)

    canadian_df=df.copy()
    canadian_df['year'] = canadian_df.cr_epdt
    canadian_df["date"] = pd.DatetimeIndex(canadian_df["cr_epdt"]).year



    df= df.reset_index()

    df = df[df["cr_record_id"].str.contains(site)]
    df.reset_index(drop=True, inplace=True)
    df['year'] = df.cr_epdt
    df["date"] = pd.DatetimeIndex(df["cr_epdt"]).year
    years = list(df.date.unique())
    years = [x for x in years if str(x) != 'nan']
    years = sorted(years)

    print(years)

    Fig1Table = []
    Fig2Table = []
    Fig4Table = []
    Fig3Table = []




    for year in years:

        year = int(year)


        number_of_rows = len(df.index)





        allrhythms_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 17) & (df['date'] == year)].index)
        all_rhythms = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 17) & (df['cr_surv'] == 1)& (df['date'] == year) ].index)

        if allrhythms_total  != 0:
            allrhythyms_percent = round((all_rhythms/allrhythms_total) * 100,2)
        else:
            allrhythyms_percent = 0
        allrhythmsList.append(allrhythyms_percent)

        allrhythms_line = {}
        allrhythms_line['EpisodeYear'] = year
        allrhythms_line['Rhythm'] = 'All rhythms'
        allrhythms_line['Survival Rate'] = allrhythyms_percent
        allrhythms_line['Alive'] = all_rhythms
        allrhythms_line['Treated'] = allrhythms_total
        allrhythmsTable.append(allrhythms_line)


        asystole_total =  len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 17) & (df['cr_frhyem'] == 0) & (df['date'] == year) ].index)
        asystole =  len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 17) & (df['cr_surv'] == 1) & (df['cr_frhyem'] == 0)& (df['date'] == year) ].index)


        if asystole_total  != 0:
            asystole_percent = round((asystole/asystole_total) * 100,2)
        else:
            asystole_percent = 0
        asystoleList.append(asystole_percent)

        print('Asystole')

        asystole_line = {}
        asystole_line['EpisodeYear'] = year
        asystole_line['Rhythm'] = 'Asystole'
        asystole_line['Survival Rate'] =  asystole_percent
        asystole_line['Alive'] = asystole
        asystole_line['Treated'] = asystole_total
        asystoleTable.append(asystole_line)

        notshockable_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 17) & (df['cr_frhyem'] == 4)& (df['date'] == year)].index)
        notshockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 17) & (df['cr_surv'] == 1) & (df['cr_frhyem'] == 4)& (df['date'] == year)].index)
        print("notshockable")


        if notshockable_total  != 0:
            notshockable_percent = round((notshockable / notshockable_total) * 100,2)
        else:
            notshockable_percent = 0
        notShockableList.append(notshockable_percent)


        notshockable_line = {}
        notshockable_line['EpisodeYear'] = year
        notshockable_line['Rhythm'] = 'Not shockable'
        notshockable_line['Survival Rate'] = notshockable_percent
        notshockable_line['Alive'] = notshockable
        notshockable_line['Treated'] = notshockable_total
        notShockableTable.append(notshockable_line)



        pea_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 17) & (df['cr_frhyem'] == 1)& (df['date'] == year)].index)
        pea = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 17) & (df['cr_surv'] == 1) & (df['cr_frhyem'] == 1)& (df['date'] == year)].index)


        if pea_total  != 0:
            pea_percent = round((pea / pea_total) * 100,2)
        else:
            pea_percent = 0
        peaList.append(pea_percent)


        pea_line = {}
        pea_line['EpisodeYear'] = year
        pea_line['Rhythm'] = 'PEA'
        pea_line['Survival Rate'] = pea_percent
        pea_line['Alive'] = pea
        pea_line['Treated'] = pea_total
        peaTable.append(pea_line)

        value_list = [2,3,5]
        vfvtshockable_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 17) &  (df['date'] == year)& (df.cr_frhyem.isin(value_list))].index)
        vfvtshockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 17) &  (df['date'] == year) &(df['cr_surv'] == 1) &(df.cr_frhyem.isin(value_list))].index)


        if vfvtshockable_total != 0:
            vfvtshockable_percent = round((vfvtshockable / vfvtshockable_total) * 100,2)
        else:
            vfvtshockable_percent = 0
        vfvtShockableList.append(vfvtshockable_percent)

        vfvtshockable_line = {}
        vfvtshockable_line['EpisodeYear'] = year
        vfvtshockable_line['Rhythm'] = 'VF/VT/Shockable'
        vfvtshockable_line['Survival Rate'] = vfvtshockable_percent
        vfvtshockable_line['Alive'] = vfvtshockable
        vfvtshockable_line['Treated'] = vfvtshockable_total
        vfvtShockableTable.append( vfvtshockable_line)



        can_vfvtshockable_total = len(canadian_df.loc[(canadian_df['cr_tx'] == 1) & (canadian_df['cr_scause'] != 17) & (canadian_df['date'] == year) & (
            canadian_df.cr_frhyem.isin(value_list))].index)
        can_vfvtshockable = len(canadian_df.loc[(canadian_df['cr_tx'] == 1) & (canadian_df['cr_scause'] != 17) & (canadian_df['date'] == year) & (
                    canadian_df['cr_surv'] == 1) & (canadian_df.cr_frhyem.isin(value_list))].index)



        if can_vfvtshockable_total != 0:
            can_vfvtshockable_percent = round((can_vfvtshockable / can_vfvtshockable_total) * 100, 2)
        else:
            can_vfvtshockable_percent = 0
        can_vfvtShockableList.append(can_vfvtshockable_percent)

        can_vfvtshockable_line = {}
        can_vfvtshockable_line['EpisodeYear'] = year
        can_vfvtshockable_line['Rhythm'] = 'Canadian VF/VT/Shockable'
        can_vfvtshockable_line['Survival Rate'] = can_vfvtshockable_percent
        can_vfvtshockable_line['Alive'] = can_vfvtshockable
        can_vfvtshockable_line['Treated'] = can_vfvtshockable
        can_vfvtShockableTable.append(can_vfvtshockable_line)





    Fig1Table.extend(allrhythmsTable)
    Fig1Table.extend(asystoleTable)
    Fig1Table.extend(notShockableTable)
    Fig1Table.extend(peaTable)
    Fig1Table.extend(vfvtShockableTable)
    Fig1Table.extend(can_vfvtShockableTable)








    x = np.arange(len(years))  # the label locations
    width = 0.05  # the width of the bars
    print("TOWWWWW")
    print(x)
    print(type(x))


    # fig, ax = plt.subplots()
    rects1 = ax.bar(x - width / 2, allrhythmsList, width, label='All Rhythms')
    rects2 = ax.bar(x + width / 2, asystoleList, width, label='Asystole')
    rects3 = ax.bar(x + 1.5 * width, notShockableList, width, label='notShockable')
    rects4 = ax.bar(x + 2.5 * width, peaList, width, label='Pea')
    rects5 = ax.bar(x + 3.5 * width, vfvtShockableList, width, label='vfvtShockableList')
    rects5 = ax.bar(x + 4.5 * width, can_vfvtShockableList, width, label='can_vfvtShockableList')

    # Add some text for labels, title and custom x-axis tick labels, etc.
    # ax.set_ylabel('Scores')

    ax.set_title('Figure 1 \n EMS Treated, All patients, Non traumatic - Survival Rate (%) by Year', fontsize = 8)
    ax.set_xticks(x)
    ax.set_xticklabels(years)
    ax.legend()
    labels = ['All Rhythms','Asystole','Not Shockable','Pea','vfvtShockableList','can_vfvtShockableList']
    plt.legend(labels, loc="lower left", bbox_to_anchor=(0, -.20) , ncol = 3,fontsize = 7) #, bbox_to_anchor=(0.5, -0.3))
    plt.xticks(fontsize=7)
    plt.yticks(fontsize=7)
    fig.subplots_adjust(bottom=.25)




    fig.tight_layout()



    plt.savefig('C:Charts/ChartImages/Survival/' + site + 'surv1.png', bbox_inches='tight')

    plt.show()
    plt.close()

    # FIGURE 2 NO OBVIOUS CAUSE #######################
    allrhythmsList = []
    asystoleList = []
    notShockableList = []
    peaList = []
    vfvtShockableList = []
    can_vfvtShockableList = []

    allrhythmsTable = []
    asystoleTable = []
    notShockableTable = []
    peaTable = []
    vfvtShockableTable = []
    can_vfvtShockableTable = []

    print(years)
    for year in years:
        year = int(year)


        allrhythms_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) &   (df['cr_witbys'] == 1) & (df['date'] == year)].index)
        all_rhythms = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_surv'] == 1) & (df['cr_witbys'] == 1) & (df['date'] == year)].index)

        if allrhythms_total != 0:
            allrhythyms_percent = round((all_rhythms / allrhythms_total) * 100, 2)
        else:
            allrhythyms_percent = 0
        allrhythmsList.append(allrhythyms_percent)

        allrhythms_line = {}
        allrhythms_line['EpisodeYear'] = year
        allrhythms_line['Rhythm'] = 'All rhythms'
        allrhythms_line['Survival Rate'] = allrhythyms_percent
        allrhythms_line['Alive'] = all_rhythms
        allrhythms_line['Treated'] = allrhythms_total
        allrhythmsTable.append(allrhythms_line)




        asystole_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) &  (df['cr_frhyem'] == 0)&   (df['cr_witbys'] == 1) & (df['date'] == year)].index)
        asystole =  len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) &  (df['cr_frhyem'] == 0)& (df['cr_surv'] == 1) & (df['cr_witbys'] == 1) & (df['date'] == year)].index)


        if asystole_total != 0:
            asystole_percent = round((asystole / asystole_total) * 100, 2)
        else:
            asystole_percent = 0
        asystoleList.append(asystole_percent)

        asystole_line = {}
        asystole_line['EpisodeYear'] = year
        asystole_line['Rhythm'] = 'Asystole'
        asystole_line['Survival Rate'] = asystole_percent
        asystole_line['Alive'] = asystole
        asystole_line['Treated'] = asystole_total
        asystoleTable.append(asystole_line)



        notshockable_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) &  (df['cr_frhyem'] == 4)&   (df['cr_witbys'] == 1) & (df['date'] == year)].index)
        notshockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) &  (df['cr_frhyem'] == 4)& (df['cr_surv'] == 1) & (df['cr_witbys'] == 1) & (df['date'] == year)].index)


        if notshockable_total != 0:
            notshockable_percent = round((notshockable / notshockable_total) * 100, 2)
        else:
            notshockable_percent = 0
        notShockableList.append(notshockable_percent)

        notshockable_line = {}
        notshockable_line['EpisodeYear'] = year
        notshockable_line['Rhythm'] = 'Not shockable'
        notshockable_line['Survival Rate'] = notshockable_percent
        notshockable_line['Alive'] = notshockable
        notshockable_line['Treated'] = notshockable_total
        notShockableTable.append(notshockable_line)

        pea_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) &  (df['cr_frhyem'] == 1)&   (df['cr_witbys'] == 1) & (df['date'] == year)].index)
        pea = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) &  (df['cr_frhyem'] == 1)& (df['cr_surv'] == 1) & (df['cr_witbys'] == 1) & (df['date'] == year)].index)

        if pea_total != 0:
            pea_percent = round((pea / pea_total) * 100, 2)
        else:
            pea_percent = 0
        peaList.append(pea_percent)

        pea_line = {}
        pea_line['EpisodeYear'] = year
        pea_line['Rhythm'] = 'PEA'
        pea_line['Survival Rate'] = pea_percent
        pea_line['Alive'] = pea
        pea_line['Treated'] = pea_total
        peaTable.append(pea_line)

        ######TO FIXXXXXXXXX
        value_list = [2, 3, 5]
        vfvtshockable_total = len(df.loc[(df['cr_tx'] == 1) &  (df['cr_witbys'] == 1) & (df['cr_scause'] ==0) & (df['date'] == year) & (
            df.cr_frhyem.isin(value_list))].index)
        vfvtshockable = len(df.loc[(df['cr_tx'] == 1) &  (df['cr_witbys'] == 1) & (df['cr_surv'] == 1) & (df['cr_scause'] ==0) & (df['date'] == year) & (
            df.cr_frhyem.isin(value_list))].index)


        if vfvtshockable_total != 0:
            vfvtshockable_percent = round((vfvtshockable / vfvtshockable_total) * 100, 2)
        else:
            vfvtshockable_percent = 0
        vfvtShockableList.append(vfvtshockable_percent)

        vfvtshockable_line = {}
        vfvtshockable_line['EpisodeYear'] = year
        vfvtshockable_line['Rhythm'] = 'VF/VT/Shockable'
        vfvtshockable_line['Survival Rate'] = vfvtshockable_percent
        vfvtshockable_line['Alive'] = vfvtshockable
        vfvtshockable_line['Treated'] = vfvtshockable_total
        vfvtShockableTable.append(vfvtshockable_line)

        can_vfvtshockable_total = len(canadian_df.loc[(canadian_df['cr_tx'] == 1) & (canadian_df['cr_scause'] == 0) & (
                    canadian_df['date'] == year) &  (df['cr_witbys'] == 1) & (canadian_df.cr_frhyem.isin(value_list))].index)

        can_vfvtshockable = len(canadian_df.loc[(canadian_df['cr_tx'] == 1) & (canadian_df['cr_scause'] ==0) & (
                    canadian_df['date'] == year) & (df['cr_witbys'] == 1) & (canadian_df['cr_surv'] == 1) & (canadian_df.cr_frhyem.isin(value_list))].index)


        if can_vfvtshockable_total != 0:
            can_vfvtshockable_percent = round((can_vfvtshockable / can_vfvtshockable_total) * 100, 2)
        else:
            can_vfvtshockable_percent = 0
        can_vfvtShockableList.append(can_vfvtshockable_percent)

        can_vfvtshockable_line = {}
        can_vfvtshockable_line['EpisodeYear'] = year
        can_vfvtshockable_line['Rhythm'] = 'Canadian VF/VT/Shockable'
        can_vfvtshockable_line['Survival Rate'] = can_vfvtshockable_percent
        can_vfvtshockable_line['Alive'] = can_vfvtshockable
        can_vfvtshockable_line['Treated'] = can_vfvtshockable
        can_vfvtShockableTable.append(can_vfvtshockable_line)




    Fig2Table.extend(allrhythmsTable)
    Fig2Table.extend(asystoleTable)
    Fig2Table.extend(notShockableTable)
    Fig2Table.extend(peaTable)
    Fig2Table.extend(vfvtShockableTable)
    Fig2Table.extend(can_vfvtShockableTable)

    print('fig 2 ')
    print(Fig2Table)





    # fig, ax = plt.subplots()
    rects1 = plt.bar(x - width / 2, allrhythmsList, width, label='All Rhythms')
    rects2 = plt.bar(x + width / 2, asystoleList, width, label='Asystole')
    rects3 = plt.bar(x + 1.5 * width, notShockableList, width, label='notShockable')
    rects4 = plt.bar(x + 2.5 * width, peaList, width, label='Pea')
    rects5 = plt.bar(x + 3.5 * width, vfvtShockableList, width, label='vfvtShockableList')
    rects5 = plt.bar(x + 4.5 * width, can_vfvtShockableList, width, label='can_vfvtShockableList')

    # Add some text for labels, title and custom x-axis tick labels, etc.
    # ax.set_ylabel('Scores')

    plt.title('Figure 2EMS Treated, Adult, No Obvious Cause, and Bystander Witnessed  Survival Rate (%) by Year', fontsize=8)
    plt.xticks(x,fontsize=7, labels = years)
    # plt.xticklabels(years)

    labels = ['All Rhythms', 'Asystole', 'Not Shockable', 'Pea', 'vfvtShockableList', 'can_vfvtShockableList']
    plt.legend(labels, loc="lower left", bbox_to_anchor=(0, -.15), ncol=4, fontsize = 7)  # , bbox_to_anchor=(0.5, -0.3))
    # plt.xticks(fontsize=7)
    plt.yticks(fontsize=7)
    fig.subplots_adjust(bottom=.25)

    fig.tight_layout()

    plt.savefig('C:/Charts/ChartImages/Survival/' + site + 'surv2.png', bbox_inches='tight')
    plt.show()
    plt.close()

    # FIGURE 3 NO OBVIOUS CAUSE AND EMS WITNESSED #######################
    allrhythmsList = []
    asystoleList = []
    notShockableList = []
    peaList = []
    vfvtShockableList = []
    can_vfvtShockableList = []

    allrhythmsTable = []
    asystoleTable = []
    notShockableTable = []
    peaTable = []
    vfvtShockableTable = []
    can_vfvtShockableTable = []

    print(years)
    for year in years:
        year = int(year)

        allrhythms_total = len(
            df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1) & (df['date'] == year)].index)
        all_rhythms = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_surv'] == 1) & (
                    df['cr_witbys'] == 2) & (df['date'] == year)].index)

        if allrhythms_total != 0:
            allrhythyms_percent = round((all_rhythms / allrhythms_total) * 100, 2)
        else:
            allrhythyms_percent = 0
        allrhythmsList.append(allrhythyms_percent)

        allrhythms_line = {}
        allrhythms_line['EpisodeYear'] = year
        allrhythms_line['Rhythm'] = 'All rhythms'
        allrhythms_line['Survival Rate'] = allrhythyms_percent
        allrhythms_line['Alive'] = all_rhythms
        allrhythms_line['Treated'] = allrhythms_total
        allrhythmsTable.append(allrhythms_line)

        asystole_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 0) & (
                    df['cr_witbys'] == 2) & (df['date'] == year)].index)
        asystole = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 0) & (
                    df['cr_surv'] == 1) & (df['cr_witbys'] == 2) & (df['date'] == year)].index)

        if asystole_total != 0:
            asystole_percent = round((asystole / asystole_total) * 100, 2)
        else:
            asystole_percent = 0
        asystoleList.append(asystole_percent)

        asystole_line = {}
        asystole_line['EpisodeYear'] = year
        asystole_line['Rhythm'] = 'Asystole'
        asystole_line['Survival Rate'] = asystole_percent
        asystole_line['Alive'] = asystole
        asystole_line['Treated'] = asystole_total
        asystoleTable.append(asystole_line)



        notshockable_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 4) & (
                    df['cr_witbys'] == 2) & (df['date'] == year)].index)
        notshockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 4) & (
                    df['cr_surv'] == 1) & (df['cr_witbys'] == 2) & (df['date'] == year)].index)

        if notshockable_total != 0:
            notshockable_percent = round((notshockable / notshockable_total) * 100, 2)
        else:
            notshockable_percent = 0
        notShockableList.append(notshockable_percent)

        notshockable_line = {}
        notshockable_line['EpisodeYear'] = year
        notshockable_line['Rhythm'] = 'Not shockable'
        notshockable_line['Survival Rate'] = notshockable_percent
        notshockable_line['Alive'] = notshockable
        notshockable_line['Treated'] = notshockable_total
        notShockableTable.append(notshockable_line)

        pea_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 1) & (
                    df['cr_witbys'] == 2) & (df['date'] == year)].index)
        pea = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 1) & (df['cr_surv'] == 1) & (
                    df['cr_witbys'] == 2) & (df['date'] == year)].index)

        if pea_total != 0:
            pea_percent = round((pea / pea_total) * 100, 2)
        else:
            pea_percent = 0
        peaList.append(pea_percent)

        pea_line = {}
        pea_line['EpisodeYear'] = year
        pea_line['Rhythm'] = 'PEA'
        pea_line['Survival Rate'] = pea_percent
        pea_line['Alive'] = pea
        pea_line['Treated'] = pea_total
        peaTable.append(pea_line)

        ######TO FIXXXXXXXXX
        value_list = [2, 3, 5]
        vfvtshockable_total = len(
            df.loc[(df['cr_tx'] == 1) & (df['cr_witbys'] == 2) & (df['cr_scause'] == 0) & (df['date'] == year) & (
                df.cr_frhyem.isin(value_list))].index)
        vfvtshockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_witbys'] == 2) & (df['cr_surv'] == 1) & (
                    df['cr_scause'] == 0) & (df['date'] == year) & (
                                       df.cr_frhyem.isin(value_list))].index)

        if vfvtshockable_total != 0:
            vfvtshockable_percent = round((vfvtshockable / vfvtshockable_total) * 100, 2)
        else:
            vfvtshockable_percent = 0
        vfvtShockableList.append(vfvtshockable_percent)

        vfvtshockable_line = {}
        vfvtshockable_line['EpisodeYear'] = year
        vfvtshockable_line['Rhythm'] = 'VF/VT/Shockable'
        vfvtshockable_line['Survival Rate'] = vfvtshockable_percent
        vfvtshockable_line['Alive'] = vfvtshockable
        vfvtshockable_line['Treated'] = vfvtshockable_total
        vfvtShockableTable.append(vfvtshockable_line)

        can_vfvtshockable_total = len(canadian_df.loc[(canadian_df['cr_tx'] == 1) & (df['cr_witbys'] == 2) & (canadian_df['cr_scause'] == 0) & (
                canadian_df['date'] == year) & (canadian_df.cr_frhyem.isin(value_list))].index)

        can_vfvtshockable = len(canadian_df.loc[(canadian_df['cr_tx'] == 1) & (df['cr_witbys'] == 2)  & (canadian_df['cr_scause'] == 0) & (
                canadian_df['date'] == year) & (canadian_df['cr_surv'] == 1) & (
                                                    canadian_df.cr_frhyem.isin(value_list))].index)

        if can_vfvtshockable_total != 0:
            can_vfvtshockable_percent = round((can_vfvtshockable / can_vfvtshockable_total) * 100, 2)
        else:
            can_vfvtshockable_percent = 0
        can_vfvtShockableList.append(can_vfvtshockable_percent)

        can_vfvtshockable_line = {}
        can_vfvtshockable_line['EpisodeYear'] = year
        can_vfvtshockable_line['Rhythm'] = 'Canadian VF/VT/Shockable'
        can_vfvtshockable_line['Survival Rate'] = can_vfvtshockable_percent
        can_vfvtshockable_line['Alive'] = can_vfvtshockable
        can_vfvtshockable_line['Treated'] = can_vfvtshockable
        can_vfvtShockableTable.append(can_vfvtshockable_line)




    Fig3Table.extend(allrhythmsTable)
    Fig3Table.extend(asystoleTable)
    Fig3Table.extend(notShockableTable)
    Fig3Table.extend(peaTable)
    Fig3Table.extend(vfvtShockableTable)
    Fig3Table.extend(can_vfvtShockableTable)

    print('fig 3 ')
    print(Fig3Table)



    # fig, ax = plt.subplots()
    rects1 = plt.bar(x - width / 2, allrhythmsList, width, label='All Rhythms')
    rects2 = plt.bar(x + width / 2, asystoleList, width, label='Asystole')
    rects3 = plt.bar(x + 1.5 * width, notShockableList, width, label='notShockable')
    rects4 = plt.bar(x + 2.5 * width, peaList, width, label='Pea')
    rects5 = plt.bar(x + 3.5 * width, vfvtShockableList, width, label='vfvtShockableList')
    rects5 = plt.bar(x + 4.5 * width, can_vfvtShockableList, width, label='can_vfvtShockableList')

    # Add some text for labels, title and custom x-axis tick labels, etc.
    # ax.set_ylabel('Scores')


    ########################################

    plt.title('Figure 3 EMS Treated, Adult, No Obvious Cause, and EMS Witnessed Survival Rate (%) by Year All rhythms', fontsize=8)
    plt.xticks(x, fontsize=7, labels=years)
    # plt.xticklabels(years)

    labels = ['All Rhythms', 'Asystole', 'Not Shockable', 'Pea', 'vfvtShockableList', 'can_vfvtShockableList']
    plt.legend(labels, loc="lower left", bbox_to_anchor=(0, -.15), ncol=4, fontsize=7)  # , bbox_to_anchor=(0.5, -0.3))
    # plt.xticks(fontsize=7)
    plt.yticks(fontsize=7)
    fig.subplots_adjust(bottom=.25)
    #
    #     # ax.bar_label(rects1, padding=3)
    #     # ax.bar_label(rects2, padding=3)
    #
    fig.tight_layout()

    plt.savefig('C:/Desktop/Charts/ChartImages/Survival/' + site + 'surv3.png', bbox_inches='tight')
    plt.show()
    plt.close()

    # FIGURE 4 NO OBVIOUS CAUSE AND EMS WITNESSED #######################
    allrhythmsList = []
    asystoleList = []
    notShockableList = []
    peaList = []
    vfvtShockableList = []
    can_vfvtShockableList = []

    allrhythmsTable = []
    asystoleTable = []
    notShockableTable = []
    peaTable = []
    vfvtShockableTable = []
    can_vfvtShockableTable = []

    print("Need to find the length of these DFS")
    print(len(df.index))
    print(len(canadian_df.index))

    print(years)
    for year in years:
        year = int(year)

        value_list = [2, 3, 5]
        vfvtshockable_total = len(
            df.loc[(df['cr_tx'] == 1) & (df['cr_witbys'] == 1)  & (df['cr_cpratt'] == 1) & (df['cr_scause'] == 0) & (df['date'] == year) & (
                df.cr_frhyem.isin(value_list))].index)
        vfvtshockable = len(df.loc[(df['cr_tx'] == 1) &  (df['cr_cpratt'] == 1)&  (df['cr_witbys'] == 1) & (df['cr_surv'] == 1) & (
                df['cr_scause'] == 0) & (df['date'] == year) & (
                                       df.cr_frhyem.isin(value_list))].index)

        if vfvtshockable_total != 0:
            vfvtshockable_percent = round((vfvtshockable / vfvtshockable_total) * 100, 2)
        else:
            vfvtshockable_percent = 0
        vfvtShockableList.append(vfvtshockable_percent)

        vfvtshockable_line = {}
        vfvtshockable_line['EpisodeYear'] = year
        vfvtshockable_line['Rhythm'] = 'VF/VT/Shockable'
        vfvtshockable_line['Survival Rate'] = vfvtshockable_percent
        vfvtshockable_line['Alive'] = vfvtshockable
        vfvtshockable_line['Treated'] = vfvtshockable_total
        vfvtShockableTable.append(vfvtshockable_line)

        can_vfvtshockable_total = len(
            canadian_df.loc[(canadian_df['cr_tx'] == 1)  & (canadian_df['cr_cpratt'] == 1) & (canadian_df['cr_witbys'] == 1) & (canadian_df['cr_scause'] == 0) & (
                    canadian_df['date'] == year) & (canadian_df.cr_frhyem.isin(value_list))].index)

        can_vfvtshockable = len(
            canadian_df.loc[(canadian_df['cr_tx'] == 1) & (canadian_df['cr_cpratt'] == 1) & (canadian_df['cr_witbys'] == 1) & (canadian_df['cr_scause'] == 0) & (
                    canadian_df['date'] == year) & (canadian_df['cr_surv'] == 1) & (
                                canadian_df.cr_frhyem.isin(value_list))].index)

        if can_vfvtshockable_total != 0:
            can_vfvtshockable_percent = round((can_vfvtshockable / can_vfvtshockable_total) * 100, 2)
        else:
            can_vfvtshockable_percent = 0
        can_vfvtShockableList.append(can_vfvtshockable_percent)

        can_vfvtshockable_line = {}
        can_vfvtshockable_line['EpisodeYear'] = year
        can_vfvtshockable_line['Rhythm'] = 'Canadian VF/VT/Shockable'
        can_vfvtshockable_line['Survival Rate'] = can_vfvtshockable_percent
        can_vfvtshockable_line['Alive'] = can_vfvtshockable
        can_vfvtshockable_line['Treated'] = can_vfvtshockable
        can_vfvtShockableTable.append(can_vfvtshockable_line)


    Fig4Table.extend(vfvtShockableTable)
    Fig4Table.extend(can_vfvtShockableTable)

    print('fig 4 ')
    print(Fig4Table)

    # fig, ax = plt.subplots()
    rects1 = plt.bar(x - width / 2, vfvtShockableList, width, label='vfvtShockable')
    rects2 = plt.bar(x + width / 2, can_vfvtShockableList, width, label='can_vfvtShockableList')


    # Add some text for labels, title and custom x-axis tick labels, etc.
    # ax.set_ylabel('Scores')

    plt.title('Figure 4 EMS Treated, Adult, No Obvious Cause, Bystander Witnessed,\n Bystander CPR, VF/VT/Shockable Survival Rate (%) by Year',
              fontsize=8)
    plt.xticks(x, fontsize=7, labels=years)
    # plt.xticklabels(years)

    labels = ['VF/VT/Shockable', 'Canadian VF/VT/Shockable']
    plt.legend(labels, loc="lower left", bbox_to_anchor=(0, -.15), ncol=4, fontsize=7)  # , bbox_to_anchor=(0.5, -0.3))
    # plt.xticks(fontsize=7)
    plt.yticks(fontsize=7)
    fig.subplots_adjust(bottom=.25)
    #
    #     # ax.bar_label(rects1, padding=3)
    #     # ax.bar_label(rects2, padding=3)
    #
    fig.tight_layout()

    plt.savefig('C:Charts/ChartImages/Survival/' + site + 'surv4.png', bbox_inches='tight')
    plt.show()
    plt.close()

    # FIGURE 5 EMS treated, all patients, non# traumatic Received Bystander CPR (%) by# Year# Toronto# Canada

    allrhythmsList = []
    asystoleList = []
    notShockableList = []
    peaList = []
    vfvtShockableList = []
    can_vfvtShockableList = []


    print(years)
    for year in years:
        year = int(year)



        value_list = [2, 3, 5]

        vfvtshockable_total = len(
            df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 17) & (df['date'] == year)].index)

        vfvtshockable = len(
            df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 17) & (df['cr_cpratt'] == 1)  & (df['date'] == year)].index)

        if vfvtshockable_total != 0:
            vfvtshockable_percent = round((vfvtshockable / vfvtshockable_total) * 100, 2)
        else:
            vfvtshockable_percent = 0
        vfvtShockableList.append(vfvtshockable_percent)

        can_vfvtshockable_total =len(
            canadian_df.loc[( canadian_df['cr_tx'] == 1) & ( canadian_df['cr_scause'] != 17) & ( canadian_df['date'] == year)].index)

        can_vfvtshockable = len(
            canadian_df.loc[(canadian_df['cr_tx'] == 1) & (canadian_df['cr_scause'] != 17) & (canadian_df['cr_cpratt'] == 1)  & (canadian_df['date'] == year)].index)

        if can_vfvtshockable_total != 0:
            can_vfvtshockable_percent = round((can_vfvtshockable / can_vfvtshockable_total) * 100, 2)
        else:
            can_vfvtshockable_percent = 0
        can_vfvtShockableList.append(can_vfvtshockable_percent)
    print("I just dont understand")
    print(allrhythmsList)
    print(asystoleList)
    print(notShockableList)
    print(peaList)
    print(vfvtShockableList)
    print(can_vfvtShockableList)

    # fig, ax = plt.subplots()
    plt.figure(figsize=(4, 4))
    plt.plot(years, vfvtShockableList, 'bo', label=site  )
    plt.plot(years, can_vfvtShockableList,'r--', label='Canada')


    plt.title('Figure 5 EMS treated, all patients, \n non-traumatic Received Bystander CPR (%) by Year',fontsize=8)
    plt.minorticks_off()
    # plt.xticks(x, fontsize=7, labels=years)
    # # plt.xticklabels(years)
    #
    labels = [site, 'Canada']
    plt.legend(labels, loc="lower left", bbox_to_anchor=(0, -.15), ncol=4, fontsize=7)  # , bbox_to_anchor=(0.5, -0.3))

    plt.xticks(fontsize=7)
    plt.yticks(fontsize=7)
    fig.subplots_adjust(bottom=.9)

    fig.tight_layout()

    plt.savefig('C:/Users/mcbarnettr/Desktop/Charts/ChartImages/Survival/' + site + 'surv5.png', bbox_inches='tight')
    plt.show()
    plt.close()

    # FIGURE 6 Figure 6 # EMS treated, all patients, non # traumatic Survival Rate (%) by Year

    allrhythmsList = []
    asystoleList = []
    notShockableList = []
    peaList = []
    vfvtShockableList = []
    can_vfvtShockableList = []

    print(years)
    for year in years:
        year = int(year)

        ######TO FIXXXXXXXXX

        value_list = [2, 3, 5]

        vfvtshockable_total = len(
            df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 17) & (df['date'] == year)].index)

        vfvtshockable = len(
            df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 17) & (df['cr_surv'] == 1) & (df['date'] == year)].index)

        if vfvtshockable_total != 0:
            vfvtshockable_percent = round((vfvtshockable / vfvtshockable_total) * 100, 2)
        else:
            vfvtshockable_percent = 0
        vfvtShockableList.append(vfvtshockable_percent)

        can_vfvtshockable_total = len(
            canadian_df.loc[
                (canadian_df['cr_tx'] == 1) & (canadian_df['cr_scause'] != 17) & (canadian_df['date'] == year)].index)

        can_vfvtshockable = len(
            canadian_df.loc[
                (canadian_df['cr_tx'] == 1) & (canadian_df['cr_scause'] != 17) & (canadian_df['cr_surv'] == 1) & (
                            canadian_df['date'] == year)].index)

        if can_vfvtshockable_total != 0:
            can_vfvtshockable_percent = round((can_vfvtshockable / can_vfvtshockable_total) * 100, 2)
        else:
            can_vfvtshockable_percent = 0
        can_vfvtShockableList.append(can_vfvtshockable_percent)
    print("I just dont understand")
    print(allrhythmsList)
    print(asystoleList)
    print(notShockableList)
    print(peaList)
    print(vfvtShockableList)
    print(can_vfvtShockableList)

    # fig, ax = plt.subplots()
    plt.figure(figsize=(4, 4))
    plt.plot(years, vfvtShockableList, 'bo', label=site)
    plt.plot(years, can_vfvtShockableList, 'r--', label='Canada')

    plt.title('Figure 6 EMS treated, all patients, \n non-traumatic Survival Rate (%) by Year', fontsize=8)
    plt.minorticks_off()
    # plt.xticks(x, fontsize=7, labels=years)
    # # plt.xticklabels(years)
    #
    plt.xticks(fontsize=7)
    plt.yticks(fontsize=7)
    labels = [site, 'Canada']
    plt.legend(labels, loc="lower left", bbox_to_anchor=(0, -.15), ncol=4, fontsize=7)  # , bbox_to_anchor=(0.5, -0.3))

    # # plt.xticks(fontsize=7)
    # plt.yticks(fontsize=7)
    # fig.subplots_adjust(bottom=.25)
    #
    #     # ax.bar_label(rects1, padding=3)
    #     # ax.bar_label(rects2, padding=3)
    #
    fig.tight_layout()

    plt.savefig('C:/Charts/ChartImages/Survival/' + site + 'surv6.png', bbox_inches='tight')
    plt.show()
    plt.close()

    # FIGURE 7 EMS treated, all patients, non # traumatic VF/VT as First Rhythm (%) by # Year

    allrhythmsList = []
    asystoleList = []
    notShockableList = []
    peaList = []
    vfvtShockableList = []
    can_vfvtShockableList = []

    print(years)
    for year in years:
        year = int(year)

        ######TO FIXXXXXXXXX

        value_list = [2, 3]

        vfvtshockable_total = len(
            df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 17) & (df['date'] == year)].index)

        vfvtshockable = len(
            df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 17) & (df['cr_surv'] == 1) & (df['date'] == year)].index)

        if vfvtshockable_total != 0:
            vfvtshockable_percent = round((vfvtshockable / vfvtshockable_total) * 100, 2)
        else:
            vfvtshockable_percent = 0
        vfvtShockableList.append(vfvtshockable_percent)

        can_vfvtshockable_total = len(
            canadian_df.loc[
                (canadian_df['cr_tx'] == 1) & (canadian_df['cr_scause'] != 17) & (canadian_df['date'] == year)].index)

        can_vfvtshockable = len(
            canadian_df.loc[
                (canadian_df['cr_tx'] == 1) & (canadian_df['cr_scause'] != 17) & (canadian_df['cr_surv'] == 1) & (
                        canadian_df['date'] == year) &  (canadian_df.cr_frhyem.isin(value_list))    ].index)

        if can_vfvtshockable_total != 0:
            can_vfvtshockable_percent = round((can_vfvtshockable / can_vfvtshockable_total) * 100, 2)
        else:
            can_vfvtshockable_percent = 0
        can_vfvtShockableList.append(can_vfvtshockable_percent)
    print("I just dont understand")
    print(allrhythmsList)
    print(asystoleList)
    print(notShockableList)
    print(peaList)
    print(vfvtShockableList)
    print(can_vfvtShockableList)

    # fig, ax = plt.subplots()
    plt.figure(figsize=(4, 4))
    plt.plot(years, vfvtShockableList, 'bo', label=site)
    plt.plot(years, can_vfvtShockableList, 'r--', label='Canada')

    plt.title('Figure 7 - EMS treated, all patients,\n nontraumatic VF/VT as First Rhythm (%) byYear', fontsize=8)
    plt.minorticks_off()
    # plt.xticks(x, fontsize=7, labels=years)
    # # plt.xticklabels(years)
    #
    labels = [site, 'Canada']
    plt.legend(labels, loc="lower left", bbox_to_anchor=(0, -.15), ncol=4, fontsize=7)  # , bbox_to_anchor=(0.5, -0.3))

    plt.xticks(fontsize=7)
    plt.yticks(fontsize=7)

    # fig.subplots_adjust(bottom=.25)
    #
    #     # ax.bar_label(rects1, padding=3)
    #     # ax.bar_label(rects2, padding=3)
    #
    fig.tight_layout()
    plt.savefig('C:/Charts/ChartImages/Survival/' + site + 'surv7.png', bbox_inches='tight')
    plt.show()
    plt.close()

    # FIGURE 8 EMS treated, all patients, non # traumatic Average Response time  # (minutes) from Call Received to 1st Arrival  # on Scene by Year


    avg_minutes_list = []
    can_avg_minutes_list = []

    print(years)
    for year in years:
        year = int(year)

        ######TO FIXXXXXXXXX


        avg_response_total =  df[(df['cr_tx'] == 1) & (df['cr_scause'] != 17) & (df['date'] == year)]
        print("avg_response_total")
        print(len(avg_response_total))
        print(avg_response_total)
        avg_response_total2 =   avg_response_total[(  avg_response_total['cr_rig1tm'] != '') & (  avg_response_total['cr_ptmrcv'] != '') &
                                                  (avg_response_total['cr_rig1tm'].notna()) & (  avg_response_total['cr_ptmrcv'].notna()) ]
        print("avg_response_total after filter")
        print(len(avg_response_total2))
        print(avg_response_total2['cr_rig1tm'])
        print(avg_response_total2['cr_ptmrcv'])
        # avg_response = relativedelta.relativedelta(
        #             datetime.strptime(avg_response_total['cr_rig1tm'], "%Y-%m-%d %H:%M:%S"),
        #             datetime.strptime(avg_response_total['cr_ptmrcv'], "%Y-%m-%d %H:%M:%S"))
        avg_response_total2.reset_index(drop = True, inplace= True)
        print(avg_response_total2['cr_rig1tm'])
        totalminutes = 0

        for i in range(len(avg_response_total2)):
            # print(i)
            # print(avg_response_total2.loc[i,'cr_rig1tm'])
            # print(avg_response_total2.loc[i, 'cr_ptmrcv'])
            avg_response = relativedelta.relativedelta(
                datetime.strptime(avg_response_total2.loc[i,'cr_rig1tm'], "%Y-%m-%d %H:%M:%S"),
                datetime.strptime(avg_response_total2.loc[i,'cr_ptmrcv'], "%Y-%m-%d %H:%M:%S"))

            diffminutes = float(avg_response.minutes) + float(avg_response.seconds/60)
            # print(diffminutes)
            totalminutes = totalminutes + diffminutes



        # print("Responses average")
        # print(avg_response_total2)
        # print(totalminutes)
        # print(len(avg_response_total2))
        avg_minutes = totalminutes /len(avg_response_total2)
        # print('average minutes')
        # print(avg_minutes)
        avg_minutes_list.append(avg_minutes)




        can_avg_response_total = canadian_df[(canadian_df['cr_tx'] == 1) & (canadian_df['cr_scause'] != 17) & (canadian_df['date'] == year)]


        can_avg_response_total2 = can_avg_response_total[
            (can_avg_response_total['cr_rig1tm'] != '') & (can_avg_response_total['cr_ptmrcv'] != '') &
            (can_avg_response_total['cr_rig1tm'].notna()) & (can_avg_response_total['cr_ptmrcv'].notna())]

        can_avg_response_total2.reset_index(drop=True, inplace=True)
        print(can_avg_response_total2['cr_rig1tm'])
        totalminutes = 0

        for i in range(len(can_avg_response_total2)):

            can_avg_response = relativedelta.relativedelta(
                datetime.strptime(can_avg_response_total2.loc[i, 'cr_rig1tm'], "%Y-%m-%d %H:%M:%S"),
                datetime.strptime(can_avg_response_total2.loc[i, 'cr_ptmrcv'], "%Y-%m-%d %H:%M:%S"))
            print("average response")
            print(can_avg_response)
            diffminutes = float(can_avg_response.minutes) + float(can_avg_response.seconds / 60)
            print(diffminutes)
            totalminutes = totalminutes + diffminutes


        avg_minutes = totalminutes / len(can_avg_response_total2)

        can_avg_minutes_list.append(avg_minutes)





    # fig, ax = plt.subplots()
    plt.figure(figsize=(4, 4))
    plt.plot(years, avg_minutes_list, 'bo', label=site)
    plt.plot(years, can_avg_minutes_list, 'r--', label='Canada')

    plt.title('Figure 8 - EMS treated, all patients, \nnon-traumatic Average Response time (minutes)\n from Call Received to 1st Arrival on Scene by Year', fontsize=8)
    plt.minorticks_off()
    # plt.xticks(x, fontsize=7, labels=years)
    # # plt.xticklabels(years)
    #
    labels = [site, 'Canada']
    plt.legend(labels, loc="lower left", bbox_to_anchor=(0, -.15), ncol=4, fontsize=7)  # , bbox_to_anchor=(0.5, -0.3))
    plt.xticks(fontsize=7)
    plt.yticks(fontsize=7)


    fig.tight_layout()

    plt.savefig('C:/Charts/ChartImages/Survival/' + site + 'surv8.png', bbox_inches='tight')
    plt.show()
    plt.close()


# CHART CREATION

    fname = "C:/Users/mcbarnettr/Desktop/Charts/SurvivalGraphs/" + site + " CASurvivalReport.xlsx"
    workbook = Workbook()
    ws = workbook.active
    ws.column_dimensions['A'].width = 10.89
    ws.column_dimensions['B'].width = 14.11
    ws.column_dimensions['C'].width = 13.5


    thin = Side(border_style="thin", color="000000")
    thinBorder = Border(top=thin, left=thin, right=thin, bottom=thin)
    double = Side(border_style="double", color="ff0000")
    # fill = PatternFill(patternType = 'solid',end_color = '34a8eb')
    fill = PatternFill(patternType='solid', start_color='5c61a1', end_color="5c61a1")
    ws['A3'] = 'Cardiac Arrest – Survival Rate Report'
    ws['A3'].font = Font(size=14, bold=True)
    ws['A5'] = 'Region:' + site
    x = datetime.now()
    ws['H5'] = x.strftime("%B %d, %Y %I:%M %p")

    img1 = Image('C:/Users/mcbarnettr/Desktop/Charts/ChartImages/Survival/' + site + 'surv1.png')
    img2 = Image('C:/Users/mcbarnettr/Desktop/Charts/ChartImages/Survival/' + site + 'surv2.png')
    img3 = Image('C:/Users/mcbarnettr/Desktop/Charts/ChartImages/Survival/' + site + 'surv3.png')
    img4 = Image('C:/Users/mcbarnettr/Desktop/Charts/ChartImages/Survival/' + site + 'surv4.png')
    img5 = Image('C:/Users/mcbarnettr/Desktop/Charts/ChartImages/Survival/' + site + 'surv5.png')
    img6 = Image('C:/Users/mcbarnettr/Desktop/Charts/ChartImages/Survival/' + site + 'surv6.png')
    img7 = Image('C:/Users/mcbarnettr/Desktop/Charts/ChartImages/Survival/' + site + 'surv7.png')
    img8 = Image('C:/Users/mcbarnettr/Desktop/Charts/ChartImages/Survival/' + site + 'surv8.png')




    ws.add_image(img1, 'A7')

    fig1length = len(Fig1Table)

    excel_rownum = 33

    ws.cell(row=excel_rownum-1, column=1).value = "Table 1 - EMS Treated, All patients, Non traumatic"

    ws.cell(row=excel_rownum, column=1).value = "Episode Year"
    ws.cell(row=excel_rownum, column=2).value = "Rhythm"
    ws.cell(row=excel_rownum, column=3).value = "Survival Rate( %)"
    ws.cell(row=excel_rownum, column=4).value = " 95 % CI"
    ws.cell(row=excel_rownum, column=5).value = "Alive #"
    ws.cell(row=excel_rownum, column=6).value = "Treated #"

    excel_rownum+=1

    for i in range(0,fig1length):
        print(i)
        ws.cell(row = excel_rownum, column = 1).value = Fig1Table[i]['EpisodeYear']
        ws.cell(row=excel_rownum, column=2).value = Fig1Table[i]['Rhythm']
        ws.cell(row=excel_rownum, column=3).value = Fig1Table[i]['Survival Rate']
        # ws.cell(row = excel_rownum, column = 4).value = Fig1Table[i]['95%CI']
        ws.cell(row=excel_rownum, column=5).value = Fig1Table[i]['Alive']
        ws.cell(row = excel_rownum, column = 6).value = Fig1Table[i]['Treated']
        excel_rownum = excel_rownum + 1

    excel_rownum = excel_rownum + 1


    ws.add_image(img2, 'A' +str(excel_rownum))

    fig2length = len(Fig2Table)

    excel_rownum = excel_rownum + 24

    ws.cell(row=excel_rownum - 1, column=1).value = "Table 2 - EMS Treated, Adult, No Obvious Cause, and Bystander Witnessed"

    ws.cell(row=excel_rownum, column=1).value = "Episode Year"
    ws.cell(row=excel_rownum, column=2).value = "Rhythm"
    ws.cell(row=excel_rownum, column=3).value = "Survival Rate( %)"
    ws.cell(row=excel_rownum, column=4).value = " 95 % CI"
    ws.cell(row=excel_rownum, column=5).value = "Alive #"
    ws.cell(row=excel_rownum, column=6).value = "Treated #"

    excel_rownum += 1

    for i in range(0, fig2length):
        print(i)
        ws.cell(row=excel_rownum, column=1).value = Fig2Table[i]['EpisodeYear']
        ws.cell(row=excel_rownum, column=2).value = Fig2Table[i]['Rhythm']
        ws.cell(row=excel_rownum, column=3).value = Fig2Table[i]['Survival Rate']
        # ws.cell(row=excel_rownum, column=4).value = Fig2Table[i]['95%CI']
        ws.cell(row=excel_rownum, column=5).value = Fig2Table[i]['Alive']
        ws.cell(row=excel_rownum, column=6).value = Fig2Table[i]['Treated']
        excel_rownum = excel_rownum + 1

    excel_rownum = excel_rownum + 1

    ws.add_image(img3, 'A' +str(excel_rownum))


    fig3length = len(Fig3Table)

    excel_rownum = excel_rownum + 24
    ws.cell(row=excel_rownum - 1,column=1).value = "Table 3 - EMS Treated, Adult, No Obvious Cause, and EMS Witnessed"

    ws.cell(row=excel_rownum, column=1).value = "Episode Year"
    ws.cell(row=excel_rownum, column=2).value = "Rhythm"
    ws.cell(row=excel_rownum, column=3).value = "Survival Rate( %)"
    ws.cell(row=excel_rownum, column=4).value = " 95 % CI"
    ws.cell(row=excel_rownum, column=5).value = "Alive #"
    ws.cell(row=excel_rownum, column=6).value = "Treated #"

    excel_rownum += 1

    for i in range(0, fig3length):
        print(i)
        ws.cell(row=excel_rownum, column=1).value = Fig3Table[i]['EpisodeYear']
        ws.cell(row=excel_rownum, column=2).value = Fig3Table[i]['Rhythm']
        ws.cell(row=excel_rownum, column=3).value = Fig3Table[i]['Survival Rate']
        # ws.cell(row=excel_rownum, column=4).value = Fig3Table[i]['95%CI']
        ws.cell(row=excel_rownum, column=5).value = Fig3Table[i]['Alive']
        ws.cell(row=excel_rownum, column=6).value = Fig3Table[i]['Treated']
        excel_rownum = excel_rownum + 1

    excel_rownum = excel_rownum + 1

    ws.add_image(img4, 'A' + str(excel_rownum))

    fig4length = len(Fig4Table)

    excel_rownum = excel_rownum + 24
    ws.cell(row=excel_rownum - 1, column=1).value = "Table 4 - EMS Treated, Adult, No Obvious Cause, Bystander Witnessed, Bystander CPR"

    ws.cell(row=excel_rownum, column=1).value = "Episode Year"
    ws.cell(row=excel_rownum, column=2).value = "Rhythm"
    ws.cell(row=excel_rownum, column=3).value = "Survival Rate( %)"
    ws.cell(row=excel_rownum, column=4).value = " 95 % CI"
    ws.cell(row=excel_rownum, column=5).value = "Alive #"
    ws.cell(row=excel_rownum, column=6).value = "Treated #"

    excel_rownum += 1

    for i in range(0, fig4length):
        print(i)
        ws.cell(row=excel_rownum, column=1).value = Fig4Table[i]['EpisodeYear']
        ws.cell(row=excel_rownum, column=2).value = Fig4Table[i]['Rhythm']
        ws.cell(row=excel_rownum, column=3).value = Fig4Table[i]['Survival Rate']
        # ws.cell(row=excel_rownum, column=4).value = Fig4Table[i]['95%CI']
        ws.cell(row=excel_rownum, column=5).value = Fig4Table[i]['Alive']
        ws.cell(row=excel_rownum, column=6).value = Fig4Table[i]['Treated']

    excel_rownum = excel_rownum + 1

    ws.add_image(img5, 'A' + str(excel_rownum))
    ws.add_image(img6, 'F' + str(excel_rownum))
    ws.add_image(img7, 'A' + str(excel_rownum + 20))
    ws.add_image(img8, 'F' + str(excel_rownum + 20))



    workbook.save(filename=fname)



fig, ax = plt.subplots()

CreateSiteLevelCharts('BC',ax)
plt.close()
fig, ax = plt.subplots()
CreateSiteLevelCharts('SK',ax)





