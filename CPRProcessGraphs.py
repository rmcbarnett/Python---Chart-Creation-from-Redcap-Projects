import pandas as pd
import matplotlib.pyplot as plt
import pandas as pd

from redcap import Project, RedcapError
from datetime import datetime
from dateutil import relativedelta

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.drawing.image import Image
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles import  Alignment,Font
from openpyxl.worksheet.pagebreak import Break
import statistics

URL = 'https://redcap.smh.ca/redcap/api/'

API_KEY = '' #CANROCPROJECT
project = Project(URL, API_KEY)

def CreateSiteLevelCharts(site,ax):

    # len(df.index) == 0
    start_date = '2018-SEP-01'
    end_date = '2019-SEP-30'
    # start_date =  input ("Enter start date as yyyy-mmm/dd:")
    # end_date = input ("Enter end date : as yyyy-mmm/dd:")
    allfields = ['cr_cmprt1','cr_cmprt2','cr_cmprt3','cr_cmprt4','cr_cmprt5','cr_cmprt6','cr_cmprt7','cr_cmprt8','cr_cmprt9','cr_cmprt10'
                 ,'cr_cprff1','cr_cprff2','cr_cprff3','cr_cprff4','cr_cprff5','cr_cprff6','cr_cprff7','cr_cprff8','cr_cprff9','cr_cprff10',
                 'cr_cdpth1','cr_cdpth2','cr_cdpth3','cr_cdpth4','cr_cdpth5','cr_cdpth6','cr_cdpth7','cr_cdpth8','cr_cdpth9','cr_cdpth10']
    df = project.export_records(format='df', fields=allfields)
    indexdf2 = df.reset_index()
    indexdf2 = indexdf2[indexdf2["cr_record_id"].str.contains(site)]
    indexdf2.reset_index(drop=True, inplace=True)
    print(indexdf2.head(200))
    allThree = 0
    ratecount=0
    fractioncount = 0
    depthcount = 0
    for i in range(len(indexdf2)):
        print('row' + str(i))
        rateavg = indexdf2.loc[i, ['cr_cmprt1', 'cr_cmprt2', 'cr_cmprt3', 'cr_cmprt4', 'cr_cmprt5', 'cr_cmprt6', 'cr_cmprt7', 'cr_cmprt8','cr_cmprt9', 'cr_cmprt10']].mean()
        fractionavg= indexdf2.loc[i, ['cr_cprff1','cr_cprff2','cr_cprff3','cr_cprff4','cr_cprff5','cr_cprff6','cr_cprff7','cr_cprff8','cr_cprff9','cr_cprff10']].mean()
        depthavg = indexdf2.loc[i, ['cr_cdpth1','cr_cdpth2','cr_cdpth3','cr_cdpth4','cr_cdpth5','cr_cdpth6','cr_cdpth7','cr_cdpth8','cr_cdpth9','cr_cdpth10']].mean()
        print(rateavg)
        print(fractionavg)
        print(depthavg)
        if 100<=rateavg<=200 :
            ratecount +=1
        if fractionavg>=.8 :
            fractioncount +=1
        if  5.0<=depthavg<=6.0:
            depthcount +=1
        if 100<=rateavg<=200 and fractionavg>=.8 and 5.0<=depthavg<=6.0:
            allThree +=1



    fields1 = ['cr_cmprt1','cr_cmprt2','cr_cmprt3','cr_cmprt4','cr_cmprt5','cr_cmprt6','cr_cmprt7','cr_cmprt8','cr_cmprt9','cr_cmprt10']

    df = project.export_records(format='df',fields=fields1 )
    pd.set_option("display.max.columns", None)
    pd.set_option('display.max_rows',None)

    indexdf2 = df.reset_index()  #shift the index from cr_recordid so we can filter by RecordId



    indexdf2=  indexdf2[indexdf2["cr_record_id"].str.contains(site)] #to get a count of only the rows with site eg BC
    indexdf2.reset_index(drop=True, inplace=True)
    index =  indexdf2.index
    number_of_rows = len(index) #Total number of analyzable records

    del indexdf2["cr_record_id"]
    total_analyzable_min = indexdf2.count().sum() #count of total analyzable minutes



    indexdf2 = indexdf2.apply(pd.to_numeric) #getting the dataset ready to check numeric conditions
    rowvaluecount = 0
    cprrowcount = 0
    for i in range(len(indexdf2)):
        goodrow = False
        if indexdf2.loc[i,'cr_cmprt1'] is not None and indexdf2.loc[i,'cr_cmprt1'] in range(100, 121):
            rowvaluecount+=1
            goodrow = True
        if indexdf2.loc[i,'cr_cmprt2'] is not None and indexdf2.loc[i,'cr_cmprt2'] in range(100, 121):
            rowvaluecount+=1
            goodrow = True
        if indexdf2.loc[i,'cr_cmprt3'] is not None and indexdf2.loc[i,'cr_cmprt3'] in range(100, 121):
            rowvaluecount+=1
            goodrow = True
        if indexdf2.loc[i,'cr_cmprt4'] is not None and indexdf2.loc[i,'cr_cmprt4'] in range(100, 121):
            rowvaluecount+=1
            goodrow = True
        if indexdf2.loc[i,'cr_cmprt5'] is not None and indexdf2.loc[i,'cr_cmprt5'] in range(100, 121):
            rowvaluecount+=1
            goodrow = True
        if indexdf2.loc[i,'cr_cmprt6'] is not None and indexdf2.loc[i,'cr_cmprt6'] in range(100, 121):
            rowvaluecount+=1
            goodrow = True
        if indexdf2.loc[i,'cr_cmprt7'] is not None and indexdf2.loc[i,'cr_cmprt7'] in range(100, 121):
            rowvaluecount+=1
            goodrow = True
        if indexdf2.loc[i,'cr_cmprt8'] is not None and indexdf2.loc[i,'cr_cmprt8'] in range(100, 121):
            rowvaluecount+=1
            goodrow = True
        if indexdf2.loc[i, 'cr_cmprt9'] is not None and indexdf2.loc[i, 'cr_cmprt9'] in range(100, 121):
            rowvaluecount += 1
            goodrow = True
        if indexdf2.loc[i, 'cr_cmprt10'] is not None and indexdf2.loc[i, 'cr_cmprt10'] in range(100, 121):
            rowvaluecount += 1
            goodrow = True
        if goodrow == True:
            cprrowcount +=1

    df1= df.stack().reset_index()

    nationalmean = df1.loc[:,0].mean()

    new_df = df1[df1["cr_record_id"].str.contains(site)]
    print(new_df.head(100))

    median = new_df.loc[:,0].median()

    mean = new_df.loc[:,0].mean()


    percent_10 = new_df.loc[:,0].quantile(0.1) # 10th percentile

    percent_90 = new_df.loc[:,0].quantile(0.90) # 10th percentile


    rateDivisions = ['Median', '10th Percentile','90th Percentile']
    rateScores = [median,percent_10,percent_90]

    #Creating the chart based on the data thats been filtered and cleaned

    plt.grid(True)
    plt.minorticks_on()

    plt.figure(figsize=(3,4))
    plt.bar(rateDivisions,rateScores,color='darkblue', width = .3)

    plt.suptitle("Compression rate (comps/min)\n ", fontsize=7)
    plt.title("Target 100-120 \n Minutes meeting target: "  + str(rowvaluecount) + "   ({:.2%})".format(rowvaluecount/total_analyzable_min),fontsize=6)

    plt.axhline(nationalmean,color ='red')
    plt.xticks(fontsize=6)
    plt.yticks(fontsize=6)
    plt.savefig('C:/Charts/ChartImages/'+site+'cpr1.png', bbox_inches='tight')

    plt.close()


    fields2 = ['cr_cprff1','cr_cprff2','cr_cprff3','cr_cprff4','cr_cprff5','cr_cprff6','cr_cprff7','cr_cprff8','cr_cprff9','cr_cprff10']
    fraction_df = project.export_records(format='df',fields=fields2 )
    pd.set_option("display.max.columns", None)
    pd.set_option('display.max_rows',None)
    fraction_indexdf2= fraction_df.reset_index()  # shift the index from cr_recordid so we can filter by RecordId
    fraction_indexdf2 = fraction_indexdf2[fraction_indexdf2["cr_record_id"].str.contains(site)]  # to get a count of only the rows with site
    fraction_indexdf2.reset_index(drop=True, inplace=True)

    index =  fraction_indexdf2.index

    del fraction_indexdf2["cr_record_id"]

    fraction_rowvaluecount = 0


    fraction_indexdf2 = fraction_indexdf2.apply(pd.to_numeric)  # getting the dataset ready to check numeric conditions
    fraction_total_analyzable_min =  fraction_indexdf2.count().sum()  # count of total analyzable minutes
    fractionrowcount = 0
    for i in range(len(fraction_indexdf2)):
        goodrow = False
        if fraction_indexdf2.loc[i, 'cr_cprff1'] is not None and fraction_indexdf2.loc[i, 'cr_cprff1'] >= .80:
                fraction_rowvaluecount += 1
                goodrow = True
        if fraction_indexdf2.loc[i, 'cr_cprff2'] is not None and fraction_indexdf2.loc[i, 'cr_cprff2'] >= .80:
                fraction_rowvaluecount += 1
                goodrow = True
        if fraction_indexdf2.loc[i, 'cr_cprff3'] is not None and fraction_indexdf2.loc[i, 'cr_cprff3'] >= .80:
                fraction_rowvaluecount += 1
                goodrow = True
        if fraction_indexdf2.loc[i, 'cr_cprff4'] is not None and fraction_indexdf2.loc[i, 'cr_cprff4'] >= .80:
                fraction_rowvaluecount += 1
                goodrow = True
        if fraction_indexdf2.loc[i, 'cr_cprff5'] is not None and fraction_indexdf2.loc[i, 'cr_cprff5'] >= .80:
                fraction_rowvaluecount += 1
                goodrow = True
        if fraction_indexdf2.loc[i, 'cr_cprff6'] is not None and fraction_indexdf2.loc[i, 'cr_cprff6'] >= .80:
                fraction_rowvaluecount += 1
                goodrow = True
        if fraction_indexdf2.loc[i, 'cr_cprff7'] is not None and fraction_indexdf2.loc[i, 'cr_cprff7'] >= .80:
                fraction_rowvaluecount += 1
                goodrow = True
        if fraction_indexdf2.loc[i, 'cr_cprff8'] is not None and fraction_indexdf2.loc[i, 'cr_cprff8'] >= .80:
                fraction_rowvaluecount += 1
                goodrow = True
        if fraction_indexdf2.loc[i, 'cr_cprff9'] is not None and fraction_indexdf2.loc[i, 'cr_cprff9'] >= .80:
                fraction_rowvaluecount += 1
                goodrow = True
        if fraction_indexdf2.loc[i, 'cr_cprff10'] is not None and fraction_indexdf2.loc[i, 'cr_cprff10'] >= .80:
                fraction_rowvaluecount += 1
                goodrow = True
        if goodrow == True:
            fractionrowcount += 1


    fraction_df1 = fraction_df.stack().reset_index()
    fraction_nationalmean = fraction_df1.loc[:, 0].mean()

    new_df = fraction_df1[fraction_df1["cr_record_id"].str.contains(site)]


    median = new_df.loc[:, 0].median()

    mean = new_df.loc[:, 0].mean()


    percent_10 = new_df.loc[:, 0].quantile(0.1)  # 10th percentile
    # print(percent_10)
    percent_90 = new_df.loc[:, 0].quantile(0.90)  # 10th percentile
    # print(percent_75)

    rateDivisions = ['Median', '10th Percentile', '90th Percentile']
    rateScores = [median, percent_10, percent_90]

    #Create the chart based on the data filtered and cleaned

    plt.grid(True)
    plt.minorticks_on()

    plt.figure(figsize=(3,4))
    plt.bar(rateDivisions,rateScores,color='darkblue', width = .3)

    plt.suptitle("Compression fraction ", fontsize=7)

    print(fraction_rowvaluecount)
    print(fraction_total_analyzable_min)
    if fraction_total_analyzable_min > 0:
        plt.title("Target >=.80 \n Minutes meeting target: " + str(fraction_rowvaluecount) + "   ({:.2%})".format(
        fraction_rowvaluecount / fraction_total_analyzable_min), fontsize=6)
    else:
        plt.title("Target >=.80 \n Minutes meeting target: NO DATA ", fontsize=6)



    plt.axhline(fraction_nationalmean,color ='red')
    plt.xticks(fontsize=7)
    plt.yticks(fontsize=7)
    plt.savefig('C:/Charts/ChartImages/'+site+'cpr2.png', bbox_inches='tight')
    plt.close()




    fields2 = ['cr_cdpth1','cr_cdpth2','cr_cdpth3','cr_cdpth4','cr_cdpth5','cr_cdpth6','cr_cdpth7','cr_cdpth8','cr_cdpth9','cr_cdpth10']
    depth_df =  depth_df = project.export_records(format='df', fields=fields2)
    depth_indexdf2 = depth_df.reset_index()  # shift the index from cr_recordid so we can filter by RecordId
    depth_indexdf2 =  depth_indexdf2[depth_indexdf2["cr_record_id"].str.contains(site)]  # to get a count of only the rows with site
    depth_indexdf2.reset_index(drop=True, inplace=True)
    del depth_indexdf2["cr_record_id"]
    depth_indexdf2 = depth_indexdf2.apply(pd.to_numeric)  # getting the dataset ready to check numeric conditions
    depth_rowvaluecount = 0
    depthrowcount = 0
    depth_total_analyzable_min = depth_indexdf2.count().sum()




    for i in range(len(fraction_indexdf2)):
        goodrow = False
        if depth_indexdf2.loc[i, 'cr_cdpth1'] is not None and depth_indexdf2.loc[i, 'cr_cdpth1'] >= .80:
                depth_rowvaluecount += 1
                goodrow = True
        if depth_indexdf2.loc[i, 'cr_cdpth2'] is not None and depth_indexdf2.loc[i, 'cr_cdpth2'] >= .80:
                depth_rowvaluecount += 1
                goodrow = True
        if depth_indexdf2.loc[i, 'cr_cdpth3'] is not None and depth_indexdf2.loc[i, 'cr_cdpth3'] >= .80:
                depth_rowvaluecount += 1
                goodrow = True
        if depth_indexdf2.loc[i, 'cr_cdpth4'] is not None and depth_indexdf2.loc[i, 'cr_cdpth4'] >= .80:
                depth_rowvaluecount += 1
                goodrow = True
        if depth_indexdf2.loc[i, 'cr_cdpth5'] is not None and depth_indexdf2.loc[i, 'cr_cdpth5'] >= .80:
                depth_rowvaluecount += 1
                goodrow = True
        if depth_indexdf2.loc[i, 'cr_cdpth6'] is not None and depth_indexdf2.loc[i, 'cr_cdpth6'] >= .80:
                depth_rowvaluecount += 1
                goodrow = True
        if depth_indexdf2.loc[i, 'cr_cdpth7'] is not None and depth_indexdf2.loc[i, 'cr_cdpth7'] >= .80:
                depth_rowvaluecount += 1
                goodrow = True
        if depth_indexdf2.loc[i, 'cr_cdpth8'] is not None and depth_indexdf2.loc[i, 'cr_cdpth8'] >= .80:
                depth_rowvaluecount += 1
                goodrow = True
        if depth_indexdf2.loc[i, 'cr_cdpth9'] is not None and depth_indexdf2.loc[i, 'cr_cdpth9'] >= .80:
                depth_rowvaluecount += 1
                goodrow = True
        if depth_indexdf2.loc[i, 'cr_cdpth10'] is not None and depth_indexdf2.loc[i, 'cr_cdpth10'] >= .80:
                depth_rowvaluecount += 1
                goodrow = True
        if goodrow == True:
            depthrowcount += 1



    depth_df1 = depth_df.stack().reset_index()

    new_df = depth_df1[depth_df1["cr_record_id"].str.contains(site)]


    median = new_df.loc[:, 0].median()

    mean = new_df.loc[:, 0].mean()


    percent_10 = new_df.loc[:, 0].quantile(0.1)  # 10th percentile
    # print(percent_10)
    percent_90 = new_df.loc[:, 0].quantile(0.9)  # 10th percentile
    # print(percent_75)

    rateDivisions = ['Median', '10th Percentile', '90th Percentile']
    rateScores = [median, percent_10, percent_90]

    # Create the chart based on the data filtered and cleaned

    plt.grid(True)
    plt.minorticks_on()

    plt.figure(figsize=(3,4))
    plt.bar(rateDivisions, rateScores, color='darkblue', width=.3)


    plt.suptitle("Compression Depth ", fontsize=7)
    plt.title("Target: 5.0-6.0 \n Minutes meeting target: 710 (31.3%))", fontsize=6)
    plt.xticks(fontsize=7)
    plt.yticks(fontsize=7)

    if  depth_total_analyzable_min > 0:
        plt.title("Target: 5.0-6.0 \n Minutes meeting target: " + str(depth_rowvaluecount) + "   ({:.2%})".format(
            depth_rowvaluecount /  depth_total_analyzable_min), fontsize=6)
    else:
        plt.title("Target >=.80 \n Minutes meeting target: NO DATA ", fontsize=6)

    plt.axhline(nationalmean,color ='red')

    plt.savefig('C:/Charts/ChartImages/'+site+'cpr3.png', bbox_inches='tight')
    plt.close()

    greatestpause= 0
    greatestreason = ''

    fields2 = ['cr_ecstrttm1','cr_ecstoptm1','cr_rsnstp1','cr_ecstrttm2','cr_ecstoptm2','cr_rsnstp2','cr_ecstrttm3','cr_ecstoptm3','cr_rsnstp3',
               'cr_ecstrttm4','cr_ecstoptm4','cr_rsnstp4','cr_ecstrttm5','cr_ecstoptm5','cr_rsnstp5','cr_ecstrttm6','cr_ecstoptm6','cr_rsnstp6',
               'cr_ecstrttm7','cr_ecstoptm7','cr_rsnstp7','cr_ecstrttm8','cr_ecstoptm8','cr_rsnstp8','cr_ecstrttm9','cr_ecstoptm9','cr_rsnstp9',]
    rhythm_ck_df = project.export_records(format='df', fields=fields2)
    rhythm_ck_df_index = rhythm_ck_df.reset_index()

    cols = ['cr_rsnstp1','cr_rsnstp2','cr_rsnstp3','cr_rsnstp4','cr_rsnstp5','cr_rsnstp6','cr_rsnstp7','cr_rsnstp8','cr_rsnstp9']

    rhythm_ck_df_index[cols] = rhythm_ck_df_index[cols].apply(pd.to_numeric, errors='coerce')
    if site == 'SK':
        print('rhythm_ck_df_index')
        # print(rhythm_ck_df_index.head(250))
    totalcount = 0
    sitecount = 0
    sitetenless = 0
    nationalmean = 0
    sitelist = []
    totallist = []
    sitetenlesslist = []
    for i in range(len(rhythm_ck_df_index)):

        for num in range(1,9):
            secondnum = num +1
            if type(rhythm_ck_df_index.loc[i, 'cr_ecstrttm'+str(secondnum)]) == str and type(
                    rhythm_ck_df_index.loc[i, 'cr_ecstoptm'+str(num)]) == str and rhythm_ck_df_index.loc[i, 'cr_rsnstp'+str(num)] == 1:
                # print('OMH')
                difference = relativedelta.relativedelta(
                    datetime.strptime(rhythm_ck_df_index.loc[i, 'cr_ecstrttm'+str(secondnum)], "%Y-%m-%d %H:%M:%S"),
                    datetime.strptime(rhythm_ck_df_index.loc[i, 'cr_ecstoptm'+str(num)], "%Y-%m-%d %H:%M:%S"))

                diffseconds = int(difference.seconds)



                if diffseconds > 0: # getting the national average
                    if diffseconds > greatestpause:
                        greatestpause = diffseconds
                        greatestreason = 'Rhythm Check Pause'
                    totalcount = totalcount + 1
                    totallist.append(diffseconds)
                    print(rhythm_ck_df_index.loc[i, 'cr_record_id'])
                    if site in rhythm_ck_df_index.loc[i, 'cr_record_id']: #getting the total number of diff seconds and <10 second
                        sitecount = sitecount + 1
                        sitelist.append(diffseconds)
                        if diffseconds <= 10:
                           sitetenless = sitetenless + 1
                           sitetenlesslist.append(diffseconds)  #calculate medians and percentiles






    nationalmean = (statistics.mean(totallist))

    if sitecount > 0 :

        median = (statistics.median(sitetenlesslist))

        mean = (statistics.mean(sitetenlesslist))


        percent_10 = np.quantile(sitetenlesslist, .10) # 10th percentile np.quantile(arr, .50)

        percent_90 = np.quantile(sitetenlesslist, .90)  # 10th percentile


    rateDivisions = ['Median', '10th Percentile', '90th Percentile']
    rateScores = [median, percent_10, percent_90]


    plt.grid(True)
    plt.minorticks_on()

    plt.figure(figsize=(3,4))
    plt.bar(rateDivisions, rateScores, color='darkblue', width=.3)


    plt.suptitle("Rhythm Check Pauses ", fontsize=7)

    if sitecount > 0 :
        plt.title("Target: <=10\n Pauses meeting target: " + str(sitetenless ) + "   ({:.2%})".format(
        sitetenless / sitecount), fontsize=6)
    else:
        plt.title("Target: <=10\n Pauses meeting target: NO DATA", fontsize = 6)


    plt.xticks(fontsize=7)
    plt.yticks(fontsize=7)

    plt.axhline(nationalmean,color ='red')

    plt.savefig('C:/Charts/ChartImages/'+site+'cpr4.png', bbox_inches='tight')
    plt.close()


    if site == 'SK':
        print('rhythm_ck_df_index')
        # print(rhythm_ck_df_index.head(250))
    totalcount = 0
    sitecount = 0
    sitetenless = 0
    nationalmean = 0
    sitelist = []
    totallist = []
    sitetenlesslist = []
    for i in range(len(rhythm_ck_df_index)):

        for num in range(1, 9):
            secondnum = num + 1
            if type(rhythm_ck_df_index.loc[i, 'cr_ecstrttm' + str(secondnum)]) == str and type(
                    rhythm_ck_df_index.loc[i, 'cr_ecstoptm' + str(num)]) == str and rhythm_ck_df_index.loc[
                i, 'cr_rsnstp' + str(num)] == 2:
                # print('OMH')
                difference = relativedelta.relativedelta(
                    datetime.strptime(rhythm_ck_df_index.loc[i, 'cr_ecstrttm' + str(secondnum)], "%Y-%m-%d %H:%M:%S"),
                    datetime.strptime(rhythm_ck_df_index.loc[i, 'cr_ecstoptm' + str(num)], "%Y-%m-%d %H:%M:%S"))

                diffseconds = int(difference.seconds)


                if diffseconds > 0:  # getting the national average
                    totalcount = totalcount + 1
                    totallist.append(diffseconds)
                    print(rhythm_ck_df_index.loc[i, 'cr_record_id'])
                    if site in rhythm_ck_df_index.loc[
                        i, 'cr_record_id']:  # getting the total number of diff seconds and <10 second
                        sitecount = sitecount + 1
                        sitelist.append(diffseconds)
                        if diffseconds < 20:
                            sitetenless = sitetenless + 1
                            sitetenlesslist.append(diffseconds)  # calculate medians and percentiles



    nationalmean = (statistics.mean(totallist))

    if sitecount > 0:
        median = (statistics.median(sitetenlesslist))
        mean = (statistics.mean(sitetenlesslist))


        percent_10 = np.quantile(sitetenlesslist, .10)  # 10th percentile np.quantile(arr, .50)
        percent_90 = np.quantile(sitetenlesslist, .90)  # 10th percentile


    rateDivisions = ['Median', '10th Percentile', '90th Percentile']
    rateScores = [median, percent_10, percent_90]

    # Create the chart based on the data filtered and cleaned

    plt.grid(True)
    plt.minorticks_on()

    plt.figure(figsize=(3, 4))
    plt.bar(rateDivisions, rateScores, color='darkblue', width=.3)

    plt.suptitle("Peri-Shock Pauses ", fontsize=7)

    if sitecount > 0:
        plt.title("Target: <20 \n Pauses meeting target: " + str(sitetenless) + "   ({:.2%})".format(
            sitetenless / sitecount), fontsize=6)
    else:
        plt.title("Target: <20\n Pauses meeting target: NO DATA", fontsize=6)

    plt.xticks(fontsize=7)
    plt.yticks(fontsize=7)

    plt.axhline(nationalmean, color='red')

    plt.savefig('C:/Charts/ChartImages/' + site + 'cpr5.png', bbox_inches='tight')
    plt.close()



    totalcount = 0
    sitecount = 0
    sitetenless = 0
    nationalmean = 0
    sitelist = []
    totallist = []
    sitetenlesslist = []
    for i in range(len(rhythm_ck_df_index)):

        for num in range(1, 9):
            secondnum = num + 1
            if type(rhythm_ck_df_index.loc[i, 'cr_ecstrttm' + str(secondnum)]) == str and type(
                    rhythm_ck_df_index.loc[i, 'cr_ecstoptm' + str(num)]) == str and rhythm_ck_df_index.loc[
                i, 'cr_rsnstp' + str(num)] == 3:
                # print('OMH')
                difference = relativedelta.relativedelta(
                    datetime.strptime(rhythm_ck_df_index.loc[i, 'cr_ecstrttm' + str(secondnum)], "%Y-%m-%d %H:%M:%S"),
                    datetime.strptime(rhythm_ck_df_index.loc[i, 'cr_ecstoptm' + str(num)], "%Y-%m-%d %H:%M:%S"))

                diffseconds = int(difference.seconds)


                if diffseconds > 0:  # getting the national average
                    if diffseconds > greatestpause:
                        greatestpause = diffseconds
                        greatestreason = 'AirWay Pause'
                    totalcount = totalcount + 1
                    totallist.append(diffseconds)
                    print(rhythm_ck_df_index.loc[i, 'cr_record_id'])
                    if site in rhythm_ck_df_index.loc[
                        i, 'cr_record_id']:  # getting the total number of diff seconds and <10 second
                        sitecount = sitecount + 1
                        sitelist.append(diffseconds)
                        if diffseconds < 30:
                            sitetenless = sitetenless + 1
                            sitetenlesslist.append(diffseconds)  # calculate medians and percentiles



    nationalmean = (statistics.mean(totallist))

    if sitecount > 0:
        median = (statistics.median(sitetenlesslist))

        mean = (statistics.mean(sitetenlesslist))


        percent_10 = np.quantile(sitetenlesslist, .10)  # 10th percentile np.quantile(arr, .50)
        # print(percent_10)
        percent_90 = np.quantile(sitetenlesslist, .90)  # 10th percentile
        # print(percent_75)

    rateDivisions = ['Median', '10th Percentile', '90th Percentile']
    rateScores = [median, percent_10, percent_90]

    plt.grid(True)
    plt.minorticks_on()

    plt.figure(figsize=(3, 4))
    plt.bar(rateDivisions, rateScores, color='darkblue', width=.3)

    plt.suptitle("Airway Pauses ", fontsize=7)

    if sitecount > 0:
        plt.title("Target: <30 \n Pauses meeting target: " + str(sitetenless) + "   ({:.2%})".format(
            sitetenless / sitecount), fontsize=6)
    else:
        plt.title("Target: <30\n Pauses meeting target: NO DATA", fontsize=6)

    plt.xticks(fontsize=7)
    plt.yticks(fontsize=7)

    plt.axhline(nationalmean, color='red')

    plt.savefig('C:/Charts/ChartImages/' + site + 'cpr6.png', bbox_inches='tight')
    plt.close()


    totalcount = 0
    sitecount = 0
    sitetenless = 0
    nationalmean = 0
    sitelist = []
    totallist = []
    sitetenlesslist = []
    for i in range(len(rhythm_ck_df_index)):

        for num in range(1, 9):
            secondnum = num + 1
            if type(rhythm_ck_df_index.loc[i, 'cr_ecstrttm' + str(secondnum)]) == str and type(
                    rhythm_ck_df_index.loc[i, 'cr_ecstoptm' + str(num)]) == str and rhythm_ck_df_index.loc[
                i, 'cr_rsnstp' + str(num)] not in [0,1,2,3]:

                difference = relativedelta.relativedelta(
                    datetime.strptime(rhythm_ck_df_index.loc[i, 'cr_ecstrttm' + str(secondnum)], "%Y-%m-%d %H:%M:%S"),
                    datetime.strptime(rhythm_ck_df_index.loc[i, 'cr_ecstoptm' + str(num)], "%Y-%m-%d %H:%M:%S"))

                diffseconds = int(difference.seconds)


                if diffseconds > 0:  # getting the national average
                    if diffseconds > greatestpause:
                        greatestpause = diffseconds
                        greatestreason = 'Other Pause'
                    totalcount = totalcount + 1
                    totallist.append(diffseconds)
                    print(rhythm_ck_df_index.loc[i, 'cr_record_id'])
                    if site in rhythm_ck_df_index.loc[
                        i, 'cr_record_id']:  # getting the total number of diff seconds and <10 second
                        sitecount = sitecount + 1
                        sitelist.append(diffseconds)
                        sitetenless = sitetenless + 1
                        sitetenlesslist.append(diffseconds)  # calculate medians and percentiles


    nationalmean = (statistics.mean(totallist))

    if sitecount > 0:
        median = (statistics.median(sitetenlesslist))
        mean = (statistics.mean(sitetenlesslist))


        percent_10 = np.quantile(sitetenlesslist, .10)  # 10th percentile np.quantile(arr, .50)
        percent_90 = np.quantile(sitetenlesslist, .90)  # 10th percentile

    rateDivisions = ['Median', '10th Percentile', '90th Percentile']
    rateScores = [median, percent_10, percent_90]

    plt.grid(True)
    plt.minorticks_on()

    plt.figure(figsize=(3, 4))
    plt.bar(rateDivisions, rateScores, color='darkblue', width=.3)

    plt.suptitle("Other Pauses (sec)", fontsize=7)

    if sitecount <= 0:
        plt.title("NO DATA ", fontsize=6)


    plt.xticks(fontsize=7)
    plt.yticks(fontsize=7)

    plt.axhline(nationalmean, color='red')

    plt.savefig('C:/Users/mcbarnettr/Desktop/Charts/ChartImages/' + site + 'cpr7.png', bbox_inches='tight')
    plt.close()




    fname = "C:/Charts/"+site+"CPRProcessSummary.xlsx"
    workbook = Workbook()
    ws = workbook.active
    thin = Side(border_style="thin", color="000000")
    thinBorder = Border(top=thin, left=thin, right=thin, bottom=thin)
    double = Side(border_style="double", color="ff0000")
    # fill = PatternFill(patternType = 'solid',end_color = '34a8eb')
    fill = PatternFill(patternType='solid',start_color='5c61a1', end_color="5c61a1")
    ws['C1'] = 'Cardiac Arrest - CPR Process Summary Report'


    ws['C1'].font = Font(size=14, bold=True)

    x = datetime.now()
    ws['G2'] = x.strftime("%B %d, %Y %I:%M %p")

    ws['A2'] = 'Region: ' + site
    ws['A2'].font = Font(size =10)
    ws['G2'].font = Font(size=10)
    ws['A8'] = 'CPR Process Summary stats - all minutes          N=' + str(total_analyzable_min)
    ws['A8'].font = Font(bold=True)

    ws.merge_cells('A8:L8')
    fill = PatternFill(patternType='solid', start_color='bcbfeb', end_color="bcbfeb")
    ws['A8'].fill = fill
    ws.merge_cells('A4:F4')

    ws['A4'] = 'Total Number of Analyzable Records'

    ws.merge_cells('A5:F5')
    ws['A5'] = 'Total Number of Analyzable CPR Minutes *'
    ws.merge_cells('A6:F6')
    ws['A6'] = 'Episodes from ' + start_date +' to ' + end_date
    ws.merge_cells('G4:L4')
    ws['G4'] = str(number_of_rows) #Calculate
    ws.merge_cells('G5:L5')
    ws['G5'] = str(total_analyzable_min)  # Calculate
    ws.merge_cells('G6:L6')



    ws['G31'] =  'CPR Process Summary Stats - all cases N=303'
    ws['G31'].font=Font(bold=True)
    ws.merge_cells('G31:L31')
    ws['G31'].fill = fill
    ws['J32'] = 'Target'
    ws['K32'] = "CPR Meeting Target"
    ws['G33'] = 'Compression Rate (comps/min)'
    ws['K33'] =  str(ratecount) +  "   ({:.2%})".format(ratecount/number_of_rows)
    ws['J33'] = '100-120'
    ws.merge_cells('G34:I34')
    ws.merge_cells('G32:I32')
    ws['G34'] = 'Compression Fraction'
    ws['J34'] = '>=0.80'
    ws['K34'] = str(fractioncount) +  "   ({:.2%})".format(fractioncount/number_of_rows)
    ws.merge_cells('G35:I35')
    ws['G35'] = 'Compression Depth (cm)'
    ws['J35'] = '5.0-6.0'
    ws['K35'] = str(depthcount) +  "   ({:.2%})".format(depthcount/number_of_rows)
    ws['H35'].alignment = Alignment(wrap_text=True)
    ws.merge_cells('G36:I36')
    ws['G36'] = 'All 3 Measures'
    ws['J36'] = str(allThree) + "   ({:.2%})".format(allThree / number_of_rows)

    ws.merge_cells('A91:K91')
    ws.merge_cells('A92:B92')
    ws.merge_cells('D92:E92')
    ws.merge_cells('D93:E93')
    ws.merge_cells('A93:B93')
    ws.merge_cells('F93:G93')
    ws.merge_cells('H93:I93')
    ws.merge_cells('F94:G94')
    ws.merge_cells('H94:I94')
    ws.merge_cells('J92:K92')
    ws['A91'] = 'Other stats - all cases      N='
    ws['A91'].fill = fill
    ws['C92'] = 'Target'
    ws['D92'] = 'Median'
    ws['F92'] = '10th percentile'
    ws['H92'] = '90th percentile'
    ws['J92'] = 'National average'
    ws['A94'] = 'Longest Pause'
    ws['D94'] =  str(greatestpause) +' (' + str(greatestreason) +')'




    for row in ws['A4:K6']:
        for cell in row:
            cell.border = thinBorder  # A5:D6 area cells set borders
            cell.font = Font(size=10)


    for row in ws['A4:L6']:
        for cell in row:
            cell.border = thinBorder  # A5:D6 area cells set borders
            cell.font = Font(size=10)

    for row in ws['G31:L36']:
        for cell in row:
            cell.font = Font(size=10)
            cell.border = thinBorder # A5:D6 area cells set borders
    ws['G31'].font = Font(size=10, bold=True)

    for row in ws['A91:K94']:
        for cell in row:
            cell.font = Font(size=10)
            cell.border = thinBorder  # A5:D6 area cells set borders
    ws['A31'].font = Font(size=10, bold=True)

    img1 = Image('C:/Charts/ChartImages/'+site+'cpr1.png')
    img2=Image('C:/Charts/ChartImages/'+site+'cpr2.png')
    img3=Image('C:/Charts/ChartImages/'+site+'cpr3.png')
    img4= Image('C:/Charts/ChartImages/'+site+'cpr4.png')
    img5= Image('C://Charts/ChartImages/' + site + 'cpr5.png')
    img6= Image('C://Charts/ChartImages/' + site + 'cpr6.png')
    img7 = Image('C:/Charts/ChartImages/' + site + 'cpr7.png')
    ws.add_image(img1, 'A10')
    ws.add_image(img2, 'G10')
    ws.add_image(img3, 'A30')
    ws.add_image(img4, 'A50')
    ws.add_image(img5,'G50')
    ws.add_image(img6,'A70')
    ws.add_image(img7, 'G70')




    ws.page_setup.paperHeight = '105mm'
    ws.page_setup.paperWidth = '148mm'

    workbook.save(filename=fname)

fig, ax = plt.subplots()
CreateSiteLevelCharts('BC',ax)









