import pandas as pd
import matplotlib.pyplot as plt
import pandas as pd
import gc
import json
# import config
from openpyxl import load_workbook
from redcap import Project, RedcapError
from datetime import datetime
from dateutil import relativedelta
from datetime import timedelta
from smtplib import SMTP
import numpy as np
import math
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.drawing.image import Image
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles import  Alignment,Font
from openpyxl.worksheet.pagebreak import Break
import statistics
import sys

URL = 'https://redcap.smh.ca/redcap/api/'
API_KEY = '' #CANROCPROJECT
project = Project(URL, API_KEY)


#This function pulls all the data necessary for the Adult Udestein charts, using the specified fields. The data is then filtered and cleaned as necessary to produce the charts.
#

def CreateSiteLevelCharts(site):



    allfields = ['cr_record_id','cr_epdt','cr_tx','cr_pyhalt','cr_witbys','cr_estageu','cr_estagev','cr_agecat','cr_scause','cr_cpratt','cr_loctyp','cr_aedapp','cr_aedshk','cr_frhyem',
                 'cr_surv','cr_rosc','cr_pdisp', 'cr_numshk','cr_ptmrcv', 'cr_rig1tm', 'cr_ptmdsp','cr_rig1dtm', 'cr_epi','cr_arryth','cr_lmasuc','cr_igelsuc',
                 'cr_ptmrcv','cr_ivtm','cr_iotm','cr_kingsuc', 'cr_ettsuc', 'cr_othawsuc','cr_rig1tm','cr_rig2tm','cr_rig3tm','cr_rig4tm','cr_v1sl','cr_v2sl','cr_v3sl',
                 'cr_v4sl','cr_lmatm','cr_kingtm','cr_igeltm','cr_etttm','cr_othawtm','cr_epitm','cr_arrythtm','cr_prosc','cr_ptmcpr']


    df = project.export_records(format='df', fields=allfields)
    # allsitesdf = df - (used for national mean etc)

    filter_date = []

    #The cr_epdt date value is used to filter the charts for start date and end date.

    for value in df["cr_epdt"]:

        if str(value) != 'nan':
            filter_date.append(datetime.strptime(value, "%Y-%m-%d"))
        else:
            filter_date.append(None)

    pd.set_option("display.max.columns", None)
    pd.set_option('display.max_rows', None)

    df["filter_date"] = filter_date
    start_date = input('Start Date YYYY-MM-DD').strip()
    start_date_changed =  datetime.strptime(start_date,"%Y-%m-%d")
    end_date = input('End Date YYYY-MM-DD').strip()
    end_date_changed = datetime.strptime(end_date, "%Y-%m-%d")

    #Filter the dataset according to the start and end date.

    df = df[(df['filter_date'] >= start_date_changed) & (df['filter_date'] <=end_date_changed) ]
    df = df[(df['cr_scause'] != 17)]


    if site != 'ALL':
        df = df.reset_index()
        df = df[df["cr_record_id"].str.contains(site)]
        df.reset_index(drop=True, inplace=True)
        print(df.head(1000))
    else:
        print('yeah ALL continues')

    pd.set_option('display.max_columns', None)

    value_list = [0, 1, '0', '1']
    value_list2 = [1, 2, '1', '2']
    allarrests_total = len(df.index)

    #These totals sre used as the denominator when calculating the percentage values for each chart

    Ntotal = allarrests_total
    Ntotal = len(df.loc[(df['cr_scause'] != 17)].index)
    Adulttotal = len(df.loc[(df['cr_estageu'] == 0) & (df['cr_estagev'] > 17) & (df['cr_scause'] != 17) ].index)
    Paediatrictotal =  len(df.loc[ (df['cr_scause'] != 17) & (((df['cr_estageu'] == 0) & (df['cr_estagev']< 18))
                                     | df.cr_estageu.isin(value_list2)  |  df.cr_agecat.isin(value_list))].index)
    Unknowntotal = len(df.loc[(df['cr_scause'] != 17) & ((df['cr_estageu'] == 3) |  (df['cr_agecat']==3))].index)

    #Populate list of values for each bar of chart.
    NList = []
    AdultList = []
    Paediatricslist = []
    UnknownList = []

    # FIG1  All Arrests Attended by Service, breakdown by age/etiology

    for x in range(1,4):

        if x == 1:
            N = len(df.loc[(df['cr_tx'] == 1) ].index)

            Npercent = 0
            if N not in [0]:
               Npercent = round((N / allarrests_total) , 4)

            NList.append(Npercent)


            Adult = len(df.loc[(df['cr_tx'] == 1) & (df['cr_estageu'] == 0) & (df['cr_estagev'] > 17)].index)
            Adultpercent = 0
            if Adult not in [0]:
                Adultpercent = round((Adult / Adulttotal) , 4)
            AdultList.append(Adultpercent)


            value_list = [0,1,'0','1']
            value_list2 = [1,2,'1','2']


            Paediatrics = len(df.loc[((df['cr_tx'] == 1) & (
                                                             ((df['cr_estageu'] == 0) & (df['cr_estagev'] < 18))  |
                                                             (df.cr_estageu.isin(value_list2))
                                                           ))
                                     ].index)
            Paediatricspercent = 0
            if Paediatrics not in [0]:
                Paediatricspercent = round((Paediatrics /  Paediatrictotal) , 4)
            Paediatricslist.append(Paediatricspercent)


            Unknown = len(df.loc[(df['cr_tx'] == 1) & ((df['cr_estageu'] == 3) |  (df['cr_agecat']==3))].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknowntotal), 4)
            UnknownList.append(Unknownpercent)


        elif x ==2:
            value_list1 = [1, 2,'1','2']
            N = len(df.loc[(df['cr_tx'] == 0)].index)
            # cr_pyhalt = 1 or 2
            Npercent = 0
            if N not in [0]:
                Npercent = round((N / allarrests_total), 4)
            NList.append(Npercent)

            Adult = len(df.loc[(df['cr_tx'] == 0) & (df['cr_estageu'] == 0) & (df['cr_estagev'] > 17)].index)
            Adultpercent = 0
            if Adult not in [0]:
                Adultpercent = round((Adult / Adulttotal), 4)
            AdultList.append(Adultpercent)

            value_list1 = [0, 1,'0','1']
            value_list2 = [1, 2,'1','2']

            Paediatrics = len(df.loc[(df['cr_tx'] == 0) & ((df['cr_estageu'] == 0) & (df['cr_estagev'] < 18) | (df.cr_estageu.isin(value_list2)))].index)

            Paediatricspercent = 0
            if Paediatrics not in [0]:
               Paediatricspercent = round((Paediatrics / Paediatrictotal), 4)
            Paediatricslist.append(Paediatricspercent)

            Unknown = len(df.loc[(df['cr_tx'] == 0) & ((df['cr_estageu'] == 3) |  (df['cr_agecat']==3))].index)

            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknowntotal) , 4)
            UnknownList.append(Unknownpercent)


        elif x == 3:
            N = len(df.loc[((df['cr_tx'] == 1) & (df['cr_witbys'] == 2))].index)

            Npercent = 0
            if N not in [0]:
               Npercent = round((N / allarrests_total), 4)
            NList.append(Npercent)


            Adult = len(df.loc[((df['cr_tx'] == 1) & (df['cr_witbys'] == 2)) & ((df['cr_estageu'] == 0) & (df['cr_estagev'] > 17))].index)

            Adultpercent = 0
            if Adult not in [0]:

                Adultpercent = round((Adult / Adulttotal), 4)
            AdultList.append(Adultpercent)

            value_list = [0, 1,'0','1']
            Paediatrics = len(df.loc[(((df['cr_tx'] == 1) & (df['cr_witbys'] == 2)) & ((df['cr_estageu'] == 0) & (df['cr_estagev'] < 18)))
                                     |(((df['cr_tx'] == 1) & (df['cr_witbys'] == 2)) & ((df.cr_estageu.isin(value_list2)) ))
                                      ].index)

            Paediatricspercent = 0
            if Paediatrics not in [0]:
                Paediatricspercent = round((Paediatrics / Paediatrictotal), 4)
            Paediatricslist.append(Paediatricspercent)

            Unknown = len(df.loc[(df['cr_tx'] == 1) & (df['cr_witbys'] == 2) &((df['cr_estageu'] == 3) | (df['cr_agecat'] == 3))].index)


            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknowntotal), 4)
            UnknownList.append(Unknownpercent)


    #UPDATE EXCEL OPTION

    AdultUstenworkbook = load_workbook(filename="C:/AdultUsteinGraphs/Adult Utstein template.xlsx")
    AdultUpsteinWorksheet = AdultUstenworkbook['Sheet2']

    AdultUpsteinWorksheet['B1'] = 'All (' + str(Ntotal)  +')'
    AdultUpsteinWorksheet['B2'] = NList[0]
    AdultUpsteinWorksheet['B3'] = NList[1]
    AdultUpsteinWorksheet['B4'] = NList[2]

    AdultUpsteinWorksheet['C1'] = 'Adult ≥18yrs (' + str(Adulttotal) + ')'
    AdultUpsteinWorksheet['C2'] = AdultList[0]
    AdultUpsteinWorksheet['C3'] = AdultList[1]
    AdultUpsteinWorksheet['C4'] = AdultList[2]

    AdultUpsteinWorksheet['D1'] = 'Paediatrics <18yrs (' + str(Paediatrictotal) + ')'
    AdultUpsteinWorksheet['D2'] = Paediatricslist[0]
    AdultUpsteinWorksheet['D3'] = Paediatricslist[1]
    AdultUpsteinWorksheet['D4'] = Paediatricslist[2]

    AdultUpsteinWorksheet['E1'] = 'Unknown (' + str(Unknowntotal) + ')'
    AdultUpsteinWorksheet['E2'] = UnknownList[0]
    AdultUpsteinWorksheet['E3'] = UnknownList[1]
    AdultUpsteinWorksheet['E4'] = UnknownList[2]


    AdultUstenworkbook.save("C:/AdultUsteinGraphs/CanRoc/Adult_Utstein_"+site+".xlsx")

    print('Completed - Fig 1')


    # FIG 2

    # Resuscitation  attempted by EMS

    allresuscitation_total = len(df.loc[(df['cr_tx'] == 1) ].index)
    Ntotal = allresuscitation_total
    Adulttotal = len(df.loc[(df['cr_estageu'] == 0) & (df['cr_estagev'] > 17) & (df['cr_tx'] == 1)].index)
    Paediatrictotal = len(df.loc[(df['cr_tx'] == 1) &(((df['cr_estageu'] == 0) & (df['cr_estagev'] < 18))
                                 | df.cr_estageu.isin(value_list2) )].index)
    Unknowntotal = len(df.loc[(df['cr_tx'] == 1)& (df['cr_estageu'] == 3) | (df['cr_agecat'] == 3)].index)

    NList = []
    AdultList = []
    Paediatricslist = []
    UnknownList = []

    for x in range(1, 5):
        if x == 1:
            N = len(df.loc[(df['cr_scause'] != 17) & (df['cr_tx'] == 1) ].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal) , 4)
            NList.append(Npercent)

            Adult = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 17) & (df['cr_estageu'] == 0) & (df['cr_estagev'] > 17)].index)

            Adultpercent = 0
            if Adult not in [0]:
                Adultpercent = round((Adult / Adulttotal) , 4)
            AdultList.append(Adultpercent)


            value_list = [0,1,'0','1']
            value_list2 = [1,2,'1','2']

            Paediatrics = len(df.loc[((df['cr_tx'] == 1)  & (df['cr_scause'] != 17) & (df['cr_estageu'] == 0) & (df['cr_estagev']< 18))
                                     | ((df['cr_tx'] == 1)  & (df['cr_scause'] != 17) & (df.cr_estageu.isin(value_list2))) | ((df['cr_tx'] == 1) & (df['cr_scause'] != 17)& (df.cr_agecat.isin(value_list)))].index)

            Paediatricspercent = 0
            if Paediatrics not in [0]:
                Paediatricspercent = round((Paediatrics / Paediatrictotal) , 4)
            Paediatricslist.append(Paediatricspercent)

            Unknown = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 17) & (
                        (df['cr_estageu'] == 3) | (df['cr_agecat'] == 3))].index)


            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknowntotal), 4)
            UnknownList.append(Unknownpercent)




        elif x ==2:
            N = len(df.loc[(df['cr_scause'] == 0) & (df['cr_tx'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            Adult = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_estageu'] == 0) & (
                        df['cr_estagev'] > 17)].index)

            Adultpercent = 0
            if Adult not in [0]:
                Adultpercent = round((Adult / Adulttotal), 4)
            AdultList.append(Adultpercent)

            value_list = [0, 1, '0', '1']
            value_list2 = [1, 2, '1', '2']

            Paediatrics = len(df.loc[((df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_estageu'] == 0) & (
                        df['cr_estagev'] < 18))
                                     | ((df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (
                df.cr_estageu.isin(value_list2))) | ((df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (
                df.cr_agecat.isin(value_list)))].index)

            Paediatricspercent = 0
            if Paediatrics not in [0]:
                Paediatricspercent = round((Paediatrics / Paediatrictotal), 4)
            Paediatricslist.append(Paediatricspercent)

            Unknown = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (
                    (df['cr_estageu'] == 3) | (df['cr_agecat'] == 3))].index)


            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknowntotal), 4)
            UnknownList.append(Unknownpercent)



        elif x == 3:
            N = len(df.loc[(df['cr_scause'] != 0) & (df['cr_scause'] != 17) & (df['cr_tx'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            Adult = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 0) &   (df['cr_scause'] != 17) &(df['cr_estageu'] == 0) & (
                    df['cr_estagev'] > 17)].index)

            Adultpercent = 0
            if Adult not in [0]:
                Adultpercent = round((Adult / Adulttotal), 4)
            AdultList.append(Adultpercent)

            value_list = [0, 1, '0', '1']
            value_list2 = [1, 2, '1', '2']

            Paediatrics = len(df.loc[((df['cr_tx'] == 1) & (df['cr_scause'] != 0)  & (df['cr_scause'] != 17)& (df['cr_estageu'] == 0) & (
                    df['cr_estagev'] < 18))
                                     | ((df['cr_tx'] == 1) & (df['cr_scause'] != 0) & (
                df.cr_estageu.isin(value_list2))) | ((df['cr_tx'] == 1) & (df['cr_scause'] != 0) & (
                df.cr_agecat.isin(value_list)))].index)

            Paediatricspercent = 0
            if Paediatrics not in [0]:
                Paediatricspercent = round((Paediatrics / Paediatrictotal), 4)
            Paediatricslist.append(Paediatricspercent)

            Unknown = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 0)  & (df['cr_scause'] != 17)& (
                    (df['cr_estageu'] == 3) | (df['cr_agecat'] == 3))].index)


            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknowntotal), 4)
            UnknownList.append(Unknownpercent)



        elif x == 4:

            value_list3 = [None,'',18,'18']


            N = len(df.loc[(df.cr_scause.isin(value_list3))& (df['cr_tx'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / allarrests_total) , 4)
            NList.append(Npercent)

            Adult = len(df.loc[(df['cr_tx'] == 1) & (df.cr_scause.isin(value_list3)) & (df['cr_estageu'] == 0) & (
                    df['cr_estagev'] > 17)].index)

            Adultpercent = 0
            if Adult not in [0]:
                Adultpercent = round((Adult / Adulttotal) , 4)
            AdultList.append(Adultpercent)

            value_list = [0, 1,'0','1']
            value_list2 = [1, 2,'1','2']

            Paediatrics = len(
                df.loc[(df['cr_tx'] == 1) & (df.cr_scause.isin(value_list3)) & (((df['cr_estageu'] == 0) & (df['cr_estagev'] < 18))
                       |  (df.cr_estageu.isin(value_list2)) | (df['cr_tx'] == 1) & (df.cr_agecat.isin(value_list)))].index)

            Paediatricspercent = 0
            if Paediatrics not in [0]:
                Paediatricspercent = round((Paediatrics / Paediatrictotal), 4)
            Paediatricslist.append(Paediatricspercent)

            Unknown = len(df.loc[(df['cr_tx'] == 1) & (df.cr_scause.isin(value_list3)) & (
                    (df['cr_estageu'] == 3) | (df['cr_agecat'] == 3))].index)


            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknowntotal), 4)
            UnknownList.append(Unknownpercent)

    #Populate the excel sheets template

    AdultUpsteinWorksheet = AdultUstenworkbook['Sheet2']

    AdultUpsteinWorksheet['B5'] = 'All (' + str(Ntotal) + ')'
    AdultUpsteinWorksheet['B6'] = NList[0]
    AdultUpsteinWorksheet['B7'] = NList[1]
    AdultUpsteinWorksheet['B8'] = NList[2]
    AdultUpsteinWorksheet['B9'] = NList[3]

    AdultUpsteinWorksheet['C5'] = 'Adult ≥18yrs (' + str(Adulttotal) + ')'
    AdultUpsteinWorksheet['C6'] = AdultList[0]
    AdultUpsteinWorksheet['C7'] = AdultList[1]
    AdultUpsteinWorksheet['C8'] = AdultList[2]
    AdultUpsteinWorksheet['C9'] = AdultList[3]

    AdultUpsteinWorksheet['D5'] = 'Paediatrics <18yrs (' + str(Paediatrictotal) + ')'
    AdultUpsteinWorksheet['D6'] = Paediatricslist[0]
    AdultUpsteinWorksheet['D7'] = Paediatricslist[1]
    AdultUpsteinWorksheet['D8'] = Paediatricslist[2]
    AdultUpsteinWorksheet['D9'] = Paediatricslist[3]

    AdultUpsteinWorksheet['E5'] = 'Unknown (' + str(Unknowntotal) + ')'
    AdultUpsteinWorksheet['E6'] = UnknownList[0]
    AdultUpsteinWorksheet['E7'] = UnknownList[1]
    AdultUpsteinWorksheet['E8'] = UnknownList[2]
    AdultUpsteinWorksheet['E9'] = UnknownList[3]

    AdultUstenworkbook.save("C:/Users/AdultUsteinGraphs/CanRoc/Adult_Utstein_" + site + ".xlsx")

    print('Completed - Fig 2')


    # FIG 3

    # Resuscitation attempted, breakdown by witnessed/bystander CPR

    value_list3 = ['', None]


    allresuscitation_total = len(df.loc[(df['cr_tx'] == 1)].index)
    Ntotal = allresuscitation_total
    Unwitnessedtotal =  len(df.loc[(df['cr_tx'] == 1)& (df['cr_witbys'] == 0)].index)
    Bystandertotal = len(df.loc[(df['cr_tx'] == 1)& (df['cr_witbys'] == 1)].index)
    EMStotal = len(df.loc[ (df['cr_tx'] == 1) & (df['cr_witbys'] == 2)].index)
    Unknowntotal  = len(df.loc[  (df['cr_tx'] == 1) & (df.cr_witbys.isin(value_list3))].index)


    NList = []
    UnwitnessedList = []
    BystanderList = []
    EMSList = []
    UnknownList=[]

    for x in range(1, 3):
        if x == 1:
            N = len(df.loc[(df['cr_cpratt'] == 1) & (df['cr_tx'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal) , 4)
            NList.append(Npercent)

            Unwitnessed = len(df.loc[(df['cr_cpratt'] == 1) & (df['cr_tx'] == 1)& (df['cr_witbys'] == 0)].index)

            Unwitnessedpercent = 0
            if Unwitnessed not in [0]:
                Unwitnessedpercent = round((Unwitnessed / Unwitnessedtotal) , 4)
            UnwitnessedList.append(Unwitnessedpercent)

            Bystander =  len(df.loc[(df['cr_cpratt'] == 1) & (df['cr_tx'] == 1)& (df['cr_witbys'] == 1)].index)

            Bystanderpercent = 0
            if Bystander not in [0]:
                Bystanderpercent = round((Bystander / Bystandertotal), 4)
            BystanderList.append(Bystanderpercent)

            EMS = len(df.loc[(df['cr_cpratt'] == 1) & (df['cr_tx'] == 1) & (df['cr_witbys'] == 2)].index)

            EMSpercent = 0
            if  EMS not in [0]:
                EMSpercent = round(( EMS / EMStotal), 4)
            EMSList.append( EMSpercent)



            Unknown = len(df.loc[(df['cr_cpratt'] == 1) & (df['cr_tx'] == 1) & (df.cr_witbys.isin(value_list3))].index)

            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknowntotal) , 4)
            UnknownList.append(Unknownpercent)

        elif x ==2:
            N = len(df.loc[(df['cr_cpratt'] == 0) & (df['cr_tx'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal) , 4)
            NList.append(Npercent)

            Unwitnessed = len(df.loc[(df['cr_cpratt'] == 0) & (df['cr_tx'] == 1) & (df['cr_witbys'] == 0)].index)

            Unwitnessedpercent = 0
            if Unwitnessed not in [0]:
                Unwitnessedpercent = round((Unwitnessed / Unwitnessedtotal), 4)
            UnwitnessedList.append(Unwitnessedpercent)

            Bystander = len(df.loc[(df['cr_cpratt'] == 0) & (df['cr_tx'] == 1) & (df['cr_witbys'] == 1)].index)

            Bystanderpercent = 0
            if Bystander not in [0]:
                Bystanderpercent = round((Bystander / Bystandertotal) , 4)
            BystanderList.append(Bystanderpercent)

            EMS = len(df.loc[(df['cr_cpratt'] == 0) & (df['cr_tx'] == 1) & (df['cr_witbys'] == 2)].index)

            EMSpercent = 0
            if EMS not in [0]:
                EMSpercent = round((EMS / EMStotal) , 4)
            EMSList.append(EMSpercent)

            value_list3 = ['', None]

            Unknown = len(df.loc[(df['cr_cpratt'] == 0) & (df['cr_tx'] == 1) & (df.cr_witbys.isin(value_list3))].index)

            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknowntotal), 4)
            UnknownList.append(Unknownpercent)

    AdultUpsteinWorksheet = AdultUstenworkbook['Sheet2']
    AdultUpsteinWorksheet['B27'] = 'All (' + str(Ntotal) + ')'
    AdultUpsteinWorksheet['B28'] = NList[0]



    AdultUpsteinWorksheet['C27'] = 'Unwitnessed (' + str(Unwitnessedtotal) + ')'
    AdultUpsteinWorksheet['C28'] = UnwitnessedList[0]


    AdultUpsteinWorksheet['D27'] = 'Bystander Witnessed (' + str(Bystandertotal) + ')'
    AdultUpsteinWorksheet['D28'] = BystanderList[0]


    AdultUpsteinWorksheet['E27'] = 'EMS Witnessed (' + str(EMStotal) + ')'
    AdultUpsteinWorksheet['E28'] = EMSList[0]


    AdultUpsteinWorksheet['F27'] = 'Unknown (' + str(Unknowntotal) + ')'
    AdultUpsteinWorksheet['F28'] = UnknownList[0]


    AdultUstenworkbook.save("C:/Charts/AdultUsteinGraphs/CanRoc/Adult_Utstein_" + site + ".xlsx")

    print('Completed - Fig 3')


    # FIG 4 Bystander resuscitation, by location (Witnessed by EMS excluded)
    value_list3 = [0, 1, '0', '1']
    value_list4 = [4, 5, 6, 7, 8, 9, 11, '4', '5', '6', '7', '8', '9', '11']


    allarrests_total = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list3)].index)
    alllocationstotal = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list3)].index)
    publiclocationstotal =  len(df.loc[ (df['cr_tx'] == 1) & df.cr_witbys.isin(value_list3) & df.cr_loctyp.isin(value_list4)].index)


    AllLocationsList = []
    PublicLocationsList = []



    for x in range(1, 3):
        if x == 1:
            AllLocations = len(df.loc[(df['cr_cpratt'] == 1) & (df['cr_tx'] == 1)& df.cr_witbys.isin(value_list3)].index)

            AllLocationspercent = 0
            if AllLocations not in [0]:
                AllLocationspercent = round(( AllLocations / alllocationstotal) , 4)
            AllLocationsList.append(AllLocationspercent)

            PublicLocations = len(df.loc[(df['cr_cpratt'] == 1) & (df['cr_tx'] == 1)& df.cr_witbys.isin(value_list3) & df.cr_loctyp.isin(value_list4)].index)

            PublicLocationspercent = 0
            if PublicLocations not in [0]:
                PublicLocationspercent = round(( PublicLocations /  publiclocationstotal) , 4)
            PublicLocationsList.append(PublicLocationspercent)


        elif x ==2:
            AllLocations = len(
                df.loc[(df['cr_aedapp'] == 1) & (df['cr_tx'] == 1) & df.cr_witbys.isin(value_list3)].index)

            AllLocationspercent = 0
            if AllLocations not in [0]:
                AllLocationspercent = round((AllLocations / alllocationstotal) , 4)
            AllLocationsList.append(AllLocationspercent)

            PublicLocations = len(df.loc[(df['cr_aedapp'] == 1) & (df['cr_tx'] == 1) & df.cr_witbys.isin(
                value_list3) & df.cr_loctyp.isin(value_list4)].index)

            PublicLocationspercent = 0
            if PublicLocations not in [0]:
                PublicLocationspercent = round((PublicLocations / publiclocationstotal) , 4)
            PublicLocationsList.append(PublicLocationspercent)

    AdultUpsteinWorksheet = AdultUstenworkbook['Sheet2']

    AdultUpsteinWorksheet['B32'] = 'All locations (' + str(alllocationstotal) + ')'
    AdultUpsteinWorksheet['B33'] =  AllLocationsList[0]
    AdultUpsteinWorksheet['B34'] =  AllLocationsList[1]

    AdultUpsteinWorksheet['C32'] = 'Public locations (' + str(publiclocationstotal) + ')'
    AdultUpsteinWorksheet['C33'] =  PublicLocationsList[0]
    AdultUpsteinWorksheet['C34'] =  PublicLocationsList[1]

    AdultUstenworkbook.save("C:/Charts/AdultUsteinGraphs/CanRoc/Adult_Utstein_" + site + ".xlsx")

    print('Completed - Fig 4')




    # FIG 5 PAD applied and shock delivered

    value_list3 = [0, 1, '0', '1']
    value_list4 = [4, 5, 6, 7, 8, 9, 11, '4', '5', '6', '7', '8', '9', '11']
    value_list_residential = [0, 1,12, '0', '1','12']
    value_list_private = [2, 10, '2', '10']

    alllocations_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_aedapp'] == 1)].index)

    AllLocationsList = []
    PublicLocationsList = []

    for x in range(1, 3):
        if x == 1:
            AllLocations = len(df.loc[(df['cr_tx'] == 1) & (df['cr_aedapp'] == 1) & (df['cr_aedshk'] == 1)].index)

            AllLocationspercent = 0
            if AllLocations not in [0]:
                AllLocationspercent = round((AllLocations / alllocations_total), 4)
            AllLocationsList.append(AllLocationspercent)

            PublicLocations = len(df.loc[(df['cr_tx'] == 1) & (df['cr_aedapp'] == 1) & (
                        df['cr_aedshk'] == 1) & df.cr_loctyp.isin(value_list4)].index)

            PublicLocationspercent = 0
            if PublicLocations not in [0]:
                PublicLocationspercent = round((PublicLocations / alllocations_total), 4)
            AllLocationsList.append(PublicLocationspercent)

            ResidentialLocations = len(df.loc[(df['cr_tx'] == 1) & (df['cr_aedapp'] == 1) & (
                        df['cr_aedshk'] == 1) & df.cr_loctyp.isin(value_list_residential)].index)
            Residentialpercent = 0
            if ResidentialLocations not in [0]:
                Residentialpercent = round((ResidentialLocations / alllocations_total), 4)
            AllLocationsList.append(Residentialpercent)

            communal_locations = len(df.loc[(df['cr_tx'] == 1) & (df['cr_aedapp'] == 1) & (df['cr_aedshk'] == 1) & (
                        df['cr_loctyp'] == 3)].index)

            communal_percent = 0
            if communal_locations not in [0]:
                communal_percent = round((communal_locations / alllocations_total), 4)
            AllLocationsList.append(communal_percent)

            private_locations = len(df.loc[(df['cr_tx'] == 1) & (df['cr_aedapp'] == 1) & (
                    df['cr_aedshk'] == 1) & df.cr_loctyp.isin(value_list_private)].index)
            private_percent = 0
            if private_locations not in [0]:
                private_percent = round((private_locations / alllocations_total), 4)
            AllLocationsList.append(private_percent)

    AdultUpsteinWorksheet = AdultUstenworkbook['Sheet2']

    AdultUpsteinWorksheet['B35'] = 'All locations (' + str(AllLocations) + ')'
    AdultUpsteinWorksheet['B36'] = AllLocationsList[0]

    AdultUpsteinWorksheet['C35'] = 'Public locations (' + str(PublicLocations) + ')'
    AdultUpsteinWorksheet['C36'] = AllLocationsList[1]

    AdultUpsteinWorksheet['D35'] = 'Residential (' + str(ResidentialLocations) + ')'
    AdultUpsteinWorksheet['D36'] = AllLocationsList[2]

    AdultUpsteinWorksheet['E35'] = 'Communal (' + str(communal_locations) + ')'
    AdultUpsteinWorksheet['E36'] = AllLocationsList[3]

    AdultUpsteinWorksheet['F35'] = 'Other Private (' + str(private_locations) + ')'
    AdultUpsteinWorksheet['F36'] = AllLocationsList[4]

    AdultUstenworkbook.save("C:/Charts/AdultUsteinGraphs/CanRoc/Adult_Utstein_" + site + ".xlsx")

    print('Completed - Fig 5')




    # Fig 6. graph Etiology, breakdown by presenting rhythm

    value_list5 = [2, 3, 5, '2', '3', '5']
    value_list6 = [18,'','18']
    value_list4 = ['', None]

    alletiology_total = len(df.loc[(df['cr_tx'] == 1)].index)
    noObvious_total= len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0)  ].index)
    TraumaObvious_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 0)  ].index)
    Unknown_total = len(df.loc[(df['cr_tx'] == 1) & df.cr_scause.isin(value_list4)  ].index)
    NList = []
    noObviousList = []
    TraumaObviousList = []
    UnknownList = []

    for x in range(1, 6):
        if x == 1:
            N = len(df.loc[(df['cr_tx'] == 1) & df.cr_frhyem.isin(value_list5) ].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / alletiology_total), 4)
            NList.append(Npercent)

            noObvious = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_frhyem.isin(value_list5) ].index)

            noObviouspercent = 0
            if noObvious not in [0]:
                noObviouspercent = round((noObvious/  noObvious_total), 4)
            noObviousList.append(noObviouspercent)

            TraumaObvious = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 0) & df.cr_frhyem.isin(value_list5) ].index)

            TraumaObviouspercent = 0
            if TraumaObvious not in [0]:
                TraumaObviouspercent = round((TraumaObvious/ TraumaObvious_total), 4)
            TraumaObviousList.append(TraumaObviouspercent)


            Unknown = len(df.loc[(df['cr_tx'] == 1) & df.cr_scause.isin(value_list4) & df.cr_frhyem.isin(value_list5) ].index)

            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)

        elif x == 2:
            N = len(df.loc[(df['cr_tx'] == 1) & (df['cr_frhyem'] == 1) ].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            noObvious = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 1)].index)

            noObviouspercent = 0
            if noObvious not in [0]:
                noObviouspercent = round((noObvious / noObvious_total), 4)
            noObviousList.append(noObviouspercent)

            TraumaObvious = len(
                df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 0) & (df['cr_frhyem'] == 1)].index)

            TraumaObviouspercent = 0
            if TraumaObvious not in [0]:
                TraumaObviouspercent = round((TraumaObvious / TraumaObvious_total), 4)
            TraumaObviousList.append(TraumaObviouspercent)

            Unknown = len(
                df.loc[(df['cr_tx'] == 1) & df.cr_scause.isin(value_list4) & (df['cr_frhyem'] == 1)].index)

            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)

        elif x == 3:
            N = len(df.loc[(df['cr_tx'] == 1) & (df['cr_frhyem'] == 0)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            noObvious = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 0)].index)

            noObviouspercent = 0
            if noObvious not in [0]:
                noObviouspercent = round((noObvious / noObvious_total), 4)
            noObviousList.append(noObviouspercent)

            TraumaObvious = len(
                df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 0) & (df['cr_frhyem'] == 0)].index)

            TraumaObviouspercent = 0
            if TraumaObvious not in [0]:
                TraumaObviouspercent = round((TraumaObvious / TraumaObvious_total), 4)
            TraumaObviousList.append(TraumaObviouspercent)

            Unknown = len(
                df.loc[(df['cr_tx'] == 1) & df.cr_scause.isin(value_list4) & (df['cr_frhyem'] == 0)].index)

            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)

        elif x == 4:
            N = len(df.loc[(df['cr_tx'] == 1) & (df['cr_frhyem'] == 4)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            noObvious = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 4)].index)

            noObviouspercent = 0
            if noObvious not in [0]:
                noObviouspercent = round((noObvious / noObvious_total), 4)
            noObviousList.append(noObviouspercent)

            TraumaObvious = len(
                df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 0) & (df['cr_frhyem'] == 4)].index)

            TraumaObviouspercent = 0
            if TraumaObvious not in [0]:
                TraumaObviouspercent = round((TraumaObvious / TraumaObvious_total), 4)
            TraumaObviousList.append(TraumaObviouspercent)

            Unknown = len(
                df.loc[(df['cr_tx'] == 1) & df.cr_scause.isin(value_list4) & (df['cr_frhyem'] == 4)].index)

            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)

        elif x ==5:
            N = len(df.loc[(df['cr_tx'] == 1) & (df['cr_frhyem'] == 6)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            noObvious = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 6)].index)

            noObviouspercent = 0
            if noObvious not in [0]:
                noObviouspercent = round((noObvious / noObvious_total), 4)
            noObviousList.append(noObviouspercent)

            TraumaObvious = len(
                df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 0) & (df['cr_frhyem'] == 6)].index)

            TraumaObviouspercent = 0
            if TraumaObvious not in [0]:
                TraumaObviouspercent = round((TraumaObvious / TraumaObvious_total), 4)
            TraumaObviousList.append(TraumaObviouspercent)

            Unknown = len(
                df.loc[(df['cr_tx'] == 1) & df.cr_scause.isin(value_list4) & (df['cr_frhyem'] == 6)].index)

            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)

    AdultUpsteinWorksheet = AdultUstenworkbook['Sheet2']

    AdultUpsteinWorksheet['B46'] = 'All (' + str(Ntotal) + ')'
    AdultUpsteinWorksheet['B47'] = NList[0]
    AdultUpsteinWorksheet['B48'] = NList[1]
    AdultUpsteinWorksheet['B49'] = NList[2]
    AdultUpsteinWorksheet['B50'] = NList[3]
    AdultUpsteinWorksheet['B51'] = NList[4]

    AdultUpsteinWorksheet['C46'] = 'Presumed cardiac etiology (' + str(noObvious_total) + ')'
    AdultUpsteinWorksheet['C47'] = noObviousList[0]
    AdultUpsteinWorksheet['C48'] = noObviousList[1]
    AdultUpsteinWorksheet['C49'] = noObviousList[2]
    AdultUpsteinWorksheet['C50'] = noObviousList[3]
    AdultUpsteinWorksheet['C51'] = noObviousList[4]

    AdultUpsteinWorksheet['D46'] = 'Obvious cause (' + str(TraumaObvious_total) + ')'
    AdultUpsteinWorksheet['D47'] = TraumaObviousList[0]
    AdultUpsteinWorksheet['D48'] = TraumaObviousList[1]
    AdultUpsteinWorksheet['D49'] = TraumaObviousList[2]
    AdultUpsteinWorksheet['D50'] = TraumaObviousList[3]
    AdultUpsteinWorksheet['D51'] = TraumaObviousList[4]

    AdultUpsteinWorksheet['E46'] = 'Unknown (' + str(Unknown_total) + ')'
    AdultUpsteinWorksheet['E47'] = UnknownList[0]
    AdultUpsteinWorksheet['E48'] = UnknownList[1]
    AdultUpsteinWorksheet['E49'] = UnknownList[2]
    AdultUpsteinWorksheet['E50'] = UnknownList[3]
    AdultUpsteinWorksheet['E51'] = UnknownList[4]

    AdultUstenworkbook.save("C:/Desktop/Charts/AdultUsteinGraphs/CanRoc/Adult_Utstein_" + site + ".xlsx")

    print('Completed - Fig 6')


    # 7th graph Witnessed arrest/bystander CPR, breakdown by outcome (Presumed cardiac etiology)
    value_list5 = [2, 3, 5, '2', '3', '5']
    value_list6 = [18, '', '18']
    value_list4 = ['', None]

    Ntotal = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0)].index)
    UnWitnessednoBystander_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 0)& (df['cr_cpratt'] == 0)].index)
    UnWitnessdBystander_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 0)& (df['cr_cpratt'] == 1)].index)
    BysWitnessednoBystander_total =  len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1)& (df['cr_cpratt'] == 0)].index)
    BysWitnessedBystander_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1)& (df['cr_cpratt'] == 1)].index)
    WitnessedEMS_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 2)].index)
    Unknown_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_witbys.isin(value_list4) & df.cr_cpratt.isin(value_list4)].index)
    BysCPR_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_cpratt'] == 1)].index)
    PADapplied_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_aedapp'] == 1)].index)


    NList = []
    UnWitnessednoBystanderList = []
    UnWitnessdBystanderList = []
    BysWitnessednoBystanderList = []
    BysWitnessedBystanderList = []
    WitnessedEMSList = []
    UnknownList = []
    BysCPRList = []
    PADappliedList = []

    for x in range(1, 6):
        if x == 1:
            N = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_rosc'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            UnWitnessednoBystander = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 0)& (df['cr_cpratt'] == 0) & (df['cr_rosc'] == 1)].index)
            UnWitnessednoBystanderpercent = 0
            if UnWitnessednoBystander not in [0]:
                UnWitnessednoBystanderpercent = round((UnWitnessednoBystander / UnWitnessednoBystander_total), 4)
            UnWitnessednoBystanderList.append(UnWitnessednoBystanderpercent)

            UnWitnessdBystander = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 0) & (df['cr_cpratt'] == 1) & (df['cr_rosc'] == 1)].index)
            UnWitnessdBystanderpercent = 0
            if UnWitnessdBystander not in [0]:
                UnWitnessdBystanderpercent = round((UnWitnessdBystander / UnWitnessdBystander_total), 4)
            UnWitnessdBystanderList.append(UnWitnessdBystanderpercent)

            BysWitnessednoBystander = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1)& (df['cr_cpratt'] == 0) & (df['cr_rosc'] == 1)].index)
            BysWitnessednoBystanderpercent = 0
            if BysWitnessednoBystander not in [0]:
                BysWitnessednoBystanderpercent = round((BysWitnessednoBystander / BysWitnessednoBystander_total), 4)
            BysWitnessednoBystanderList.append(BysWitnessednoBystanderpercent)

            BysWitnessedBystander =  len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1)& (df['cr_cpratt'] == 1)  & (df['cr_rosc'] == 1)].index)
            BysWitnessedBystanderpercent = 0
            if BysWitnessedBystander not in [0]:
                BysWitnessedBystanderpercent = round((BysWitnessedBystander / BysWitnessedBystander_total), 4)
            BysWitnessedBystanderList.append(BysWitnessedBystanderpercent)

            WitnessedEMS = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 2)  & (df['cr_rosc'] == 1)].index)
            WitnessedEMSpercent = 0
            if WitnessedEMS not in [0]:
                WitnessedEMSpercent = round((WitnessedEMS/ WitnessedEMS_total), 4)
            WitnessedEMSList.append(WitnessedEMSpercent)

            Unknown = len(df.loc[len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_witbys.isin(value_list4) & df.cr_cpratt.isin(value_list4)].index) & (df['cr_rosc'] == 1)].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)

            BysCPR = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_cpratt'] == 1) &(df['cr_rosc'] == 1)].index)
            BysCPRpercent = 0
            if BysCPR not in [0]:
                BysCPRpercent = round((BysCPR / BysCPR_total), 4)
            BysCPRList.append(BysCPRpercent)

            PADapplied = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_aedapp'] == 1)  &(df['cr_rosc'] == 1)].index)

            PADappliedpercent = 0
            if PADapplied not in [0]:
                PADappliedpercent = round((PADappliedpercent / PADapplied_total), 4)
            PADappliedList.append(PADappliedpercent)

        if x == 2:
            N = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_prosc'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            UnWitnessednoBystander = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 0) & (
                        df['cr_cpratt'] == 0) & (df['cr_prosc'] == 1)].index)
            UnWitnessednoBystanderpercent = 0
            if UnWitnessednoBystander not in [0]:
                UnWitnessednoBystanderpercent = round((UnWitnessednoBystander / UnWitnessednoBystander_total), 4)
            UnWitnessednoBystanderList.append(UnWitnessednoBystanderpercent)

            UnWitnessdBystander = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 0) & (
                        df['cr_cpratt'] == 1) & (df['cr_prosc'] == 1)].index)
            UnWitnessdBystanderpercent = 0
            if UnWitnessdBystander not in [0]:
                UnWitnessdBystanderpercent = round((UnWitnessdBystander / UnWitnessdBystander_total), 4)
            UnWitnessdBystanderList.append(UnWitnessdBystanderpercent)

            BysWitnessednoBystander = len(df.loc[
                                              (df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1) & (
                                                          df['cr_cpratt'] == 0) & (df['cr_prosc'] == 1)].index)
            BysWitnessednoBystanderpercent = 0
            if BysWitnessednoBystander not in [0]:
                BysWitnessednoBystanderpercent = round((BysWitnessednoBystander / BysWitnessednoBystander_total), 4)
            BysWitnessednoBystanderList.append(BysWitnessednoBystanderpercent)

            BysWitnessedBystander = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1) & (
                        df['cr_cpratt'] == 1) & (df['cr_prosc'] == 1)].index)
            BysWitnessedBystanderpercent = 0
            if BysWitnessedBystander not in [0]:
                BysWitnessedBystanderpercent = round((BysWitnessedBystander / BysWitnessedBystander_total), 4)
            BysWitnessedBystanderList.append(BysWitnessedBystanderpercent)

            WitnessedEMS = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 2) & (df['cr_prosc'] == 1)].index)
            WitnessedEMSpercent = 0
            if WitnessedEMS not in [0]:
                WitnessedEMSpercent = round((WitnessedEMS / WitnessedEMS_total), 4)
            WitnessedEMSList.append(WitnessedEMSpercent)

            Unknown = len(
                df.loc[len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_witbys.isin(value_list4) & df.cr_cpratt.isin(value_list4)].index) & (df['cr_prosc'] == 1)].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)

            BysCPR = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_cpratt'] == 1) & (df['cr_prosc'] == 1)].index)
            BysCPRpercent = 0
            if BysCPR not in [0]:
                BysCPRpercent = round((BysCPR / BysCPR_total), 4)
            BysCPRList.append(BysCPRpercent)

            PADapplied = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_aedapp'] == 1) &  (df['cr_prosc'] == 1)].index)

            PADappliedpercent = 0
            if PADapplied not in [0]:
                PADappliedpercent = round((PADappliedpercent / PADapplied_total), 4)
            PADappliedList.append(PADappliedpercent)

        if x == 3:
            N = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_prosc'] == 0)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            UnWitnessednoBystander = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 0) & (
                        df['cr_cpratt'] == 0) & (df['cr_prosc'] == 0)].index)
            UnWitnessednoBystanderpercent = 0
            if UnWitnessednoBystander not in [0]:
                UnWitnessednoBystanderpercent = round((UnWitnessednoBystander / UnWitnessednoBystander_total), 4)
            UnWitnessednoBystanderList.append(UnWitnessednoBystanderpercent)

            UnWitnessdBystander = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 0) & (
                        df['cr_cpratt'] == 1) & (df['cr_prosc'] == 0)].index)
            UnWitnessdBystanderpercent = 0
            if UnWitnessdBystander not in [0]:
                UnWitnessdBystanderpercent = round((UnWitnessdBystander / UnWitnessdBystander_total), 4)
            UnWitnessdBystanderList.append(UnWitnessdBystanderpercent)

            BysWitnessednoBystander = len(df.loc[
                                              (df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1) & (
                                                          df['cr_cpratt'] == 0) & (df['cr_prosc'] == 0)].index)
            BysWitnessednoBystanderpercent = 0
            if BysWitnessednoBystander not in [0]:
                BysWitnessednoBystanderpercent = round((BysWitnessednoBystander / BysWitnessednoBystander_total), 4)
            BysWitnessednoBystanderList.append(BysWitnessednoBystanderpercent)

            BysWitnessedBystander = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1) & (
                        df['cr_cpratt'] == 1) & (df['cr_prosc'] == 0)].index)
            BysWitnessedBystanderpercent = 0
            if BysWitnessedBystander not in [0]:
                BysWitnessedBystanderpercent = round((BysWitnessedBystander / BysWitnessedBystander_total), 4)
            BysWitnessedBystanderList.append(BysWitnessedBystanderpercent)

            WitnessedEMS = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 2) & (df['cr_prosc'] == 0)].index)
            WitnessedEMSpercent = 0
            if WitnessedEMS not in [0]:
                WitnessedEMSpercent = round((WitnessedEMS / WitnessedEMS_total), 4)
            WitnessedEMSList.append(WitnessedEMSpercent)

            Unknown = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_witbys.isin(value_list4) & df.cr_cpratt.isin(value_list4) & (df['cr_prosc'] == 0)].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)

            BysCPR = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_cpratt'] == 1) & (df['cr_prosc'] == 0)].index)
            BysCPRpercent = 0
            if BysCPR not in [0]:
                BysCPRpercent = round((BysCPR / BysCPR_total), 4)
            BysCPRList.append(BysCPRpercent)

            PADapplied = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_aedapp'] == 1) &  (df['cr_prosc'] == 0)].index)

            PADappliedpercent = 0
            if PADapplied not in [0]:
                PADappliedpercent = round((PADappliedpercent / PADapplied_total), 4)
            PADappliedList.append(PADappliedpercent)

        if x == 4:
            N = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_surv'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            UnWitnessednoBystander = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 0) & (
                        df['cr_cpratt'] == 0) & (df['cr_surv'] == 1)].index)
            UnWitnessednoBystanderpercent = 0
            if UnWitnessednoBystander not in [0]:
                UnWitnessednoBystanderpercent = round((UnWitnessednoBystander / UnWitnessednoBystander_total), 4)
            UnWitnessednoBystanderList.append(UnWitnessednoBystanderpercent)

            UnWitnessdBystander = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 0) & (
                        df['cr_cpratt'] == 1) & (df['cr_surv'] == 1)].index)
            UnWitnessdBystanderpercent = 0
            if UnWitnessdBystander not in [0]:
                UnWitnessdBystanderpercent = round((UnWitnessdBystander / UnWitnessdBystander_total), 4)
            UnWitnessdBystanderList.append(UnWitnessdBystanderpercent)

            BysWitnessednoBystander = len(df.loc[
                                              (df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1) & (
                                                          df['cr_cpratt'] == 0) & (df['cr_surv'] == 1)].index)
            BysWitnessednoBystanderpercent = 0
            if BysWitnessednoBystander not in [0]:
                BysWitnessednoBystanderpercent = round((BysWitnessednoBystander / BysWitnessednoBystander_total), 4)
            BysWitnessednoBystanderList.append(BysWitnessednoBystanderpercent)

            BysWitnessedBystander = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1) & (
                        df['cr_cpratt'] == 1) & (df['cr_surv'] == 1)].index)
            BysWitnessedBystanderpercent = 0
            if BysWitnessedBystander not in [0]:
                BysWitnessedBystanderpercent = round((BysWitnessedBystander / BysWitnessedBystander_total), 4)
            BysWitnessedBystanderList.append(BysWitnessedBystanderpercent)

            WitnessedEMS = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 2) & (df['cr_surv'] == 1)].index)
            WitnessedEMSpercent = 0
            if WitnessedEMS not in [0]:
                WitnessedEMSpercent = round((WitnessedEMS / WitnessedEMS_total), 4)
            WitnessedEMSList.append(WitnessedEMSpercent)

            Unknown = len(
                df.loc[len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_witbys.isin(value_list4) & df.cr_cpratt.isin(value_list4)].index) & (df['cr_surv'] == 1)].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)

            BysCPR = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_cpratt'] == 1) & (df['cr_surv'] == 1)].index)
            BysCPRpercent = 0
            if BysCPR not in [0]:
                BysCPRpercent = round((BysCPR / BysCPR_total), 4)
            BysCPRList.append(BysCPRpercent)

            PADapplied = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_aedapp'] == 1) &  (df['cr_surv'] == 1)].index)

            PADappliedpercent = 0
            if PADapplied not in [0]:
                PADappliedpercent = round((PADappliedpercent / PADapplied_total), 4)
            PADappliedList.append(PADappliedpercent)
        if x == 5:
            N = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_surv'] == 2)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            UnWitnessednoBystander = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 0) & (
                        df['cr_cpratt'] == 0) & (df['cr_surv'] == 2)].index)
            UnWitnessednoBystanderpercent = 0
            if UnWitnessednoBystander not in [0]:
                UnWitnessednoBystanderpercent = round((UnWitnessednoBystander / UnWitnessednoBystander_total), 4)
            UnWitnessednoBystanderList.append(UnWitnessednoBystanderpercent)

            UnWitnessdBystander = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 0) & (
                        df['cr_cpratt'] == 1) & (df['cr_surv'] == 2)].index)
            UnWitnessdBystanderpercent = 0
            if UnWitnessdBystander not in [0]:
                UnWitnessdBystanderpercent = round((UnWitnessdBystander / UnWitnessdBystander_total), 4)
            UnWitnessdBystanderList.append(UnWitnessdBystanderpercent)

            BysWitnessednoBystander = len(df.loc[
                                              (df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1) & (
                                                          df['cr_cpratt'] == 0) & (df['cr_surv'] == 2)].index)
            BysWitnessednoBystanderpercent = 0
            if BysWitnessednoBystander not in [0]:
                BysWitnessednoBystanderpercent = round((BysWitnessednoBystander / BysWitnessednoBystander_total), 4)
            BysWitnessednoBystanderList.append(BysWitnessednoBystanderpercent)

            BysWitnessedBystander = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1) & (
                        df['cr_cpratt'] == 1) & (df['cr_surv'] == 2)].index)
            BysWitnessedBystanderpercent = 0
            if BysWitnessedBystander not in [0]:
                BysWitnessedBystanderpercent = round((BysWitnessedBystander / BysWitnessedBystander_total), 4)
            BysWitnessedBystanderList.append(BysWitnessedBystanderpercent)

            WitnessedEMS = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 2) & (df['cr_surv'] == 2)].index)
            WitnessedEMSpercent = 0
            if WitnessedEMS not in [0]:
                WitnessedEMSpercent = round((WitnessedEMS / WitnessedEMS_total), 4)
            WitnessedEMSList.append(WitnessedEMSpercent)

            Unknown = len(
                df.loc[len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_witbys.isin(value_list4) & df.cr_cpratt.isin(value_list4)].index) & (df['cr_surv'] == 2)].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)

            BysCPR = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_cpratt'] == 1) & (df['cr_surv'] == 2)].index)
            BysCPRpercent = 0
            if BysCPR not in [0]:
                BysCPRpercent = round((BysCPR / BysCPR_total), 4)
            BysCPRList.append(BysCPRpercent)

            PADapplied = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_aedapp'] == 1) &  (df['cr_surv'] == 2)].index)

            PADappliedpercent = 0
            if PADapplied not in [0]:
                PADappliedpercent = round((PADappliedpercent / PADapplied_total), 4)
            PADappliedList.append(PADappliedpercent)





    AdultUpsteinWorksheet = AdultUstenworkbook['Sheet2']

    AdultUpsteinWorksheet['B59'] = 'All (' + str(Ntotal) + ')'
    AdultUpsteinWorksheet['B60'] = NList[0]
    AdultUpsteinWorksheet['B61'] = NList[1]
    AdultUpsteinWorksheet['B62'] = NList[2]
    AdultUpsteinWorksheet['B63'] = NList[3]
    AdultUpsteinWorksheet['B64'] = NList[4]

    AdultUpsteinWorksheet['C59'] = 'Unwitnessed, No Bys CPR (' + str(UnWitnessednoBystander_total) + ')'
    AdultUpsteinWorksheet['C60'] = UnWitnessednoBystanderList[0]
    AdultUpsteinWorksheet['C61'] = UnWitnessednoBystanderList[1]
    AdultUpsteinWorksheet['C62'] = UnWitnessednoBystanderList[2]
    AdultUpsteinWorksheet['C63'] = UnWitnessednoBystanderList[3]


    AdultUpsteinWorksheet['D59'] = 'Unwitnessed, Bys CPR (' + str(UnWitnessdBystander_total) + ')'
    AdultUpsteinWorksheet['D60'] = UnWitnessdBystanderList[0]
    AdultUpsteinWorksheet['D61'] = UnWitnessdBystanderList[1]
    AdultUpsteinWorksheet['D62'] = UnWitnessdBystanderList[2]
    AdultUpsteinWorksheet['D63'] = UnWitnessdBystanderList[3]


    AdultUpsteinWorksheet['E59'] = 'Bys Witnessed, No Bys CPR (' + str(BysWitnessednoBystander_total) + ')'
    AdultUpsteinWorksheet['E60'] = BysWitnessednoBystanderList[0]
    AdultUpsteinWorksheet['E61'] = BysWitnessednoBystanderList[1]
    AdultUpsteinWorksheet['E62'] = BysWitnessednoBystanderList[2]
    AdultUpsteinWorksheet['E63'] = BysWitnessednoBystanderList[3]


    AdultUpsteinWorksheet['F59'] = 'Bys Witnessed, Bys CPR (' + str( BysWitnessedBystander_total) + ')'
    AdultUpsteinWorksheet['F60'] = BysWitnessedBystanderList[0]
    AdultUpsteinWorksheet['F61'] = BysWitnessedBystanderList[1]
    AdultUpsteinWorksheet['F62'] = BysWitnessedBystanderList[2]
    AdultUpsteinWorksheet['F63'] = BysWitnessedBystanderList[3]




    AdultUpsteinWorksheet['G59'] = 'Witnessed by EMS/Fire (' + str(WitnessedEMS_total) + ')'
    AdultUpsteinWorksheet['G60'] = WitnessedEMSList[0]
    AdultUpsteinWorksheet['G61'] = WitnessedEMSList[1]
    AdultUpsteinWorksheet['G62'] = WitnessedEMSList[2]
    AdultUpsteinWorksheet['G63'] = WitnessedEMSList[3]


    AdultUpsteinWorksheet['H59'] = 'Unknown (' + str(Unknown_total) + ')'
    AdultUpsteinWorksheet['H60'] = UnknownList[0]
    AdultUpsteinWorksheet['H61'] = UnknownList[1]
    AdultUpsteinWorksheet['H62'] = UnknownList[2]
    AdultUpsteinWorksheet['H63'] = UnknownList[3]


    AdultUpsteinWorksheet['I59'] = 'Bys CPR (' + str(BysCPR_total) + ')'
    AdultUpsteinWorksheet['I60'] = BysCPRList[0]
    AdultUpsteinWorksheet['I61'] = BysCPRList[1]
    AdultUpsteinWorksheet['I62'] = BysCPRList[2]
    AdultUpsteinWorksheet['I63'] = BysCPRList[3]


    AdultUpsteinWorksheet['J59'] = 'PAD applied (' + str(PADapplied_total) + ')'
    AdultUpsteinWorksheet['J60'] = PADappliedList[0]
    AdultUpsteinWorksheet['J61'] = PADappliedList[1]
    AdultUpsteinWorksheet['J62'] = PADappliedList[2]
    AdultUpsteinWorksheet['J63'] = PADappliedList[3]




    AdultUstenworkbook.save("C:/Charts/AdultUsteinGraphs/CanRoc/Adult_Utstein_"+site+".xlsx")
    print('Completed - Fig 7')




    #   8th graph Presenting rhythm, breakdown by outcome (Presumed cardiac etiology)
    value_list5 = [2, 3, 5, '2', '3', '5']
    value_list6 = [0, 1, 4, 6]
    value_list4 = ['', None]

    Ntotal = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0)].index)
    VFVTshockable_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_frhyem.isin(value_list5)].index)
    PEA_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 1)].index)
    Asystole_total =   len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 0)].index)
    NotShockable_total =  len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 4)].index)
    Unknown_total =  len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 6)].index)


    NList = []
    VFVTshockableList = []
    PEAList = []
    AsystoleList = []
    NotShockableList = []
    UnknownList = []



    for x in range(1, 6):
        if x == 1:
            N = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_rosc'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            VFVTshockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_frhyem.isin(value_list5) & (df['cr_rosc'] == 1)].index)
            VFVTshockablepercent = 0
            if VFVTshockable not in [0]:
                VFVTshockablepercent = round((VFVTshockable / VFVTshockable_total), 4)
            VFVTshockableList.append(VFVTshockablepercent)

            PEA = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 1)  & (df['cr_rosc'] == 1)].index)
            PEA_percent = 0
            if PEA not in [0]:
                PEA_percent = round((PEA/ PEA_total), 4)
            PEAList.append(PEA_percent)

            Asystole = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 0) & (df['cr_rosc'] == 1)].index)
            Asystolepercent = 0
            if Asystole not in [0]:
                Asystolepercent = round((Asystole / Asystole_total), 4)
            AsystoleList.append(Asystolepercent)

            NotShockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 4) & (df['cr_rosc'] == 1)].index)
            NotShockablepercent = 0
            if NotShockable not in [0]:
                NotShockablepercent = round((NotShockable / NotShockable_total), 4)
            NotShockableList.append(NotShockablepercent)

            Unknown = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 6) & (df['cr_rosc'] == 1)].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)


        if x == 2:
            N = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_pdisp'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            VFVTshockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_frhyem.isin(value_list5) & (df['cr_pdisp'] == 1)].index)
            VFVTshockablepercent = 0
            if VFVTshockable not in [0]:
                VFVTshockablepercent = round((VFVTshockable / VFVTshockable_total), 4)
            VFVTshockableList.append(VFVTshockablepercent)

            PEA = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 1)  & (df['cr_pdisp'] == 1)].index)
            PEA_percent = 0
            if PEA not in [0]:
                PEA_percent = round((PEA/ PEA_total), 4)
            PEAList.append(PEA_percent)

            Asystole = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 0) & (df['cr_pdisp'] == 1)].index)
            Asystolepercent = 0
            if Asystole not in [0]:
                Asystolepercent = round((Asystole / Asystole_total), 4)
            AsystoleList.append(Asystolepercent)

            NotShockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 4) & (df['cr_pdisp'] == 1)].index)
            NotShockablepercent = 0
            if NotShockable not in [0]:
                NotShockablepercent = round((NotShockable / NotShockable_total), 4)
            NotShockableList.append(NotShockablepercent)

            Unknown = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 6) & (df['cr_pdisp'] == 1)].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)



        if x == 3:
            N = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_surv'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            VFVTshockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_frhyem.isin(value_list5) & (df['cr_surv'] == 1)].index)
            VFVTshockablepercent = 0
            if VFVTshockable not in [0]:
                VFVTshockablepercent = round((VFVTshockable / VFVTshockable_total), 4)
            VFVTshockableList.append(VFVTshockablepercent)

            PEA = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 1)  & (df['cr_surv'] == 1)].index)
            PEA_percent = 0
            if PEA not in [0]:
                PEA_percent = round((PEA/ PEA_total), 4)
            PEAList.append(PEA_percent)

            Asystole = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 0) & (df['cr_surv'] == 1)].index)
            Asystolepercent = 0
            if Asystole not in [0]:
                Asystolepercent = round((Asystole / Asystole_total), 4)
            AsystoleList.append(Asystolepercent)

            NotShockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 4) & (df['cr_surv'] == 1)].index)
            NotShockablepercent = 0
            if NotShockable not in [0]:
                NotShockablepercent = round((NotShockable / NotShockable_total), 4)
            NotShockableList.append(NotShockablepercent)

            Unknown = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 6) & (df['cr_surv'] == 1)].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)


        if x == 4:
            N = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_surv'] == 2)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)





    AdultUpsteinWorksheet = AdultUstenworkbook['Sheet2']

    AdultUpsteinWorksheet['B70'] = 'All (' + str(Ntotal) + ')'
    AdultUpsteinWorksheet['B71'] = NList[0]
    AdultUpsteinWorksheet['B72'] = NList[1]
    AdultUpsteinWorksheet['B73'] = NList[2]
    AdultUpsteinWorksheet['B74'] = NList[3]


    AdultUpsteinWorksheet['C70'] = 'VF/VT or shockable (' + str(VFVTshockable_total) + ')'
    AdultUpsteinWorksheet['C71'] = VFVTshockableList[0]
    AdultUpsteinWorksheet['C72'] = VFVTshockableList[1]
    AdultUpsteinWorksheet['C73'] = VFVTshockableList[2]



    AdultUpsteinWorksheet['D70'] = 'PEA (' + str(PEA_total) + ')'
    AdultUpsteinWorksheet['D71'] = PEAList[0]
    AdultUpsteinWorksheet['D72'] = PEAList[1]
    AdultUpsteinWorksheet['D73'] = PEAList[2]


    AdultUpsteinWorksheet['E70'] = 'Asystole (' + str(Asystole_total) + ')'
    AdultUpsteinWorksheet['E71'] = AsystoleList[0]
    AdultUpsteinWorksheet['E72'] = AsystoleList[1]
    AdultUpsteinWorksheet['E73'] = AsystoleList[2]


    AdultUpsteinWorksheet['F70'] = 'Not Shockable (' + str(NotShockable_total ) + ')'
    AdultUpsteinWorksheet['F71'] = NotShockableList[0]
    AdultUpsteinWorksheet['F72'] = NotShockableList[1]
    AdultUpsteinWorksheet['F73'] = NotShockableList[2]



    AdultUpsteinWorksheet['G70'] = 'Unknown Rhythm (' + str(Unknown_total) + ')'
    AdultUpsteinWorksheet['G71'] = UnknownList[0]


    AdultUstenworkbook.save("C:/Users/mcbarnettr/Desktop/Charts/AdultUsteinGraphs/CanRoc/Adult_Utstein_" + site + ".xlsx")
    print('Completed - Fig 8')

    # 9th Graph   Presenting rhythm, breakdown by outcome (Bystander Witnessed Arrests, Without # Bystander CPR - Presumed cardiac etiology)
    value_list5 = [2, 3, 5, '2', '3', '5']
    value_list6 = [0, 1, 4, 6]
    value_list4 = ['', None]
    value_list7 = [0,1,'0','1']

    Ntotal = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1)].index)
    VFVTshockable_total = len(
        df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0)& (df['cr_witbys'] == 1) & df.cr_frhyem.isin(value_list5)].index)
    PEA_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1)& (df['cr_frhyem'] == 1)].index)
    Asystole_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1)& (df['cr_frhyem'] == 0)].index)
    NotShockable_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1)& (df['cr_frhyem'] == 4)].index)
    Unknown_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1)& (df['cr_frhyem'] == 6)].index)
    NotVFVTshockableshocked_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) &  (df['cr_witbys'] == 1) & (
                df['cr_numshk'] > 0) & df.cr_frhyem.isin(value_list6)].index)

    NList = []
    VFVTshockableList = []
    PEAList = []
    AsystoleList = []
    NotShockableList = []
    UnknownList = []
    NotVFVTshockableshockedList = []


    for x in range(1, 6):
        if x == 1:
            N = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1)& (df['cr_rosc'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            VFVTshockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1)& df.cr_frhyem.isin(value_list5) & (df['cr_rosc'] == 1)].index)
            VFVTshockablepercent = 0
            if VFVTshockable not in [0]:
                VFVTshockablepercent = round((VFVTshockable / VFVTshockable_total), 4)
            VFVTshockableList.append(VFVTshockablepercent)

            PEA = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0)& (df['cr_witbys'] == 1) & (df['cr_frhyem'] == 1)  & (df['cr_rosc'] == 1)].index)
            PEA_percent = 0
            if PEA not in [0]:
                PEA_percent = round((PEA/ PEA_total), 4)
            PEAList.append(PEA_percent)

            Asystole = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0)& (df['cr_witbys'] == 1) & (df['cr_frhyem'] == 0) & (df['cr_rosc'] == 1)].index)
            Asystolepercent = 0
            if Asystole not in [0]:
                Asystolepercent = round((Asystole / Asystole_total), 4)
            AsystoleList.append(Asystolepercent)

            NotShockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0)& (df['cr_witbys'] == 1) & (df['cr_frhyem'] == 4) & (df['cr_rosc'] == 1)].index)
            NotShockablepercent = 0
            if NotShockable not in [0]:
                NotShockablepercent = round((NotShockable / NotShockable_total), 4)
            NotShockableList.append(NotShockablepercent)

            Unknown = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0)& (df['cr_witbys'] == 1) & (df['cr_frhyem'] == 6) & (df['cr_rosc'] == 1)].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)

        if x == 2: # Changed to "Transported to hospital - pdsip = 1
            N = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0)& (df['cr_witbys'] == 1) & (df['cr_pdisp'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            VFVTshockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0)& (df['cr_witbys'] == 1) & df.cr_frhyem.isin(value_list5) & (df['cr_pdisp'] == 1)].index)
            VFVTshockablepercent = 0
            if VFVTshockable not in [0]:
                VFVTshockablepercent = round((VFVTshockable / VFVTshockable_total), 4)
            VFVTshockableList.append(VFVTshockablepercent)

            PEA = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 1)& (df['cr_witbys'] == 1)  & (df['cr_pdisp'] == 1)].index)
            PEA_percent = 0
            if PEA not in [0]:
                PEA_percent = round((PEA/ PEA_total), 4)
            PEAList.append(PEA_percent)

            Asystole = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 0)& (df['cr_witbys'] == 1) & (df['cr_pdisp'] == 1)].index)
            Asystolepercent = 0
            if Asystole not in [0]:
                Asystolepercent = round((Asystole / Asystole_total), 4)
            AsystoleList.append(Asystolepercent)

            NotShockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 4)& (df['cr_witbys'] == 1) & (df['cr_pdisp'] == 1)].index)
            NotShockablepercent = 0
            if NotShockable not in [0]:
                NotShockablepercent = round((NotShockable / NotShockable_total), 4)
            NotShockableList.append(NotShockablepercent)

            Unknown = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 6)& (df['cr_witbys'] == 1) & (df['cr_pdisp'] == 1)].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)



        if x == 3:
            N = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0)& (df['cr_witbys'] == 1) & (df['cr_surv'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            VFVTshockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_frhyem.isin(value_list5)& (df['cr_witbys'] == 1) & (df['cr_surv'] == 1)].index)
            VFVTshockablepercent = 0
            if VFVTshockable not in [0]:
                VFVTshockablepercent = round((VFVTshockable / VFVTshockable_total), 4)
            VFVTshockableList.append(VFVTshockablepercent)

            PEA = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 1) & (df['cr_witbys'] == 1) & (df['cr_surv'] == 1)].index)
            PEA_percent = 0
            if PEA not in [0]:
                PEA_percent = round((PEA/ PEA_total), 4)
            PEAList.append(PEA_percent)

            Asystole = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 0)& (df['cr_witbys'] == 1) & (df['cr_surv'] == 1)].index)
            Asystolepercent = 0
            if Asystole not in [0]:
                Asystolepercent = round((Asystole / Asystole_total), 4)
            AsystoleList.append(Asystolepercent)

            NotShockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 4)& (df['cr_witbys'] == 1) & (df['cr_surv'] == 1)].index)
            NotShockablepercent = 0
            if NotShockable not in [0]:
                NotShockablepercent = round((NotShockable / NotShockable_total), 4)
            NotShockableList.append(NotShockablepercent)

            Unknown = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 6)& (df['cr_witbys'] == 1) & (df['cr_surv'] == 1)].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)


        if x == 4:
            N = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0)& df.cr_witbys.isin(value_list7) & (df['cr_surv'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)





    AdultUpsteinWorksheet = AdultUstenworkbook['Sheet2']

    AdultUpsteinWorksheet['B85'] = 'All (' + str(Ntotal) + ')'
    AdultUpsteinWorksheet['B86'] = NList[0]
    AdultUpsteinWorksheet['B87'] = NList[1]
    AdultUpsteinWorksheet['B88'] = NList[2]
    AdultUpsteinWorksheet['B89'] = NList[3]


    AdultUpsteinWorksheet['C85'] = 'VF/VT or shockable (' + str(VFVTshockable_total) + ')'
    AdultUpsteinWorksheet['C86'] = VFVTshockableList[0]
    AdultUpsteinWorksheet['C87'] = VFVTshockableList[1]
    AdultUpsteinWorksheet['C88'] = VFVTshockableList[2]



    AdultUpsteinWorksheet['D85'] = 'PEA (' + str(PEA_total) + ')'
    AdultUpsteinWorksheet['D86'] = PEAList[0]
    AdultUpsteinWorksheet['D87'] = PEAList[1]
    AdultUpsteinWorksheet['D88'] = PEAList[2]



    AdultUpsteinWorksheet['E85'] = 'Asystole (' + str(Asystole_total) + ')'
    AdultUpsteinWorksheet['E86'] = AsystoleList[0]
    AdultUpsteinWorksheet['E87'] = AsystoleList[1]
    AdultUpsteinWorksheet['E88'] = AsystoleList[2]



    AdultUpsteinWorksheet['F85'] = 'Not Shockable (' + str(NotShockable_total ) + ')'
    AdultUpsteinWorksheet['F86'] = NotShockableList[0]
    AdultUpsteinWorksheet['F87'] = NotShockableList[1]
    AdultUpsteinWorksheet['F88'] = NotShockableList[2]





    AdultUpsteinWorksheet['G85'] = 'Unknown Rhythm (' + str(Unknown_total) + ')'
    AdultUpsteinWorksheet['G86'] = UnknownList[0]


    AdultUstenworkbook.save("C:/Users/mcbarnettr/Desktop/Charts/AdultUsteinGraphs/CanRoc/Adult_Utstein_" + site + ".xlsx")
    print('Completed - Fig 9')



    ##################################################################

    # 10th Graph Copy with cprpratt in denominator   Presenting rhythm, breakdown by outcome (Bystander # Witnessed Arrests - Presumed cardiac etiology)

    value_list5 = [2, 3, 5, '2', '3', '5']
    value_list6 = [0, 1, 4, 6]
    value_list4 = ['', None]

    Ntotal = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1)].index)
    VFVTshockable_total = len(
        df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1) & df.cr_frhyem.isin(
            value_list5)].index)
    PEA_total = len(
        df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1) & (df['cr_frhyem'] == 1)].index)
    Asystole_total = len(
        df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1) & (df['cr_frhyem'] == 0)].index)
    NotShockable_total = len(
        df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1) & (df['cr_frhyem'] == 4)].index)
    Unknown_total = len(
        df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1) & (df['cr_frhyem'] == 6)].index)
    NotVFVTshockableshocked_total = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1) & (
            df['cr_numshk'] > 0) & df.cr_frhyem.isin(value_list6)].index)

    NList = []
    VFVTshockableList = []
    PEAList = []
    AsystoleList = []
    NotShockableList = []
    UnknownList = []
    NotVFVTshockableshockedList = []

    for x in range(1, 6):
        if x == 1:
            N = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1) & (
                        df['cr_rosc'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            VFVTshockable = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (
                        df['cr_witbys'] == 1) & df.cr_frhyem.isin(value_list5) & (df['cr_rosc'] == 1)].index)
            VFVTshockablepercent = 0
            if VFVTshockable not in [0]:
                VFVTshockablepercent = round((VFVTshockable / VFVTshockable_total), 4)
            VFVTshockableList.append(VFVTshockablepercent)

            PEA = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1) & (
                        df['cr_frhyem'] == 1) & (df['cr_rosc'] == 1)].index)
            PEA_percent = 0
            if PEA not in [0]:
                PEA_percent = round((PEA / PEA_total), 4)
            PEAList.append(PEA_percent)

            Asystole = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1) & (
                        df['cr_frhyem'] == 0) & (df['cr_rosc'] == 1)].index)
            Asystolepercent = 0
            if Asystole not in [0]:
                Asystolepercent = round((Asystole / Asystole_total), 4)
            AsystoleList.append(Asystolepercent)

            NotShockable = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1) & (
                        df['cr_frhyem'] == 4) & (df['cr_rosc'] == 1)].index)
            NotShockablepercent = 0
            if NotShockable not in [0]:
                NotShockablepercent = round((NotShockable / NotShockable_total), 4)
            NotShockableList.append(NotShockablepercent)

            Unknown = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1) & (
                        df['cr_frhyem'] == 6) & (df['cr_rosc'] == 1)].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)

            NotVFVTshockableshocked = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (
                        df['cr_witbys'] == 1) & df.cr_frhyem.isin(value_list6) & (df['cr_numshk'] > 0) & (
                                                             df['cr_rosc'] == 1)].index)
            NotVFVTshockableshockedpercent = 0
            if NotVFVTshockableshocked not in [0]:
                NotVFVTshockableshockedpercent = round((NotVFVTshockableshocked / NotVFVTshockableshocked_total), 4)
            NotVFVTshockableshockedList.append(NotVFVTshockableshockedpercent)

        if x == 2:
            N = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1) & (
                        df['cr_pdisp'] == 0)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            VFVTshockable = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (
                        df['cr_witbys'] == 1) & df.cr_frhyem.isin(value_list5) & (df['cr_pdisp'] == 0)].index)
            VFVTshockablepercent = 0
            if VFVTshockable not in [0]:
                VFVTshockablepercent = round((VFVTshockable / VFVTshockable_total), 4)
            VFVTshockableList.append(VFVTshockablepercent)

            PEA = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 1) & (
                        df['cr_witbys'] == 1) & (df['cr_pdisp'] == 0)].index)
            PEA_percent = 0
            if PEA not in [0]:
                PEA_percent = round((PEA / PEA_total), 4)
            PEAList.append(PEA_percent)

            Asystole = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 0) & (
                        df['cr_witbys'] == 1) & (df['cr_pdisp'] == 0)].index)
            Asystolepercent = 0
            if Asystole not in [0]:
                Asystolepercent = round((Asystole / Asystole_total), 4)
            AsystoleList.append(Asystolepercent)

            NotShockable = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 4) & (
                        df['cr_witbys'] == 1) & (df['cr_pdisp'] == 0)].index)
            NotShockablepercent = 0
            if NotShockable not in [0]:
                NotShockablepercent = round((NotShockable / NotShockable_total), 4)
            NotShockableList.append(NotShockablepercent)

            Unknown = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 6) & (
                        df['cr_witbys'] == 1) & (df['cr_pdisp'] == 0)].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)

            NotVFVTshockableshocked = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (
                        df['cr_witbys'] == 1) & df.cr_frhyem.isin(value_list6) & (df['cr_numshk'] > 0) & (
                                                             df['cr_pdisp'] == 0)].index)
            NotVFVTshockableshockedpercent = 0
            if NotVFVTshockableshocked not in [0]:
                NotVFVTshockableshockedpercent = round((NotVFVTshockableshocked / NotVFVTshockableshocked_total), 4)
            NotVFVTshockableshockedList.append(NotVFVTshockableshockedpercent)

        if x == 3:
            N = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1) & (
                        df['cr_surv'] == 0)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            VFVTshockable = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (
                        df['cr_witbys'] == 1) & df.cr_frhyem.isin(value_list5) & (df['cr_surv'] == 0)].index)
            VFVTshockablepercent = 0
            if VFVTshockable not in [0]:
                VFVTshockablepercent = round((VFVTshockable / VFVTshockable_total), 4)
            VFVTshockableList.append(VFVTshockablepercent)

            PEA = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 1) & (
                        df['cr_witbys'] == 1) & (df['cr_surv'] == 0)].index)
            PEA_percent = 0
            if PEA not in [0]:
                PEA_percent = round((PEA / PEA_total), 4)
            PEAList.append(PEA_percent)

            Asystole = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 0) & (
                        df['cr_witbys'] == 1) & (df['cr_surv'] == 0)].index)
            Asystolepercent = 0
            if Asystole not in [0]:
                Asystolepercent = round((Asystole / Asystole_total), 4)
            AsystoleList.append(Asystolepercent)

            NotShockable = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 4) & (
                        df['cr_witbys'] == 1) & (df['cr_surv'] == 0)].index)
            NotShockablepercent = 0
            if NotShockable not in [0]:
                NotShockablepercent = round((NotShockable / NotShockable_total), 4)
            NotShockableList.append(NotShockablepercent)

            Unknown = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 6) & (
                        df['cr_witbys'] == 1) & (df['cr_surv'] == 0)].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)

            NotVFVTshockableshocked = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_frhyem.isin(
                value_list6) & (df['cr_witbys'] == 1) & (df['cr_numshk'] > 0) & (df['cr_surv'] == 0)].index)
            NotVFVTshockableshockedpercent = 0
            if NotVFVTshockableshocked not in [0]:
                NotVFVTshockableshockedpercent = round((NotVFVTshockableshocked / NotVFVTshockableshocked_total), 4)
            NotVFVTshockableshockedList.append(NotVFVTshockableshockedpercent)

        if x == 4:
            N = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1) & (
                        df['cr_surv'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            VFVTshockable = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_frhyem.isin(value_list5) & (
                        df['cr_witbys'] == 1) & (df['cr_surv'] == 1)].index)
            VFVTshockablepercent = 0
            if VFVTshockable not in [0]:
                VFVTshockablepercent = round((VFVTshockable / VFVTshockable_total), 4)
            VFVTshockableList.append(VFVTshockablepercent)

            PEA = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 1) & (
                        df['cr_witbys'] == 1) & (df['cr_surv'] == 1)].index)
            PEA_percent = 0
            if PEA not in [0]:
                PEA_percent = round((PEA / PEA_total), 4)
            PEAList.append(PEA_percent)

            Asystole = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 0) & (
                        df['cr_witbys'] == 1) & (df['cr_surv'] == 1)].index)
            Asystolepercent = 0
            if Asystole not in [0]:
                Asystolepercent = round((Asystole / Asystole_total), 4)
            AsystoleList.append(Asystolepercent)

            NotShockable = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 4) & (
                        df['cr_witbys'] == 1) & (df['cr_surv'] == 1)].index)
            NotShockablepercent = 0
            if NotShockable not in [0]:
                NotShockablepercent = round((NotShockable / NotShockable_total), 4)
            NotShockableList.append(NotShockablepercent)

            Unknown = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 6) & (
                        df['cr_witbys'] == 1) & (df['cr_surv'] == 1)].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)

            NotVFVTshockableshocked = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_frhyem.isin(
                value_list6) & (df['cr_witbys'] == 1) & (df['cr_numshk'] > 0) & (df['cr_surv'] == 1)].index)
            NotVFVTshockableshockedpercent = 0
            if NotVFVTshockableshocked not in [0]:
                NotVFVTshockableshockedpercent = round((NotVFVTshockableshocked / NotVFVTshockableshocked_total), 4)
            NotVFVTshockableshockedList.append(NotVFVTshockableshockedpercent)

        if x == 5:
            N = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_witbys'] == 1) & (
                        df['cr_surv'] == 2)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            VFVTshockable = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_frhyem.isin(value_list5) & (
                        df['cr_witbys'] == 1) & (df['cr_surv'] == 2)].index)
            VFVTshockablepercent = 0
            if VFVTshockable not in [0]:
                VFVTshockablepercent = round((VFVTshockable / VFVTshockable_total), 4)
            VFVTshockableList.append(VFVTshockablepercent)

            PEA = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 1) & (
                        df['cr_witbys'] == 1) & (df['cr_surv'] == 2)].index)
            PEA_percent = 0
            if PEA not in [0]:
                PEA_percent = round((PEA / PEA_total), 4)
            PEAList.append(PEA_percent)

            Asystole = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 0) & (
                        df['cr_witbys'] == 1) & (df['cr_surv'] == 2)].index)
            Asystolepercent = 0
            if Asystole not in [0]:
                Asystolepercent = round((Asystole / Asystole_total), 4)
            AsystoleList.append(Asystolepercent)

            NotShockable = len(df.loc[(df['cr_cpratt'] == 1)& (df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 4) & (
                        df['cr_witbys'] == 1) & (df['cr_surv'] == 2)].index)
            NotShockablepercent = 0
            if NotShockable not in [0]:
                NotShockablepercent = round((NotShockable / NotShockable_total), 4)
            NotShockableList.append(NotShockablepercent)

            Unknown = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 6) & (
                        df['cr_witbys'] == 1) & (df['cr_surv'] == 2)].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)

            NotVFVTshockableshocked = len(df.loc[(df['cr_cpratt'] == 1)&(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_frhyem.isin(
                value_list6) & (df['cr_witbys'] == 1) & (df['cr_numshk'] > 0) & (df['cr_surv'] == 2)].index)
            NotVFVTshockableshockedpercent = 0
            if NotVFVTshockableshocked not in [0]:
                NotVFVTshockableshockedpercent = round((NotVFVTshockableshocked / NotVFVTshockableshocked_total), 4)
            NotVFVTshockableshockedList.append(NotVFVTshockableshockedpercent)

    AdultUpsteinWorksheet = AdultUstenworkbook['Sheet2']

    AdultUpsteinWorksheet['B92'] = 'All (' + str(Ntotal) + ')'
    AdultUpsteinWorksheet['B93'] = NList[0]
    AdultUpsteinWorksheet['B94'] = NList[1]
    AdultUpsteinWorksheet['B95'] = NList[2]
    AdultUpsteinWorksheet['B96'] = NList[3]
    AdultUpsteinWorksheet['B97'] = NList[4]

    AdultUpsteinWorksheet['C92'] = 'VF/VT or shockable (' + str(VFVTshockable_total) + ')'
    AdultUpsteinWorksheet['C93'] = VFVTshockableList[0]
    AdultUpsteinWorksheet['C94'] = VFVTshockableList[1]
    AdultUpsteinWorksheet['C95'] = VFVTshockableList[2]
    AdultUpsteinWorksheet['C96'] = VFVTshockableList[3]
    AdultUpsteinWorksheet['C97'] = VFVTshockableList[4]

    AdultUpsteinWorksheet['D92'] = 'PEA (' + str(PEA_total) + ')'
    AdultUpsteinWorksheet['D93'] = PEAList[0]
    AdultUpsteinWorksheet['D94'] = PEAList[1]
    AdultUpsteinWorksheet['D95'] = PEAList[2]
    AdultUpsteinWorksheet['D96'] = PEAList[3]
    AdultUpsteinWorksheet['D97'] = PEAList[4]

    AdultUpsteinWorksheet['E92'] = 'Asystole (' + str(Asystole_total) + ')'
    AdultUpsteinWorksheet['E93'] = AsystoleList[0]
    AdultUpsteinWorksheet['E94'] = AsystoleList[1]
    AdultUpsteinWorksheet['E95'] = AsystoleList[2]
    AdultUpsteinWorksheet['E96'] = AsystoleList[3]
    AdultUpsteinWorksheet['E97'] = AsystoleList[4]

    AdultUpsteinWorksheet['F92'] = 'Not Shockable (' + str(NotShockable_total) + ')'
    AdultUpsteinWorksheet['F93'] = NotShockableList[0]
    AdultUpsteinWorksheet['F94'] = NotShockableList[1]
    AdultUpsteinWorksheet['F95'] = NotShockableList[2]
    AdultUpsteinWorksheet['F96'] = NotShockableList[3]
    AdultUpsteinWorksheet['F97'] = NotShockableList[4]

    AdultUpsteinWorksheet['G92'] = 'Unknown Rhythm (' + str(Unknown_total) + ')'
    AdultUpsteinWorksheet['G93'] = UnknownList[0]




    AdultUstenworkbook.save("C:/Users/mcbarnettr/Desktop/Charts/AdultUsteinGraphs/CanRoc/Adult_Utstein_" + site + ".xlsx")
    print('Completed - 10')


    ###################################################################



    # 11th   Presenting rhythm, breakdown by outcome (Not  # Witnessed by EMS - Presumed cardiac etiology)
    value_list5 = [2, 3, 5, '2', '3', '5']
    value_list6 = [0, 1, 4, 6]
    value_list7 = [0,1]
    value_list4 = ['', None]

    Ntotal = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_witbys.isin(value_list7)].index)
    VFVTshockable_total = len(
        df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_witbys.isin(value_list7) & df.cr_frhyem.isin(
            value_list5)].index)
    PEA_total = len(
        df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_witbys.isin(value_list7) & (df['cr_frhyem'] == 1)].index)
    Asystole_total = len(
        df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_witbys.isin(value_list7) & (df['cr_frhyem'] == 0)].index)
    NotShockable_total = len(
        df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_witbys.isin(value_list7) & (df['cr_frhyem'] == 4)].index)
    Unknown_total = len(
        df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_witbys.isin(value_list7) & (df['cr_frhyem'] == 6)].index)

    NList = []
    VFVTshockableList = []
    PEAList = []
    AsystoleList = []
    NotShockableList = []
    UnknownList = []
    NotVFVTshockableshockedList = []

    for x in range(1, 6):
        if x == 1:
            N = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_witbys.isin(value_list7) & (
                        df['cr_rosc'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            VFVTshockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_witbys.isin(value_list7) & df.cr_frhyem.isin(value_list5) & (df['cr_rosc'] == 1)].index)
            VFVTshockablepercent = 0
            if VFVTshockable not in [0]:
                VFVTshockablepercent = round((VFVTshockable / VFVTshockable_total), 4)
            VFVTshockableList.append(VFVTshockablepercent)

            PEA = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_witbys.isin(value_list7) & (
                        df['cr_frhyem'] == 1) & (df['cr_rosc'] == 1)].index)
            PEA_percent = 0
            if PEA not in [0]:
                PEA_percent = round((PEA / PEA_total), 4)
            PEAList.append(PEA_percent)

            Asystole = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_witbys.isin(value_list7) & (
                        df['cr_frhyem'] == 0) & (df['cr_rosc'] == 1)].index)
            Asystolepercent = 0
            if Asystole not in [0]:
                Asystolepercent = round((Asystole / Asystole_total), 4)
            AsystoleList.append(Asystolepercent)

            NotShockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_witbys.isin(value_list7) & (
                        df['cr_frhyem'] == 4) & (df['cr_rosc'] == 1)].index)
            NotShockablepercent = 0
            if NotShockable not in [0]:
                NotShockablepercent = round((NotShockable / NotShockable_total), 4)
            NotShockableList.append(NotShockablepercent)

            Unknown = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_witbys.isin(value_list7) & (
                        df['cr_frhyem'] == 6) & (df['cr_rosc'] == 1)].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)


        if x == 2:
            N = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_witbys.isin(value_list7) & (
                        df['cr_pdisp'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            VFVTshockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_witbys.isin(value_list7) & df.cr_frhyem.isin(value_list5) & (df['cr_pdisp'] == 1)].index)
            VFVTshockablepercent = 0
            if VFVTshockable not in [0]:
                VFVTshockablepercent = round((VFVTshockable / VFVTshockable_total), 4)
            VFVTshockableList.append(VFVTshockablepercent)

            PEA = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 1) & df.cr_witbys.isin(value_list7) & (df['cr_pdisp'] == 1)].index)
            PEA_percent = 0
            if PEA not in [0]:
                PEA_percent = round((PEA / PEA_total), 4)
            PEAList.append(PEA_percent)

            Asystole = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 0) & df.cr_witbys.isin(value_list7) & (df['cr_pdisp'] == 1)].index)
            Asystolepercent = 0
            if Asystole not in [0]:
                Asystolepercent = round((Asystole / Asystole_total), 4)
            AsystoleList.append(Asystolepercent)

            NotShockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 4) & df.cr_witbys.isin(value_list7) & (df['cr_pdisp'] == 1)].index)
            NotShockablepercent = 0
            if NotShockable not in [0]:
                NotShockablepercent = round((NotShockable / NotShockable_total), 4)
            NotShockableList.append(NotShockablepercent)

            Unknown = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 6) & df.cr_witbys.isin(value_list7) & (df['cr_pdisp'] == 1)].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)

            NotVFVTshockableshocked = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_witbys.isin(value_list7) & df.cr_frhyem.isin(value_list6) & (df['cr_numshk'] > 0) & (
                                                             df['cr_pdisp'] == 1)].index)



        if x == 3:
            N = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_witbys.isin(value_list7) & (
                        df['cr_surv'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            VFVTshockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_frhyem.isin(value_list5) & df.cr_witbys.isin(value_list7) & (df['cr_surv'] == 1)].index)
            VFVTshockablepercent = 0
            if VFVTshockable not in [0]:
                VFVTshockablepercent = round((VFVTshockable / VFVTshockable_total), 4)
            VFVTshockableList.append(VFVTshockablepercent)

            PEA = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_witbys.isin(value_list7) & (
                        df['cr_witbys'] == 1) & (df['cr_surv'] == 1)].index)
            PEA_percent = 0
            if PEA not in [0]:
                PEA_percent = round((PEA / PEA_total), 4)
            PEAList.append(PEA_percent)

            Asystole = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 0) & df.cr_witbys.isin(value_list7) & (df['cr_surv'] == 1)].index)
            Asystolepercent = 0
            if Asystole not in [0]:
                Asystolepercent = round((Asystole / Asystole_total), 4)
            AsystoleList.append(Asystolepercent)

            NotShockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 4) & df.cr_witbys.isin(value_list7) & (df['cr_surv'] == 1)].index)
            NotShockablepercent = 0
            if NotShockable not in [0]:
                NotShockablepercent = round((NotShockable / NotShockable_total), 4)
            NotShockableList.append(NotShockablepercent)

            Unknown = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 6) & df.cr_witbys.isin(value_list7) & (df['cr_surv'] == 1)].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)


        if x == 4:
            N = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_witbys.isin(value_list7) & (
                        df['cr_surv'] == 2)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            VFVTshockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & df.cr_frhyem.isin(value_list5) & df.cr_witbys.isin(value_list7) & (df['cr_surv'] == 2)].index)
            VFVTshockablepercent = 0
            if VFVTshockable not in [0]:
                VFVTshockablepercent = round((VFVTshockable / VFVTshockable_total), 4)
            VFVTshockableList.append(VFVTshockablepercent)

            PEA = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 1) & df.cr_witbys.isin(value_list7) & (df['cr_surv'] == 2)].index)
            PEA_percent = 0
            if PEA not in [0]:
                PEA_percent = round((PEA / PEA_total), 4)
            PEAList.append(PEA_percent)

            Asystole = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 0) & df.cr_witbys.isin(value_list7) & (df['cr_surv'] == 2)].index)
            Asystolepercent = 0
            if Asystole not in [0]:
                Asystolepercent = round((Asystole / Asystole_total), 4)
            AsystoleList.append(Asystolepercent)

            NotShockable = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 4) & df.cr_witbys.isin(value_list7) & (df['cr_surv'] == 2)].index)
            NotShockablepercent = 0
            if NotShockable not in [0]:
                NotShockablepercent = round((NotShockable / NotShockable_total), 4)
            NotShockableList.append(NotShockablepercent)

            Unknown = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0) & (df['cr_frhyem'] == 6) & df.cr_witbys.isin(value_list7) & (df['cr_surv'] == 2)].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)

    AdultUpsteinWorksheet = AdultUstenworkbook['Sheet2']

    AdultUpsteinWorksheet['B99'] = 'All (' + str(Ntotal) + ')'
    AdultUpsteinWorksheet['B100'] = NList[0]
    AdultUpsteinWorksheet['B101'] = NList[1]
    AdultUpsteinWorksheet['B102'] = NList[2]
    AdultUpsteinWorksheet['B103'] = NList[3]


    AdultUpsteinWorksheet['C99'] = 'VF/VT or shockable (' + str(VFVTshockable_total) + ')'
    AdultUpsteinWorksheet['C100'] = VFVTshockableList[0]
    AdultUpsteinWorksheet['C101'] = VFVTshockableList[1]
    AdultUpsteinWorksheet['C102'] = VFVTshockableList[2]
    AdultUpsteinWorksheet['C103'] = VFVTshockableList[3]


    AdultUpsteinWorksheet['D99'] = 'PEA (' + str(PEA_total) + ')'
    AdultUpsteinWorksheet['D100'] = PEAList[0]
    AdultUpsteinWorksheet['D101'] = PEAList[1]
    AdultUpsteinWorksheet['D102'] = PEAList[2]
    AdultUpsteinWorksheet['D103'] = PEAList[3]


    AdultUpsteinWorksheet['E99'] = 'Asystole (' + str(Asystole_total) + ')'
    AdultUpsteinWorksheet['E100'] = AsystoleList[0]
    AdultUpsteinWorksheet['E101'] = AsystoleList[1]
    AdultUpsteinWorksheet['E102'] = AsystoleList[2]
    AdultUpsteinWorksheet['E103'] = AsystoleList[3]


    AdultUpsteinWorksheet['F99'] = 'Not Shockable (' + str(NotShockable_total) + ')'
    AdultUpsteinWorksheet['F100'] = NotShockableList[0]
    AdultUpsteinWorksheet['F101'] = NotShockableList[1]
    AdultUpsteinWorksheet['F102'] = NotShockableList[2]
    AdultUpsteinWorksheet['F103'] = NotShockableList[3]


    AdultUpsteinWorksheet['G99'] = 'Unknown Rhythm (' + str(Unknown_total) + ')'
    AdultUpsteinWorksheet['G100'] = UnknownList[0]



    AdultUstenworkbook.save("C:/AdultUsteinGraphs/CanRoc/Adult_Utstein_" + site + ".xlsx")
    print('Completed - Fig 11')



    # 12th   Etiology, breakdown by outcome
    value_list5 = [2, 3, 5, '2', '3', '5']
    value_list6 = [0, 1, 4, 6]
    value_list7 = [0, 1]
    value_list8 = [18, '',None]


    Ntotal = len(df.loc[(df['cr_tx'] == 1)].index)
    NoObviouscause_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] == 0)].index)
    Trauma_total = len(df.loc[(df['cr_tx'] == 1) & (df['cr_scause'] != 0)].index)
    Unknown_total = len(df.loc[(df['cr_tx'] == 1) & df.cr_scause.isin(value_list8)].index)


    NList = []
    NoObviouscauseList = []
    TraumaList = []
    UnknownList = []

    for x in range(1, 6):
        if x == 1:
            N  = len(df.loc[(df['cr_tx'] == 1) & (df['cr_rosc'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            NoObviouscause =  len(df.loc[(df['cr_tx'] == 1) & (df['cr_rosc'] == 1) & (df['cr_scause'] == 0)].index )
            NoObviouscausepercent = 0
            if NoObviouscause not in [0]:
                NoObviouscausepercent = round((NoObviouscause / NoObviouscause_total), 4)
            NoObviouscauseList.append(NoObviouscausepercent)

            Trauma = len(df.loc[(df['cr_tx'] == 1) & (df['cr_rosc'] == 1) &(df['cr_scause'] != 0)].index)
            Trauma_percent = 0
            if Trauma not in [0]:
                Trauma_percent = round((Trauma / Trauma_total), 4)
            TraumaList.append(Trauma_percent)

            Unknown = len(df.loc[(df['cr_tx'] == 1) & (df['cr_rosc'] == 1) & df.cr_scause.isin(value_list8)].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)


        if x == 2:
            N  = len(df.loc[(df['cr_tx'] == 1) & (df['cr_pdisp'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            NoObviouscause = len(df.loc[(df['cr_tx'] == 1) & (df['cr_pdisp'] == 1) & (df['cr_scause'] == 0)].index)
            NoObviouscausepercent = 0
            if NoObviouscause not in [0]:
                NoObviouscausepercent = round((NoObviouscause / NoObviouscause_total), 4)
            NoObviouscauseList.append(NoObviouscausepercent)

            Trauma = len(df.loc[(df['cr_tx'] == 1) & (df['cr_pdisp'] == 1) & (df['cr_scause'] != 0)].index)
            Trauma_percent = 0
            if Trauma not in [0]:
                Trauma_percent = round((Trauma / Trauma_total), 4)
            TraumaList.append(Trauma_percent)

            Unknown = len(df.loc[(df['cr_tx'] == 1) & (df['cr_pdisp'] == 1) & df.cr_scause.isin(value_list8)].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)



        if x == 3:
            N =  len(df.loc[(df['cr_tx'] == 1) & (df['cr_surv'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            NoObviouscause = len(df.loc[(df['cr_tx'] == 1) & (df['cr_surv'] == 1) & (df['cr_scause'] == 0)].index)
            NoObviouscausepercent = 0
            if NoObviouscause not in [0]:
                NoObviouscausepercent = round((NoObviouscause / NoObviouscause_total), 4)
            NoObviouscauseList.append(NoObviouscausepercent)

            Trauma = len(df.loc[(df['cr_tx'] == 1) & (df['cr_surv'] == 1) & (df['cr_scause'] != 0)].index)
            Trauma_percent = 0
            if Trauma not in [0]:
                Trauma_percent = round((Trauma / Trauma_total), 4)
            TraumaList.append(Trauma_percent)

            Unknown = len(
                df.loc[(df['cr_tx'] == 1) & (df['cr_surv'] == 1) & df.cr_scause.isin(value_list8)].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)

        if x == 4:
            N =  len(df.loc[(df['cr_tx'] == 1) & (df['cr_surv'] == 2)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            NoObviouscause = len(df.loc[(df['cr_tx'] == 1) & (df['cr_surv'] == 2)].index & (df['cr_scause'] == 0))
            NoObviouscausepercent = 0
            if NoObviouscause not in [0]:
                NoObviouscausepercent = round((NoObviouscause / NoObviouscause_total), 4)
            NoObviouscauseList.append(NoObviouscausepercent)

            Trauma = len(df.loc[(df['cr_tx'] == 1) & (df['cr_surv'] == 2)& (df['cr_scause'] != 0)].index)
            Trauma_percent = 0
            if Trauma not in [0]:
                Trauma_percent = round((Trauma / Trauma_total), 4)
            TraumaList.append(Trauma_percent)

            Unknown = len(
                df.loc[(df['cr_tx'] == 1) & (df['cr_surv'] == 2) & df.cr_scause.isin(value_list8)].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)

    AdultUpsteinWorksheet = AdultUstenworkbook['Sheet2']

    AdultUpsteinWorksheet['B114'] = 'All (' + str(Ntotal) + ')'
    AdultUpsteinWorksheet['B115'] = NList[0]
    AdultUpsteinWorksheet['B116'] = NList[1]
    AdultUpsteinWorksheet['B117'] = NList[2]
    AdultUpsteinWorksheet['B118'] = NList[3]


    AdultUpsteinWorksheet['C114'] = 'Presumed cardiac etiology  (' + str(NoObviouscause_total) + ')'
    AdultUpsteinWorksheet['C115'] = NoObviouscauseList[0]
    AdultUpsteinWorksheet['C116'] = NoObviouscauseList[1]
    AdultUpsteinWorksheet['C117'] = NoObviouscauseList[2]



    AdultUpsteinWorksheet['D114'] = 'Trauma or obvious cause (' + str(Trauma_total) + ')'
    AdultUpsteinWorksheet['D115'] = TraumaList[0]
    AdultUpsteinWorksheet['D116'] = TraumaList[1]
    AdultUpsteinWorksheet['D117'] = TraumaList[2]



    AdultUpsteinWorksheet['E114'] = 'Unknown Etiology (' + str(Unknown_total) + ')'
    AdultUpsteinWorksheet['E115'] = UnknownList[0]


    AdultUstenworkbook.save("C:Charts/AdultUsteinGraphs/CanRoc/Adult_Utstein_" + site + ".xlsx")
    print('Completed - Fig 12')


    # 13th   Call received at dispatch to 1st vehicle arrival interval,# breakdown by outcome (Witnessed by EMS excluded)

    TimeCalculation = []

    for index, row in df.iterrows():

        if  (str(row["cr_ptmrcv"]) != 'nan' and str(row["cr_rig1tm"]) != 'nan'):

            ptmrcv = datetime.strptime(row["cr_ptmrcv"], "%Y-%m-%d %H:%M:%S")
            rig1tm = datetime.strptime(row["cr_rig1tm"], "%Y-%m-%d %H:%M:%S")
            timediff = rig1tm - ptmrcv
            timeDiffSecond = timediff.total_seconds()/60


            TimeCalculation.append( timeDiffSecond)
        elif (str(row["cr_ptmrcv"]) == 'nan' or str(row["cr_rig1tm"]) == 'nan'):
            TimeCalculation.append(-1)


    df['TimeCalculation']= TimeCalculation




    value_list5 = [2, 3, 5, '2', '3', '5']
    value_list6 = [0, 1, 4, 6]
    value_list7 = [0, 1]
    value_list8 = [18, '', None]




    #For each row, rig1tm - ptmcrv in a new column, in minutes

    Ntotal = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)].index)
    zerototwo_total = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7) & (df['TimeCalculation'] >= 0) & (df['TimeCalculation'] <= 2)].index)
    twotofour_total = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['TimeCalculation'] > 2) & (df['TimeCalculation'] <= 4)].index)
    fourtosix_total = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['TimeCalculation'] > 4) & (df['TimeCalculation'] <= 6)].index)
    sixtoeight_total= len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['TimeCalculation'] > 6) & (df['TimeCalculation'] <= 8)].index)
    eighttoten_total = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['TimeCalculation'] > 8) & (df['TimeCalculation'] <= 10)].index)
    greaterthanten_total =  len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['TimeCalculation'] > 10) ].index)
    Unknown_total =         len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['TimeCalculation'] == -1) ].index)

    NList = []
    zerototwoList = []
    twotofourList = []
    fourtosixList = []
    sixtoeightList = []
    eighttotenList = []
    greaterthantenList = []
    UnknownList = []

    for x in range(1, 6):
        if x == 1:
            N = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7) & (df['cr_rosc'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            zerototwo = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_rosc'] == 1) & (df['TimeCalculation'] >= 0) & (df['TimeCalculation'] <= 2)].index)
            zerototwopercent = 0
            if zerototwo not in [0]:
                zerototwopercent = round((zerototwo / zerototwo_total), 4)
            zerototwoList.append(zerototwopercent)

            twotofour = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_rosc'] == 1)& (df['TimeCalculation'] > 2) & (df['TimeCalculation'] <= 4)].index)
            twotofour_percent = 0
            if twotofour not in [0]:
                twotofour_percent = round((twotofour / twotofour_total), 4)
            twotofourList.append(twotofour_percent)

            fourtosix = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_rosc'] == 1)& (df['TimeCalculation'] > 4) & (df['TimeCalculation'] <= 6)].index)
            fourtosixpercent = 0
            if fourtosix not in [0]:
                fourtosixpercent = round((fourtosix / fourtosix_total), 4)
            fourtosixList.append(fourtosixpercent)

            sixtoeight = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_rosc'] == 1)& (df['TimeCalculation'] > 6) & (df['TimeCalculation'] <= 8)].index)
            sixtoeightpercent = 0
            if  sixtoeight not in [0]:
                sixtoeightpercent = round(( sixtoeight /  sixtoeight_total), 4)
            sixtoeightList.append( sixtoeightpercent)

            eighttoten = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_rosc'] == 1)& (df['TimeCalculation'] > 8) & (df['TimeCalculation'] <= 10)].index)
            eighttotenpercent = 0
            if  eighttoten not in [0]:
                eighttotenpercent = round(( eighttoten /  eighttoten_total), 4)
            eighttotenList.append( eighttotenpercent)

            greaterthanten = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_rosc'] == 1)& (df['TimeCalculation'] > 10) ].index)
            greaterthantenpercent = 0
            if greaterthanten not in [0]:
                greaterthantenpercent = round((greaterthanten / greaterthanten_total), 4)
            greaterthantenList.append(greaterthantenpercent)

            Unknown = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_rosc'] == 1)& (df['TimeCalculation'] == -1) ].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)

        if x == 2:
            N = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7) & (df['cr_pdisp'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            zerototwo = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_pdisp'] == 1) & (df['TimeCalculation'] >= 0) & (df['TimeCalculation'] <= 2)].index)
            zerototwopercent = 0
            if zerototwo not in [0]:
                zerototwopercent = round((zerototwo / zerototwo_total), 4)
            zerototwoList.append(zerototwopercent)

            twotofour = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_pdisp'] == 1)& (df['TimeCalculation'] > 2) & (df['TimeCalculation'] <= 4)].index)
            twotofour_percent = 0
            if twotofour not in [0]:
                twotofour_percent = round((twotofour / twotofour_total), 4)
            twotofourList.append(twotofour_percent)

            fourtosix = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_pdisp'] == 1)& (df['TimeCalculation'] > 4) & (df['TimeCalculation'] <= 6)].index)
            fourtosixpercent = 0
            if fourtosix not in [0]:
                fourtosixpercent = round((fourtosix / fourtosix_total), 4)
            fourtosixList.append(fourtosixpercent)

            sixtoeight = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_pdisp'] == 1)& (df['TimeCalculation'] > 6) & (df['TimeCalculation'] <= 8)].index)
            sixtoeightpercent = 0
            if sixtoeight not in [0]:
                sixtoeightpercent = round((sixtoeight / sixtoeight_total), 4)
            sixtoeightList.append(sixtoeightpercent)

            eighttoten = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_pdisp'] == 1)& (df['TimeCalculation'] > 8) & (df['TimeCalculation'] <= 10)].index)
            eighttotenpercent = 0
            if eighttoten not in [0]:
                eighttotenpercent = round((eighttoten / eighttoten_total), 4)
            eighttotenList.append(eighttotenpercent)

            greaterthanten = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_pdisp'] == 1)& (df['TimeCalculation'] > 10) ].index)
            greaterthantenpercent = 0
            if greaterthanten not in [0]:
                greaterthantenpercent = round((greaterthanten / greaterthanten_total), 4)
            greaterthantenList.append(greaterthantenpercent)

            Unknown = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_pdisp'] == 1)& (df['TimeCalculation'] == -1) ].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)



        if x == 3:
            N = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7) & (df['cr_surv'] == 1)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            zerototwo = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_surv'] == 1) & (df['TimeCalculation'] >= 0) & (df['TimeCalculation'] <= 2)].index)
            zerototwopercent = 0
            if zerototwo not in [0]:
                zerototwopercent = round((zerototwo / zerototwo_total), 4)
            zerototwoList.append(zerototwopercent)

            twotofour = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_surv'] == 1)& (df['TimeCalculation'] > 2) & (df['TimeCalculation'] <= 4)].index)
            twotofour_percent = 0
            if twotofour not in [0]:
                twotofour_percent = round((twotofour / twotofour_total), 4)
            twotofourList.append(twotofour_percent)

            fourtosix = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_surv'] == 1)& (df['TimeCalculation'] > 4) & (df['TimeCalculation'] <= 6)].index)
            fourtosixpercent = 0
            if fourtosix not in [0]:
                fourtosixpercent = round((fourtosix / fourtosix_total), 4)
            fourtosixList.append(fourtosixpercent)

            sixtoeight = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_surv'] == 1)& (df['TimeCalculation'] > 6) & (df['TimeCalculation'] <= 8)].index)
            sixtoeightpercent = 0
            if sixtoeight not in [0]:
                sixtoeightpercent = round((sixtoeight / sixtoeight_total), 4)
            sixtoeightList.append(sixtoeightpercent)

            eighttoten = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_surv'] == 1)& (df['TimeCalculation'] > 8) & (df['TimeCalculation'] <= 10)].index)
            eighttotenpercent = 0
            if eighttoten not in [0]:
                eighttotenpercent = round((eighttoten / eighttoten_total), 4)
            eighttotenList.append(eighttotenpercent)

            greaterthanten = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_surv'] == 1)& (df['TimeCalculation'] > 10) ].index)
            greaterthantenpercent = 0
            if greaterthanten not in [0]:
                greaterthantenpercent = round((greaterthanten / greaterthanten_total), 4)
            greaterthantenList.append(greaterthantenpercent)

            Unknown = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_surv'] == 1)& (df['TimeCalculation'] == -1) ].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)

        if x == 4:
            N = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7) & (df['cr_surv'] == 2)].index)

            Npercent = 0
            if N not in [0]:
                Npercent = round((N / Ntotal), 4)
            NList.append(Npercent)

            zerototwo = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_surv'] == 2) & (df['TimeCalculation'] >= 0) & (df['TimeCalculation'] <= 2)].index)
            zerototwopercent = 0
            if zerototwo not in [0]:
                zerototwopercent = round((zerototwo / zerototwo_total), 4)
            zerototwoList.append(zerototwopercent)

            twotofour = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_surv'] == 2)& (df['TimeCalculation'] > 2) & (df['TimeCalculation'] <= 4)].index)
            twotofour_percent = 0
            if twotofour not in [0]:
                twotofour_percent = round((twotofour / twotofour_total), 4)
            twotofourList.append(twotofour_percent)

            fourtosix = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_surv'] == 2)& (df['TimeCalculation'] > 4) & (df['TimeCalculation'] <= 6)].index)
            fourtosixpercent = 0
            if fourtosix not in [0]:
                fourtosixpercent = round((fourtosix / fourtosix_total), 4)
            fourtosixList.append(fourtosixpercent)

            sixtoeight = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_surv'] == 2)& (df['TimeCalculation'] > 6) & (df['TimeCalculation'] <= 8)].index)
            sixtoeightpercent = 0
            if sixtoeight not in [0]:
                sixtoeightpercent = round((sixtoeight / sixtoeight_total), 4)
            sixtoeightList.append(sixtoeightpercent)

            eighttoten = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_surv'] == 2)& (df['TimeCalculation'] > 8) & (df['TimeCalculation'] <= 10)].index)
            eighttotenpercent = 0
            if eighttoten not in [0]:
                eighttotenpercent = round((eighttoten / eighttoten_total), 4)
            eighttotenList.append(eighttotenpercent)

            greaterthanten = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_surv'] == 2)& (df['TimeCalculation'] > 10) ].index)
            greaterthantenpercent = 0
            if greaterthanten not in [0]:
                greaterthantenpercent = round((greaterthanten / greaterthanten_total), 4)
            greaterthantenList.append(greaterthantenpercent)

            Unknown =  len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)& (df['cr_surv'] == 2)& (df['TimeCalculation'] == -1) ].index)
            Unknownpercent = 0
            if Unknown not in [0]:
                Unknownpercent = round((Unknown / Unknown_total), 4)
            UnknownList.append(Unknownpercent)


    AdultUpsteinWorksheet = AdultUstenworkbook['Sheet2']
    AdultUpsteinNextWorksheet = AdultUstenworkbook['Sheet3']

    AdultUpsteinWorksheet['B128'] = 'All (' + str(Ntotal) + ')'
    AdultUpsteinWorksheet['B129'] = NList[0]
    AdultUpsteinWorksheet['B130'] = NList[1]
    AdultUpsteinWorksheet['B131'] = NList[2]
    AdultUpsteinWorksheet['B132'] = NList[3]


    AdultUpsteinWorksheet['C128'] = '0-2 (' + str(zerototwo_total) + ')'
    AdultUpsteinWorksheet['C129'] = zerototwoList[0]
    AdultUpsteinWorksheet['C130'] = zerototwoList[1]
    AdultUpsteinWorksheet['C131'] = zerototwoList[2]


    AdultUpsteinWorksheet['D128'] = '>2-4 (' + str(twotofour_total) + ')'
    AdultUpsteinWorksheet['D129'] = twotofourList[0]
    AdultUpsteinWorksheet['D130'] = twotofourList[1]
    AdultUpsteinWorksheet['D131'] = twotofourList[2]



    AdultUpsteinWorksheet['E128'] = '>4-6 (' + str(fourtosix_total) + ')'
    AdultUpsteinWorksheet['E129'] = fourtosixList[0]
    AdultUpsteinWorksheet['E130'] = fourtosixList[1]
    AdultUpsteinWorksheet['E131'] = fourtosixList[2]
    AdultUpsteinWorksheet['E132'] = fourtosixList[3]


    AdultUpsteinWorksheet['F128'] = '>6-8 (' + str(sixtoeight_total) + ')'
    AdultUpsteinWorksheet['F129'] = sixtoeightList[0]
    AdultUpsteinWorksheet['F130'] = sixtoeightList[1]
    AdultUpsteinWorksheet['F131'] = sixtoeightList[2]



    AdultUpsteinWorksheet['G128'] = '>8-10 (' + str(eighttoten_total) + ')'
    AdultUpsteinWorksheet['G129'] = eighttotenList[0]
    AdultUpsteinWorksheet['G130'] = eighttotenList[1]
    AdultUpsteinWorksheet['G131'] = eighttotenList[2]



    AdultUpsteinWorksheet['H128'] = '>10 (' + str(greaterthanten_total) + ')'
    AdultUpsteinWorksheet['H129'] = greaterthantenList[0]
    AdultUpsteinWorksheet['H130'] = greaterthantenList[1]
    AdultUpsteinWorksheet['H131'] = greaterthantenList[2]


    AdultUpsteinWorksheet['I128'] = 'Unknown Interval (' + str(Unknown_total) + ')'
    AdultUpsteinWorksheet['I129'] = UnknownList[0]



    AdultUpsteinNextWorksheet['C298'] = 'All = '+ str(Ntotal)

    AdultUstenworkbook.save("C:/Charts/AdultUsteinGraphs/CanRoc/Adult_Utstein_" + site + ".xlsx")
    print('Completed - Fig 13')

    # 14th   Response Intervals - Call received at dispatch to 1st vehicle arrival interval (Witnessed by EMS excluded) (N=339)

    Ntotal = len( df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)].index)
    zerototwo_total = len(
        df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7) & (df['TimeCalculation'] >= 0) & (df['TimeCalculation'] <= 2)].index)
    twotofour_total = len(
        df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7) & (df['TimeCalculation'] > 2) & (df['TimeCalculation'] <= 4)].index)
    fourtosix_total = len(
        df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7) & (df['TimeCalculation'] > 4) & (df['TimeCalculation'] <= 6)].index)
    sixtoeight_total = len(
        df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7) & (df['TimeCalculation'] > 6) & (df['TimeCalculation'] <= 8)].index)
    eighttoten_total = len(
        df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7) & (df['TimeCalculation'] > 8) & (df['TimeCalculation'] <= 10)].index)
    greaterthanten_total = len( df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7) & (df['TimeCalculation'] > 10)].index)
    Unknown_total = len( df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7) & (df['TimeCalculation'] == -1)].index)

    zerototwo_percent = round(zerototwo_total / Ntotal, 4)
    twotofour_percent = round(twotofour_total / Ntotal, 4)
    fourtosix_tpercent = round(fourtosix_total / Ntotal, 4)
    sixtoeight_percent = round(sixtoeight_total / Ntotal, 4)
    eighttoten_percent = round( eighttoten_total / Ntotal, 4)
    greaterthanten_percent = round( greaterthanten_total / Ntotal, 4)
    Unknown_percent = round( Unknown_total / Ntotal, 4)


    AdultUpsteinWorksheet = AdultUstenworkbook['Sheet2']


    AdultUpsteinWorksheet['A145'] = zerototwo_percent
    AdultUpsteinWorksheet['B145'] = twotofour_percent
    AdultUpsteinWorksheet['C145'] = fourtosix_tpercent
    AdultUpsteinWorksheet['D145'] = sixtoeight_percent
    AdultUpsteinWorksheet['E145'] = eighttoten_percent
    AdultUpsteinWorksheet['F145'] = greaterthanten_percent
    AdultUpsteinWorksheet['G145'] = Unknown_percent
    AdultUpsteinNextWorksheet['L298'] = 'All = '+ str(Ntotal)
    AdultUstenworkbook.save("C:/Charts/AdultUsteinGraphs/CanRoc/Adult_Utstein_" + site + ".xlsx")
    print('Completed - Fig 12')


    # 15th   Response Intervals - Call received at dispatch to 1st vehicle arrival interval (Witnessed by EMS excluded) (N=339)
    allminutes = TimeCalculation

    newallminutes =  [x if x != -1 else np.nan for x in allminutes]

    graphmean =np.nanmean(newallminutes)
    graphmedian = np.nanmedian(newallminutes)
    graphsd = np.nanstd(newallminutes)
    graphicstenpercentile = np.nanpercentile(newallminutes, 10)
    graphicspercentile =  np.nanpercentile(newallminutes, 90)


    pd.set_option('display.max_rows', None)

    # print(df[['cr_rig1tm', 'cr_ptmrcv', 'TimeCalculation']])

    AdultUpsteinWorksheet['A147'] =  round(graphmean,4)
    AdultUpsteinWorksheet['B147'] = round(graphmedian,4)
    AdultUpsteinWorksheet['C147'] = round(graphicstenpercentile,4)
    AdultUpsteinWorksheet['D147'] = round(graphicspercentile,4)
    AdultUpsteinNextWorksheet['K292'] = 'All =' + str(Ntotal)
    AdultUstenworkbook.save("C:/Charts/AdultUsteinGraphs/CanRoc/Adult_Utstein_" + site + ".xlsx")
    print('Completed - Fig 13')


    # 16th   Service Response Intervals    # 1st Crew    # notified to 1st vehicle arrival interval    # (Witnessed by EMS excluded)

    TimeCalculation = []




    for index, row in df.iterrows():


        if (str(row["cr_rig1tm"]) not in ['nan',''] and str(row["cr_ptmcpr"]) not in ['nan','']):


            rig1tm = datetime.strptime(row["cr_rig1tm"], "%Y-%m-%d %H:%M:%S")

            cr_ptmcpr = datetime.strptime(row["cr_ptmcpr"], "%Y-%m-%d %H:%M:%S")


            timediff =  cr_ptmcpr - rig1tm

            timeDiffSecond = timediff.total_seconds() / 60
            print(timeDiffSecond)

            if timediff.total_seconds()<0 and timediff.total_seconds()>-60:
                timeDiffSecond*=-1
                print(timeDiffSecond)
            elif timediff.total_seconds() < -60:
                timeDiffSecond = -1
                print(timeDiffSecond)


            TimeCalculation.append(timeDiffSecond)


        else:
            TimeCalculation.append(-1)





    df['TimeCalculation2'] = TimeCalculation



    Ntotal = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)].index)
    zerototwo_total = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7) & (df['TimeCalculation2'] >= 0) & (
                            df['TimeCalculation2'] <= 2)].index)
    twotofour_total = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7) & (df['TimeCalculation2'] > 2) & (
                            df['TimeCalculation2'] <= 4)].index)
    fourtosix_total = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7) & (df['TimeCalculation2'] > 4) & (
                            df['TimeCalculation2'] <= 6)].index)
    sixtoeight_total = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7) & (df['TimeCalculation2'] > 6) & (
                            df['TimeCalculation2'] <= 8)].index)
    eighttoten_total = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7) & (df['TimeCalculation2'] > 8) & (
                            df['TimeCalculation2'] <= 10)].index)
    greaterthanten_total = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7) & (df['TimeCalculation2'] > 10)].index)
    Unknown_total = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7) & (df['TimeCalculation2'] == -1)].index)

    zerototwo_percent = round(zerototwo_total / Ntotal, 4)
    twotofour_percent = round(twotofour_total / Ntotal, 4)
    fourtosix_tpercent = round(fourtosix_total / Ntotal, 4)
    sixtoeight_percent = round(sixtoeight_total / Ntotal, 4)
    eighttoten_percent = round(eighttoten_total / Ntotal, 4)
    greaterthanten_percent = round(greaterthanten_total / Ntotal, 4)
    Unknown_percent = round(Unknown_total / Ntotal, 4)

    AdultUpsteinWorksheet['A166'] = zerototwo_percent
    AdultUpsteinWorksheet['B166'] = twotofour_percent
    AdultUpsteinWorksheet['C166'] = fourtosix_tpercent
    AdultUpsteinWorksheet['D166'] = sixtoeight_percent
    AdultUpsteinWorksheet['E166'] = eighttoten_percent
    AdultUpsteinWorksheet['F166'] = greaterthanten_percent
    AdultUpsteinWorksheet['G166'] = Unknown_percent
    # AdultUpsteinNextWorksheet['B319'] = 'N =' + str(Ntotal)
    AdultUstenworkbook.save("C:/Charts/AdultUsteinGraphs/CanRoc/Adult_Utstein_" + site + ".xlsx")
    print('Completed - Fig 16')




    # 17th   Service Response Intervals    # 1st Crew    # notified to 1st vehicle arrival interval    # (Witnessed by EMS excluded) Averages
    allminutes = TimeCalculation

    newallminutes = [x if x != -1 else np.nan for x in allminutes]

    print(newallminutes)
    # sys.exit()


    graphmean = np.nanmean(newallminutes)
    graphmedian = np.nanmedian(newallminutes)
    graphsd = np.nanstd(newallminutes)
    graphicstenpercentile = np.nanpercentile(newallminutes, 10)
    graphicspercentile = np.nanpercentile(newallminutes, 90)


    # pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)


    # print(df[['cr_rig1tm', 'cr_ptmrcv', 'TimeCalculation']])


    AdultUpsteinWorksheet['A168'] = round(graphmean, 4)
    AdultUpsteinWorksheet['B168'] = round(graphmedian, 4)
    AdultUpsteinWorksheet['C168'] = round(graphicstenpercentile, 4)
    AdultUpsteinWorksheet['D168'] = round(graphicspercentile, 4)
    AdultUpsteinNextWorksheet['B309'] = 'N =' + str(Ntotal)
    AdultUstenworkbook.save("C:/Users/mcbarnettr/Desktop/Charts/AdultUsteinGraphs/CanRoc/Adult_Utstein_" + site + ".xlsx")

    print('Completed - Fig 17')
    value_list7 = [0, 1]
    value_list8=[2,3,5]

    # 18th   Advanced Interventions Performed (Witnessed by EMS    # excluded) * **s
    Ntotal = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7)].index)
    Epinephrine_total = len(df.loc[(df['cr_tx'] == 1) & df.cr_witbys.isin(value_list7) & df['cr_epi'] == 1].index)
    Antiarryth_total = len(df.loc[(df['cr_tx'] == 1)  & (df.cr_witbys.isin(value_list7)) & (df['cr_arryth'] == 1 ) & (df.cr_frhyem.isin(value_list8))].index)
    Antiarryth_Ntotal = len(df.loc[(df['cr_tx'] == 1) & (df.cr_witbys.isin(value_list7))& (df.cr_frhyem.isin(value_list8))].index)

    print (Antiarryth_total )
    print(Antiarryth_Ntotal)



    Advanced_totaldf = df[(df['cr_lmasuc___1']== 1) | (df['cr_igelsuc___1'] == 1)| (df['cr_kingsuc___1'] == 1) | (df['cr_ettsuc___1'] == 1) |
                          (df['cr_othawsuc___1'] == 1)]

    Advanced_total = len(Advanced_totaldf.loc[(Advanced_totaldf['cr_tx'] == 1) & Advanced_totaldf.cr_witbys.isin(value_list7)].index)


    Epinephrine_percent = 0

    if Ntotal not in [0]:
        Epinephrine_percent = round((Epinephrine_total / Ntotal), 4)

    Antiarryth_percent = 0
    if Antiarryth_Ntotal not in [0]:
        Antiarryth_percent = round((Antiarryth_total / Antiarryth_Ntotal), 4) # *******TO FIX! denominator needs additional filters: cr_frhyem=2, 3, or 5 to be relevant

    Advanced_percent = 0
    if Ntotal not in [0]:
        Advanced_percent = round((Advanced_total / Ntotal), 4)

    AdultUpsteinWorksheet['B210'] =  Epinephrine_percent
    AdultUpsteinWorksheet['B211'] = Antiarryth_percent
    AdultUpsteinWorksheet['B212'] = Advanced_percent
    AdultUpsteinNextWorksheet['B326'] = 'N =' + str(Ntotal)

    AdultUstenworkbook.save("C:/Charts/AdultUsteinGraphs/CanRoc/Adult_Utstein_" + site + ".xlsx")
    print('Completed - Fig 16')





    # Compression Rate , Fraction, Depth

    # Compression Rate

    fields1 = ['cr_epdt','cr_cprprc', 'cr_cmprt1', 'cr_cmprt2', 'cr_cmprt3', 'cr_cmprt4', 'cr_cmprt5', 'cr_cmprt6', 'cr_cmprt7',
               'cr_cmprt8', 'cr_cmprt9', 'cr_cmprt10','cr_scause']

    compdf = project.export_records(format='df', fields=fields1)
    pd.set_option("display.max.columns", None)
    pd.set_option('display.max_rows', None)

    print(compdf.head(100))

    compdf = compdf[(compdf['cr_scause'] != 17)]
    indexdf2 = compdf.reset_index()  # shift the index from cr_recordid so we can filter by RecordId

    indexdf2 = indexdf2[indexdf2["cr_cprprc"] == 1.0]
    filter_date = []

    for value in indexdf2["cr_epdt"]:

        if str(value) != 'nan':
            filter_date.append(datetime.strptime(value, "%Y-%m-%d"))
        else:
            filter_date.append(None)

    pd.set_option("display.max.columns", None)
    pd.set_option('display.max_rows', None)

    indexdf2 ["filter_date"] = filter_date



    indexdf2 = indexdf2[(indexdf2['filter_date'] >= start_date_changed) & (indexdf2['filter_date'] <= end_date_changed) ]
    nationaldf = indexdf2.copy()
    if site != 'ALL':
        indexdf2 = indexdf2[indexdf2["cr_record_id"].str.contains(site)]  # to get a count of only the rows with site eg BC

    indexdf2.reset_index(drop=True, inplace=True)
    index = indexdf2.index
    number_of_rows = len(index)  # Total number of analyzable records



    del indexdf2["cr_record_id"]
    del indexdf2["cr_cprprc"]
    del indexdf2["filter_date"]
    del indexdf2["cr_epdt"]

    total_analyzable_min_old = indexdf2.count().sum()  # count of total analyzable minutes

    indexdf2 = indexdf2.apply(pd.to_numeric)  # getting the dataset ready to check numeric conditions
    print(indexdf2.head(100))

    total_records  = 0 # count of records with at least one minute
    range_minute_count = 0

    total_analyzable_min = 0
    for i in range(len(indexdf2)):
        goodrow = False
        if indexdf2.loc[i, 'cr_cmprt1'] not in [None, ''] and not (pd.isna(indexdf2.loc[i, 'cr_cmprt1']))  and indexdf2.loc[i, 'cr_cmprt1'] in range(0, 226):

            total_analyzable_min += 1
            if indexdf2.loc[i, 'cr_cmprt1'] in range(100, 121):
                range_minute_count += 1
            goodrow = True
        if indexdf2.loc[i, 'cr_cmprt2'] not in [None, '', 'nan'] and not (pd.isna(indexdf2.loc[i, 'cr_cmprt2']))  and indexdf2.loc[i, 'cr_cmprt2'] in range(0, 226):

            total_analyzable_min += 1
            if indexdf2.loc[i, 'cr_cmprt2'] in range(100, 121):
                range_minute_count += 1
            goodrow = True
        if indexdf2.loc[i, 'cr_cmprt3'] not in [None, '', 'nan'] and not (pd.isna(indexdf2.loc[i, 'cr_cmprt3']))  and indexdf2.loc[i, 'cr_cmprt3'] in range(0, 226):

            total_analyzable_min += 1
            if indexdf2.loc[i, 'cr_cmprt3'] in range(100, 121):
                range_minute_count += 1
            goodrow = True
        if indexdf2.loc[i, 'cr_cmprt4'] not in [None, '', 'nan'] and not (pd.isna(indexdf2.loc[i, 'cr_cmprt4']))  and indexdf2.loc[i, 'cr_cmprt4'] in range(0, 226):

            total_analyzable_min += 1
            if indexdf2.loc[i, 'cr_cmprt4'] in range(100, 121):
                range_minute_count += 1
            goodrow = True
        if indexdf2.loc[i, 'cr_cmprt5'] not in [None, '', 'nan'] and not (pd.isna(indexdf2.loc[i, 'cr_cmprt5']))  and indexdf2.loc[i, 'cr_cmprt5'] in range(0, 226):

            total_analyzable_min += 1
            if indexdf2.loc[i, 'cr_cmprt5'] in range(100, 121):
                range_minute_count += 1
            goodrow = True
        if indexdf2.loc[i, 'cr_cmprt6'] not in [None, '', 'nan'] and not (pd.isna(indexdf2.loc[i, 'cr_cmprt6']))  and indexdf2.loc[i, 'cr_cmprt6'] in range(0, 226):

            total_analyzable_min += 1
            if indexdf2.loc[i, 'cr_cmprt6'] in range(100, 121):
                range_minute_count += 1
            goodrow = True
        if indexdf2.loc[i, 'cr_cmprt7'] not in [None, '', 'nan'] and not (pd.isna(indexdf2.loc[i, 'cr_cmprt7']))  and indexdf2.loc[i, 'cr_cmprt7'] in range(0, 226):

            total_analyzable_min += 1
            if indexdf2.loc[i, 'cr_cmprt7'] in range(100, 121):
                range_minute_count += 1
            goodrow = True
        if indexdf2.loc[i, 'cr_cmprt8'] not in [None, '', 'nan'] and not (pd.isna(indexdf2.loc[i, 'cr_cmprt8']))  and indexdf2.loc[i, 'cr_cmprt8'] in range(0, 226):

            total_analyzable_min += 1
            if indexdf2.loc[i, 'cr_cmprt8'] in range(100, 121):
                range_minute_count += 1
            goodrow = True
        if indexdf2.loc[i, 'cr_cmprt9'] not in [None, '', 'nan'] and not (pd.isna(indexdf2.loc[i, 'cr_cmprt9']))  and indexdf2.loc[i, 'cr_cmprt9'] in range(0, 226):

            total_analyzable_min += 1
            if indexdf2.loc[i, 'cr_cmprt9'] in range(100, 121):
                range_minute_count += 1
            goodrow = True
        if indexdf2.loc[i, 'cr_cmprt10'] not in [None, '', 'nan'] and not (pd.isna(indexdf2.loc[i, 'cr_cmprt10']))   and indexdf2.loc[i, 'cr_cmprt10'] in range(0, 226):

            total_analyzable_min += 1
            if indexdf2.loc[i, 'cr_cmprt10'] in range(100, 121):
                range_minute_count += 1
            goodrow = True
        if goodrow == True:
            total_records += 1 #total number of records

    minutepercent = 0
    if total_analyzable_min != 0:
        minutepercent = round(range_minute_count/total_analyzable_min,4)


    median = 0
    percent_10 = 0
    percent_90 = 0
    nationalmean = 0


    del nationaldf["cr_record_id"]
    del nationaldf["cr_cprprc"]
    del nationaldf["filter_date"]
    del nationaldf["cr_epdt"]
    nationaldf = nationaldf.stack().reset_index()

    nationaldf = nationaldf[(nationaldf[0] <= 255) & (nationaldf[0] >=0)]
    print(nationaldf)

    nationalmean = nationaldf.loc[:, 0].mean()
    nationalmedian = nationaldf.loc[:, 0].median()

    df1 = indexdf2.stack().reset_index()
    df1 =  df1[( df1[0] <= 255) & ( df1[0] >= 0)]

    median = df1.loc[:, 0].median()

    mean = df1.loc[:, 0].mean()


    percent_10 = df1.loc[:, 0].quantile(0.1)  # 10th percentile
    print(percent_10 )
    percent_90 = df1.loc[:, 0].quantile(0.90)  # 10th percentile
    print(percent_90 )



    rateDivisions = ['Median', '10th Percentile', '90th Percentile']
    rateScores = [median, percent_10, percent_90]

    AdultUpsteinWorksheet['B186'] = round(median, 4)
    AdultUpsteinWorksheet['C186'] = round(percent_10, 4)
    AdultUpsteinWorksheet['D186'] = round(percent_90, 4)
    AdultUpsteinWorksheet['B187'] = round(nationalmedian, 4)
    AdultUpsteinWorksheet['C187'] = round(nationalmedian, 4)
    AdultUpsteinWorksheet['D187'] = round(nationalmedian, 4)
    if site == "ALL":
        AdultUpsteinWorksheet['B187'] = ''
        AdultUpsteinWorksheet['C187'] = ''
        AdultUpsteinWorksheet['D187'] = ''
        AdultUpsteinNextWorksheet['C356'] =''
        AdultUpsteinNextWorksheet['B356'].fill = PatternFill(start_color="FFFFFF", fill_type="solid")
    else:
        AdultUpsteinNextWorksheet['C356'] = 'Canadian National Median'


    AdultUpsteinNextWorksheet['B338'] = f'({total_records} records, {round(minutepercent*100,2)} % of minutes meet target of 100 - 120)'

    AdultUstenworkbook.save("C:/Users/mcbarnettr/Desktop/Charts/AdultUsteinGraphs/CanRoc/Adult_Utstein_" + site + ".xlsx")

    print('Completed - Fig 17')


    # Compression Fraction

    # plt.show()

    fields2 = ["cr_epdt","cr_cprprc",'cr_cprff1', 'cr_cprff2', 'cr_cprff3', 'cr_cprff4', 'cr_cprff5', 'cr_cprff6', 'cr_cprff7', 'cr_cprff8',
               'cr_cprff9', 'cr_cprff10','cr_scause']
    fraction_df = project.export_records(format='df', fields=fields2)
    pd.set_option("display.max.columns", None)
    pd.set_option('display.max_rows', None)
    fraction_df = fraction_df[(fraction_df['cr_scause'] != 17)]
    fraction_indexdf2 = fraction_df.reset_index()  # shift the index from cr_recordid so we can filter by RecordId
    fraction_indexdf2 = fraction_indexdf2[fraction_indexdf2["cr_cprprc"] == 1.0]

    filter_date = []

    for value in fraction_indexdf2["cr_epdt"]:

        if str(value) != 'nan':
            filter_date.append(datetime.strptime(value, "%Y-%m-%d"))
        else:
            filter_date.append(None)

    pd.set_option("display.max.columns", None)
    pd.set_option('display.max_rows', None)

    fraction_indexdf2["filter_date"] = filter_date

    fraction_indexdf2 =fraction_indexdf2[(fraction_indexdf2['filter_date'] >= start_date_changed) & (fraction_indexdf2['filter_date'] <= end_date_changed)]




    fraction_nationaldf = fraction_indexdf2.copy()

    if site != 'ALL':
        fraction_indexdf2 = fraction_indexdf2[
            fraction_indexdf2["cr_record_id"].str.contains(site)]  # to get a count of only the rows with site eg BC

    fraction_indexdf2.reset_index(drop=True, inplace=True)

    index = fraction_indexdf2.index
    fraction_number_of_rows = len(index)

    print( 'fraction_number_of_rows')
    print(fraction_number_of_rows)
    del fraction_indexdf2["cr_record_id"]
    del  fraction_indexdf2["cr_cprprc"]
    del fraction_indexdf2["filter_date"]
    del fraction_indexdf2["cr_epdt"]


    fraction_total_analyzable_min = 0

    fractiontotal_records = 0  # count of records with at least one minute
    fractionrange_minute_count = 0

    fraction_indexdf2 = fraction_indexdf2.apply(pd.to_numeric)  # getting the dataset ready to check numeric conditions
    fraction_total_analyzable_min = fraction_indexdf2.count().sum()  # count of total analyzable minutes
    total_analyzable_min = 0

    print( fraction_indexdf2 )




    for i in range(len(fraction_indexdf2)):
        goodrow = False
        if fraction_indexdf2.loc[i, 'cr_cprff1']  not in [None, '', 'nan'] and not (pd.isna(fraction_indexdf2.loc[i, 'cr_cprff1'])) and fraction_indexdf2.loc[i, 'cr_cprff1'] in range(0, 2):
            total_analyzable_min += 1
            if fraction_indexdf2.loc[i, 'cr_cprff1'] >= .60:
                fractionrange_minute_count += 1
            goodrow = True
        if fraction_indexdf2.loc[i, 'cr_cprff2'] not in [None, '', 'nan'] and not (pd.isna(fraction_indexdf2.loc[i, 'cr_cprff2'])) and fraction_indexdf2.loc[i, 'cr_cprff2'] in range(0, 2):
            total_analyzable_min += 1
            if fraction_indexdf2.loc[i, 'cr_cprff2'] >= .60:
                fractionrange_minute_count += 1
            goodrow = True
        if fraction_indexdf2.loc[i, 'cr_cprff3'] not in [None, '', 'nan'] and not (pd.isna(fraction_indexdf2.loc[i, 'cr_cprff3'])) and fraction_indexdf2.loc[i, 'cr_cprff3'] in range(0, 2):
            total_analyzable_min += 1
            if fraction_indexdf2.loc[i, 'cr_cprff3'] >= .60:
                fractionrange_minute_count += 1
            goodrow = True
        if fraction_indexdf2.loc[i, 'cr_cprff4'] not in [None, '', 'nan'] and not (pd.isna(fraction_indexdf2.loc[i, 'cr_cprff4'])) and fraction_indexdf2.loc[i, 'cr_cprff4'] in range(0, 2):
            total_analyzable_min += 1
            if fraction_indexdf2.loc[i, 'cr_cprff4'] >= .60:
                fractionrange_minute_count += 1
            goodrow = True
        if fraction_indexdf2.loc[i, 'cr_cprff5'] not in [None, '', 'nan'] and not (pd.isna(fraction_indexdf2.loc[i, 'cr_cprff5']))and fraction_indexdf2.loc[i, 'cr_cprff5'] in range(0, 2):
            total_analyzable_min += 1
            if fraction_indexdf2.loc[i, 'cr_cprff5'] >= .60:
                fractionrange_minute_count += 1
            goodrow = True
        if fraction_indexdf2.loc[i, 'cr_cprff6'] not in [None, '', 'nan'] and not (pd.isna(fraction_indexdf2.loc[i, 'cr_cprff6'])) and fraction_indexdf2.loc[i, 'cr_cprff6'] in range(0, 2):
            total_analyzable_min += 1
            if fraction_indexdf2.loc[i, 'cr_cprff6'] >= .60:
                fractionrange_minute_count += 1
            goodrow = True
        if fraction_indexdf2.loc[i, 'cr_cprff7'] not in [None, '', 'nan'] and not (pd.isna(fraction_indexdf2.loc[i, 'cr_cprff7']))and fraction_indexdf2.loc[i, 'cr_cprff7'] in range(0, 2):
            total_analyzable_min += 1
            if fraction_indexdf2.loc[i, 'cr_cprff7'] >= .60:
                fractionrange_minute_count += 1
            goodrow = True
        if fraction_indexdf2.loc[i, 'cr_cprff8'] not in [None, '', 'nan'] and not (pd.isna(fraction_indexdf2.loc[i, 'cr_cprff8']))and fraction_indexdf2.loc[i, 'cr_cprff8'] in range(0, 2):
            total_analyzable_min += 1
            if fraction_indexdf2.loc[i, 'cr_cprff8'] >= .60:
                fractionrange_minute_count += 1
            goodrow = True
        if fraction_indexdf2.loc[i, 'cr_cprff9'] not in [None, '', 'nan'] and not (pd.isna(fraction_indexdf2.loc[i, 'cr_cprff9']))and fraction_indexdf2.loc[i, 'cr_cprff9'] in range(0, 2):
            total_analyzable_min += 1
            if fraction_indexdf2.loc[i, 'cr_cprff9'] >= .60:
                fractionrange_minute_count += 1
            goodrow = True
        if fraction_indexdf2.loc[i, 'cr_cprff10'] not in [None, '', 'nan'] and not (pd.isna(fraction_indexdf2.loc[i, 'cr_cprff10']))and fraction_indexdf2.loc[i, 'cr_cprff10'] in range(0, 2):
            total_analyzable_min += 1
            if fraction_indexdf2.loc[i, 'cr_cprff10'] >= .60:
                 fractionrange_minute_count += 1
            goodrow = True

        if goodrow == True:
            fractiontotal_records += 1
    fractionminutepercent = 0
    print('total_analyzable_min')
    print( total_analyzable_min)
    print(fraction_total_analyzable_min)

    if total_analyzable_min !=0:
        fractionminutepercent = round(fractionrange_minute_count / total_analyzable_min, 4)
    median = 0
    percent_10 = 0
    percent_90 = 0
    nationalmean = 0

    del fraction_nationaldf["cr_record_id"]
    del fraction_nationaldf["cr_cprprc"]
    del fraction_nationaldf["filter_date"]
    del fraction_nationaldf["cr_epdt"]
    fraction_nationaldf = fraction_nationaldf.stack().reset_index()
    fraction_nationaldf =  fraction_nationaldf[( fraction_nationaldf[0] <= 1) & ( fraction_nationaldf[0] >= 0)]

    fraction_nationalmean = fraction_nationaldf.loc[:, 0].mean()
    fraction_nationalmedian = fraction_nationaldf.loc[:, 0].median()

    fraction_df1 =fraction_indexdf2.stack().reset_index()
    fraction_df1 = fraction_df1[(fraction_df1[0] <= 1) & (fraction_df1[0] >= 0)]



    median = fraction_df1.loc[:, 0].median()

    mean = fraction_df1.loc[:, 0].mean()

    percent_10 = fraction_df1.loc[:, 0].quantile(0.1)  # 10th percentile

    percent_90 = fraction_df1.loc[:, 0].quantile(0.90)  # 10th percentile



    rateDivisions = ['Median', '10th Percentile', '90th Percentile']
    rateScores = [median, percent_10, percent_90]

    AdultUpsteinWorksheet['B189'] = round(median, 4)
    AdultUpsteinWorksheet['C189'] = round(percent_10, 4)
    AdultUpsteinWorksheet['D189'] = round(percent_90, 4)
    AdultUpsteinWorksheet['B190'] = round(fraction_nationalmedian, 4)
    AdultUpsteinWorksheet['C190'] = round(fraction_nationalmedian, 4)
    AdultUpsteinWorksheet['D190'] = round(fraction_nationalmedian, 4)
    if site =='ALL':
        AdultUpsteinWorksheet['B190'] = ''
        AdultUpsteinWorksheet['C190'] = ''
        AdultUpsteinWorksheet['D190'] = ''
        AdultUpsteinNextWorksheet['K356'] = ''
        AdultUpsteinNextWorksheet['J356'].fill = PatternFill(start_color="FFFFFF", fill_type="solid")
    else:
        AdultUpsteinNextWorksheet['K356'] = 'Canadian National Median'



    AdultUpsteinNextWorksheet['J338'] = f'({fractiontotal_records } records, {round(fractionminutepercent * 100,2)} % of minutes meet target of >=.60)'

    AdultUstenworkbook.save("C:/Charts/AdultUsteinGraphs/CanRoc/Adult_Utstein_" + site + ".xlsx")

    # Compression Depth

    fields2 = ["cr_epdt","cr_cprprc",'cr_cdpth1', 'cr_cdpth2', 'cr_cdpth3', 'cr_cdpth4', 'cr_cdpth5', 'cr_cdpth6', 'cr_cdpth7', 'cr_cdpth8',
               'cr_cdpth9', 'cr_cdpth10','cr_scause']



    depth_df = project.export_records(format='df', fields=fields2)
    pd.set_option("display.max.columns", None)
    pd.set_option('display.max_rows', None)
    depth_df = depth_df[(depth_df['cr_scause'] != 17)]
    depth_indexdf2 = depth_df.reset_index()  # shift the index from cr_recordid so we can filter by RecordId
    depth_indexdf2 = depth_indexdf2[depth_indexdf2["cr_cprprc"] == 1.0]
    filter_date = []

    for value in depth_indexdf2["cr_epdt"]:

        if str(value) != 'nan':
            filter_date.append(datetime.strptime(value, "%Y-%m-%d"))
        else:
            filter_date.append(None)

    pd.set_option("display.max.columns", None)
    pd.set_option('display.max_rows', None)

    depth_indexdf2["filter_date"] = filter_date

    depth_indexdf2 = depth_indexdf2[(depth_indexdf2['filter_date'] >= start_date_changed) & (
                depth_indexdf2['filter_date'] <= end_date_changed)]

    depth_nationaldf = depth_indexdf2.copy()


    if site != 'ALL':
        depth_indexdf2 = depth_indexdf2[
            depth_indexdf2["cr_record_id"].str.contains(site)]  # to get a count of only the rows with site eg BC

    depth_indexdf2.reset_index(drop=True, inplace=True)

    print(depth_indexdf2)


    index = depth_indexdf2.index
    depth_number_of_rows = len(index)

    print('depth_number_of_rows')
    print(depth_number_of_rows)
    del depth_indexdf2["cr_record_id"]
    del depth_indexdf2["cr_cprprc"]
    del depth_indexdf2["cr_epdt"]
    del depth_indexdf2["filter_date"]

    print(depth_indexdf2)


    depth_total_analyzable_min = 0

    depthtotal_records = 0  # count of records with at least one minute
    depthrange_minute_count = 0

    depth_indexdf2 = depth_indexdf2.apply(pd.to_numeric)  # getting the dataset ready to check numeric conditions
    depth_total_analyzable_min = depth_indexdf2.count().sum()  # count of total analyzable minutes
    print(depth_total_analyzable_min)

    total_analyzable_min = 0
    for i in range(len(depth_indexdf2)):
        goodrow = False
        if depth_indexdf2.loc[i, 'cr_cdpth1'] not in [None, ''] and not (pd.isna(depth_indexdf2.loc[i, 'cr_cdpth1'])) and depth_indexdf2.loc[i, 'cr_cdpth1'] in range(0, 14):

            total_analyzable_min += 1
            print('1')
            if depth_indexdf2.loc[i, 'cr_cdpth1'] >= 5 and depth_indexdf2.loc[i, 'cr_cdpth1'] <=6 :
                depthrange_minute_count += 1
            goodrow = True
        if depth_indexdf2.loc[i, 'cr_cdpth2'] not in [None, ''] and not (pd.isna(depth_indexdf2.loc[i, 'cr_cdpth2']))and depth_indexdf2.loc[i, 'cr_cdpth2'] in range(0, 14):

            total_analyzable_min += 1
            print('2')
            if depth_indexdf2.loc[i, 'cr_cdpth2'] >= 5 and depth_indexdf2.loc[i, 'cr_cdpth2'] <=6 :
                depthrange_minute_count += 1
            goodrow = True
        if depth_indexdf2.loc[i, 'cr_cdpth3'] not in [None, ''] and not (pd.isna(depth_indexdf2.loc[i, 'cr_cdpth3']))and depth_indexdf2.loc[i, 'cr_cdpth3'] in range(0, 14):

            total_analyzable_min += 1
            print('3')
            if depth_indexdf2.loc[i, 'cr_cdpth3'] >= 5 and depth_indexdf2.loc[i, 'cr_cdpth3'] <=6 :
                depthrange_minute_count += 1
            goodrow = True

        if depth_indexdf2.loc[i, 'cr_cdpth4'] not in [None, ''] and not (pd.isna(depth_indexdf2.loc[i, 'cr_cdpth4']))and depth_indexdf2.loc[i, 'cr_cdpth4'] in range(0, 14):

            total_analyzable_min += 1
            print('4')
            print(depth_indexdf2.loc[i, 'cr_cdpth4'])
            if depth_indexdf2.loc[i, 'cr_cdpth4'] >= 5 and depth_indexdf2.loc[i, 'cr_cdpth4'] <=6 :
                depthrange_minute_count += 1
            goodrow = True
        if depth_indexdf2.loc[i, 'cr_cdpth5'] not in [None, ''] and not (pd.isna(depth_indexdf2.loc[i, 'cr_cdpth5']))and depth_indexdf2.loc[i, 'cr_cdpth5'] in range(0, 14):

            total_analyzable_min += 1
            print(depth_indexdf2.loc[i, 'cr_cdpth5'])
            print('5')
            if depth_indexdf2.loc[i, 'cr_cdpth5'] >= 5 and depth_indexdf2.loc[i, 'cr_cdpth5'] <=6 :
                depthrange_minute_count += 1
            goodrow = True
        if depth_indexdf2.loc[i, 'cr_cdpth6'] not in [None, ''] and not (pd.isna(depth_indexdf2.loc[i, 'cr_cdpth6']))and depth_indexdf2.loc[i, 'cr_cdpth6'] in range(0, 14):

            total_analyzable_min += 1
            print('6')
            if depth_indexdf2.loc[i, 'cr_cdpth6'] >= 5 and depth_indexdf2.loc[i, 'cr_cdpth6'] <=6 :
                depthrange_minute_count += 1
            goodrow = True
        if depth_indexdf2.loc[i, 'cr_cdpth7'] not in [None, ''] and not (pd.isna(depth_indexdf2.loc[i, 'cr_cdpth7']))and depth_indexdf2.loc[i, 'cr_cdpth7'] in range(0, 14):

            total_analyzable_min += 1
            print('7')
            if depth_indexdf2.loc[i, 'cr_cdpth7'] >= 5 and depth_indexdf2.loc[i, 'cr_cdpth7'] <=6 :
                depthrange_minute_count += 1
            goodrow = True
        if depth_indexdf2.loc[i, 'cr_cdpth8'] not in [None, ''] and not (pd.isna(depth_indexdf2.loc[i, 'cr_cdpth8']))and depth_indexdf2.loc[i, 'cr_cdpth8'] in range(0, 14):

            total_analyzable_min += 1
            print('8')
            if depth_indexdf2.loc[i, 'cr_cdpth8'] >= 5 and depth_indexdf2.loc[i, 'cr_cdpth8'] <=6 :
                depthrange_minute_count += 1
            goodrow = True
        if depth_indexdf2.loc[i, 'cr_cdpth9'] not in [None, ''] and not (pd.isna(depth_indexdf2.loc[i, 'cr_cdpth9']))and depth_indexdf2.loc[i, 'cr_cdpth9'] in range(0, 14):

            total_analyzable_min += 1
            print('9')
            if depth_indexdf2.loc[i, 'cr_cdpth9'] >= 5 and depth_indexdf2.loc[i, 'cr_cdpth9'] <=6 :
                depthrange_minute_count += 1
            goodrow = True
        if depth_indexdf2.loc[i, 'cr_cdpth10'] not in [None, ''] and not (pd.isna(depth_indexdf2.loc[i, 'cr_cdpth10']))and depth_indexdf2.loc[i, 'cr_cdpth10'] in range(0, 14):

            total_analyzable_min += 1
            print('10')
            if depth_indexdf2.loc[i, 'cr_cdpth10'] >= 5 and depth_indexdf2.loc[i, 'cr_cdpth10'] <=6 :
                depthrange_minute_count += 1
            goodrow = True

        if goodrow == True:
            depthtotal_records += 1

    print(depthrange_minute_count)
    print(depth_total_analyzable_min)
    print(total_analyzable_min)






    #########
    depthminutepercent = 0
    if total_analyzable_min != 0:
        depthminutepercent = round(depthrange_minute_count /  total_analyzable_min , 4)
    median = 0
    percent_10 = 0
    percent_90 = 0
    nationalmean = 0

    del depth_nationaldf["cr_record_id"]
    del depth_nationaldf["cr_cprprc"]
    del depth_nationaldf["cr_epdt"]
    del depth_nationaldf["filter_date"]
    depth_nationaldf = depth_nationaldf.stack().reset_index()
    depth_nationaldf = depth_nationaldf[(depth_nationaldf[0] <= 13) & (depth_nationaldf[0] >= .1)]

    depth_nationalmean = depth_nationaldf.loc[:, 0].mean()
    depth_nationalmedian = depth_nationaldf.loc[:, 0].median()

    depth_df1 = depth_indexdf2.stack().reset_index()
    print(depth_df1)

    depth_df1 = depth_df1[(depth_df1[0] <= 13) &  (depth_df1[0] >= 0)]

    print(depth_df1)

    # sys.exit()

    median = depth_df1.loc[:, 0].median()

    mean = depth_df1.loc[:, 0].mean()

    percent_10 = depth_df1.loc[:, 0].quantile(0.1)  # 10th percentile

    percent_90 = depth_df1.loc[:, 0].quantile(0.90)  # 10th percentile

    ############
    rateDivisions = ['Median', '10th Percentile', '90th Percentile']
    rateScores = [median, percent_10, percent_90]

    AdultUpsteinWorksheet['B192'] = round(median, 4)
    AdultUpsteinWorksheet['C192'] = round(percent_10, 4)
    AdultUpsteinWorksheet['D192'] = round(percent_90, 4)
    AdultUpsteinWorksheet['B193'] = round(depth_nationalmedian, 4)
    AdultUpsteinWorksheet['C193'] = round(depth_nationalmedian, 4)
    AdultUpsteinWorksheet['D193'] = round(depth_nationalmedian, 4)
    if site == 'ALL':
        AdultUpsteinWorksheet['B193'] = ''
        AdultUpsteinWorksheet['C193'] = ''
        AdultUpsteinWorksheet['D193'] = ''
        AdultUpsteinNextWorksheet['F374'] = ''
        AdultUpsteinNextWorksheet['E374'].fill = PatternFill(start_color="FFFFFF", fill_type="solid")
    else:
        AdultUpsteinNextWorksheet['F374'] = 'Canadian National Median'

    AdultUpsteinNextWorksheet[
        'F357'] = f'({depthtotal_records} records, {round(depthminutepercent * 100,2)} % of minutes meet target of 5.0 - 6.0)'
    if depth_df1.empty:
        AdultUpsteinNextWorksheet[
        'F357'] = 'Chart has no data for period: N/A'
        AdultUpsteinWorksheet['B193'] = ''
        AdultUpsteinWorksheet['C193'] = ''
        AdultUpsteinWorksheet['D193'] = ''


    AdultUstenworkbook.save("C:/Charts/AdultUsteinGraphs/CanRoc/Adult_Utstein_" + site + ".xlsx")

    allfields = [ 'cr_scause','cr_tx', 'cr_pyhalt', 'cr_witbys', 'cr_estageu', 'cr_estagev', 'cr_agecat',
                 'cr_scause', 'cr_cpratt', 'cr_loctyp', 'cr_aedapp', 'cr_aedshk', 'cr_frhyem',
                 'cr_surv', 'cr_rosc', 'cr_pdisp', 'cr_numshk', 'cr_ptmrcv', 'cr_rig1tm', 'cr_ptmdsp', 'cr_rig1dtm',
                 'cr_epi', 'cr_arryth', 'cr_lmasuc', 'cr_igelsuc',
                 'cr_ptmrcv', 'cr_ivtm', 'cr_iotm', 'cr_kingsuc', 'cr_ettsuc', 'cr_othawsuc', 'cr_rig1tm', 'cr_rig2tm',
                 'cr_rig3tm', 'cr_rig4tm', 'cr_v1sl', 'cr_v2sl', 'cr_v3sl',
                 'cr_v4sl', 'cr_lmatm', 'cr_kingtm', 'cr_igeltm', 'cr_etttm', 'cr_othawtm', 'cr_epitm', 'cr_arrythtm','cr_ptmcpr']

    df = project.export_records(format='df', fields=allfields)

    df = df[(df['cr_scause'] != 17)]
    df = df.reset_index()
    print(df)




    #Last Chart - Intervals

    ATimeCalculation = []
    BlueTimeCalculation = []
    OrangeTimeCalculation =[]
    GrayTimeCalculation = []
    YellowTimeCalculation  = []
    print('#Last Chart - Intervals')

    for index, row in df.iterrows():
        print('check initial time data')
        print(row["cr_ptmrcv"], row["cr_iotm"],row["cr_ivtm"] )

        if (str(row["cr_ptmrcv"]) != 'nan' and ((str(row["cr_iotm"]) != 'nan') or str(row["cr_ivtm"]) != 'nan')):
            Avalue = 0
            Bvalue = 0
            subtractedvalue = 0
            timediffblue = 0
            if (str(row["cr_iotm"]) != 'nan' or  (str(row["cr_ivtm"]) != 'nan')):
                if (str(row["cr_iotm"]) != 'nan' and  (str(row["cr_ivtm"]) == 'nan')):
                    subtractedvalue = datetime.strptime(row["cr_iotm"], "%Y-%m-%d %H:%M:%S")
                elif(str(row["cr_iotm"]) == 'nan' and  (str(row["cr_ivtm"]) != 'nan')):
                    subtractedvalue = datetime.strptime(row["cr_ivtm"], "%Y-%m-%d %H:%M:%S")
                elif (str(row["cr_iotm"]) != 'nan' and  (str(row["cr_ivtm"]) != 'nan')):
                    value_iotm = datetime.strptime(row["cr_iotm"], "%Y-%m-%d %H:%M:%S")
                    value_ivtm = datetime.strptime(row["cr_ivtm"], "%Y-%m-%d %H:%M:%S")
                    timediff =  value_iotm - value_ivtm
                    print('value_iotm')

                    print(value_iotm)
                    print('value_ivtm')
                    print(value_ivtm)
                    print('timediff')
                    print(timediff)
                    if timediff.total_seconds() < 0:
                        subtractedvalue = value_iotm
                    else:
                        subtractedvalue = value_ivtm
            Avalue = subtractedvalue
            print('Avalue')
            print(Avalue)

            for i in range(1,5):
                print(i)
                print('rigxtm and vsxl')
                print(row["cr_rig"+str(i)+"tm"])
                print(row["cr_v"+str(i)+"sl"])
                if str(row["cr_rig"+str(i)+"tm"]) != 'nan':# and str(row["cr_v"+str(i)+"sl"]) in ['1','2','1.0','2.0']:
                    Bvalue = datetime.strptime(row["cr_rig"+str(i)+"tm"], "%Y-%m-%d %H:%M:%S")
                    print('yes if works')
                    break

            if Avalue != 0 and Bvalue!= 0:
                timediffblue = (Avalue - Bvalue).total_seconds()/60
                if timediffblue > 0 and timediffblue <= 70:
                # BlueTimeCalculation.append(row["cr_record_id"])
                # BlueTimeCalculation.append(row["cr_ivtm"])
                # BlueTimeCalculation.append(row["cr_iotm"])
                # BlueTimeCalculation.append(row["cr_rig1tm"])
                # BlueTimeCalculation.append(row["cr_v1sl"])
                # BlueTimeCalculation.append(row["cr_rig2tm"])
                # BlueTimeCalculation.append(row["cr_v2sl"])
                # BlueTimeCalculation.append(row["cr_rig3tm"])
                # BlueTimeCalculation.append(row["cr_v3sl"])
                # BlueTimeCalculation.append(row["cr_rig4tm"])
                # BlueTimeCalculation.append(row["cr_v4sl"])
                    BlueTimeCalculation.append(timediffblue)



            # print(row["cr_ptmrcv"], row["cr_iotm"], row["cr_ivtm"],Avalue,Bvalue,timediffblue)


        if (str(row["cr_ptmrcv"]) != 'nan' and (str(row["cr_lmatm"]) != 'nan' or str(row["cr_kingtm"]) != 'nan'
                                        or str(row["cr_igeltm"]) != 'nan' or str(row["cr_etttm"]) != 'nan' or str(row["cr_othawtm"]) != 'nan'  )):
            Avalue = 0
            Bvalue = 0

            smallesttimediff = None

            cr_ptmrcv = datetime.strptime(str(row["cr_ptmrcv"]), "%Y-%m-%d %H:%M:%S")
            if str(row["cr_lmatm"])!='nan':
                cr_lmatm = datetime.strptime(str(row["cr_lmatm"]), "%Y-%m-%d %H:%M:%S")
                smallesttimediff =  cr_lmatm
                # OrangeTimeCalculationList = (row["cr_lmatm"]-(row["cr_ptmrcv"]

            if str(row["cr_kingtm"]) != 'nan':
                cr_kingtm = datetime.strptime(row["cr_kingtm"], "%Y-%m-%d %H:%M:%S")
                if smallesttimediff != None:
                    if smallesttimediff > cr_kingtm:
                        smallesttimediff = cr_kingtm
                elif smallesttimediff == None:
                    smallesttimediff = cr_kingtm
            if str(row["cr_igeltm"]) != 'nan':
                cr_igeltm = datetime.strptime(row["cr_igeltm"], "%Y-%m-%d %H:%M:%S")
                if smallesttimediff != None:
                    if smallesttimediff > cr_igeltm:
                        smallesttimediff = cr_igeltm
                elif smallesttimediff == None:
                    smallesttimediff == cr_igeltm
            if str(row["cr_etttm"]) != 'nan':
                cr_etttm = datetime.strptime(row["cr_etttm"], "%Y-%m-%d %H:%M:%S")
                if smallesttimediff!= None:
                    if smallesttimediff > cr_etttm:
                        smallesttimediff = cr_etttm
                elif smallesttimediff == None:
                    smallesttimediff = cr_etttm
            if str(row["cr_othawtm"]) != 'nan':
                cr_othawtm = datetime.strptime(row["cr_othawtm"], "%Y-%m-%d %H:%M:%S")
                if smallesttimediff != None:
                    if smallesttimediff > cr_othawtm:
                        smallesttimediff = cr_othawtm
                elif smallesttimediff == None:
                    smallesttimediff = cr_othawtm
            if smallesttimediff !=None:
                Avalue = smallesttimediff

            for i in range(1,5):
                print(i)
                print('rigxtm and vsxl')
                print(row["cr_rig"+str(i)+"tm"])
                print(row["cr_v"+str(i)+"sl"])
                if str(row["cr_rig"+str(i)+"tm"]) != 'nan' and str(row["cr_v"+str(i)+"sl"]) in ['1','2','1.0','2.0']:
                    Bvalue = datetime.strptime(row["cr_rig"+str(i)+"tm"], "%Y-%m-%d %H:%M:%S")
                    print('yes if works')
                    break


            if Avalue != 0 and Bvalue!= 0:

                # print(Avalue)
                # print(Bvalue)
                timedifforange = (Avalue - Bvalue).total_seconds()/60
                # if timedifforange > 0 and timedifforange <= 70:
                if timedifforange > 0 and timedifforange <= 70:
                    # OrangeTimeCalculation.append(row["cr_record_id"])

                    OrangeTimeCalculation.append(timedifforange)



        if (str(row["cr_ptmrcv"]) != 'nan' and (str(row["cr_epitm"]) != 'nan' or str(row["cr_arrythtm"]) != 'nan')):
            Avalue = 0
            Bvalue = 0
            subtractedvalue = 0
            timediffgray = 0
            if (str(row["cr_epitm"]) != 'nan' or (str(row["cr_arrythtm"]) != 'nan')):
                if (str(row["cr_epitm"]) != 'nan' and (str(row["cr_arrythtm"]) == 'nan')):
                    subtractedvalue = datetime.strptime(row["cr_epitm"], "%Y-%m-%d %H:%M:%S")
                elif (str(row["cr_epitm"]) == 'nan' and (str(row["cr_arrythtm"]) != 'nan')):
                    subtractedvalue = datetime.strptime(row["cr_arrythtm"], "%Y-%m-%d %H:%M:%S")
                elif (str(row["cr_epitm"]) != 'nan' and (str(row["cr_arrythtm"]) != 'nan')):
                    value_epitm = datetime.strptime(row["cr_epitm"], "%Y-%m-%d %H:%M:%S")
                    value_arrythtm = datetime.strptime(row["cr_arrythtm"], "%Y-%m-%d %H:%M:%S")
                    timediff = value_epitm - value_arrythtm
                    print('value_arrythtm')

                    print(value_arrythtm)
                    print('value_epitm')
                    print(value_epitm)
                    print('timediff')
                    print(timediff)
                    if timediff.total_seconds() < 0:
                        subtractedvalue = value_epitm
                    else:
                        subtractedvalue = value_arrythtm
            Avalue = subtractedvalue
            print('Avalue')
            print(Avalue)

            for i in range(1, 5):
                print(i)
                print('rigxtm and vsxl')
                print(row["cr_rig" + str(i) + "tm"])
                print(row["cr_v" + str(i) + "sl"])
                if str(row["cr_rig" + str(i) + "tm"]) != 'nan' and str(row["cr_v" + str(i) + "sl"]) in [ '2','2.0']:
                    Bvalue = datetime.strptime(row["cr_rig" + str(i) + "tm"], "%Y-%m-%d %H:%M:%S")
                    print('yes if works')
                    break

            if Avalue != 0 and Bvalue != 0:
                timediffgray = (Avalue - Bvalue).total_seconds() / 60
                if timediffgray > 0 and timediffgray <= 70:

                    # GrayTimeCalculation.append(row["cr_record_id"])
                    GrayTimeCalculation.append(timediffgray)

            if Avalue!=0 and (str(row["cr_ptmcpr"]) != 'nan'):
                cr_ptmcpr = datetime.strptime(row["cr_ptmcpr"], "%Y-%m-%d %H:%M:%S")

                timediffyellow =  (Avalue - cr_ptmcpr).total_seconds() / 60
                if timediffyellow  > 0 and timediffyellow  <= 70:


                    YellowTimeCalculation.append(timediffyellow)

    print('YellowTimeCalculation\n')
    for i in range( 1 , 300):
        print(YellowTimeCalculation[i])




    bluegraphmedian = np.nanmedian(BlueTimeCalculation)
    bluegraphtenpercentile = np.nanpercentile(BlueTimeCalculation, 10)
    bluegraphpercentile = np.nanpercentile(BlueTimeCalculation, 90)

    orangegraphmedian = np.nanmedian(OrangeTimeCalculation)
    orangegraphtenpercentile = np.nanpercentile(OrangeTimeCalculation, 10)
    orangegraphpercentile = np.nanpercentile(OrangeTimeCalculation, 90)

    graygraphmedian = np.nanmedian(GrayTimeCalculation)
    graygraphtenpercentile = np.nanpercentile(GrayTimeCalculation, 10)
    graygraphpercentile = np.nanpercentile(GrayTimeCalculation, 90)

    yellowgraphmedian = np.nanmedian(YellowTimeCalculation)
    yellowgraphtenpercentile = np.nanpercentile(YellowTimeCalculation, 10)
    yellowgraphpercentile = np.nanpercentile(YellowTimeCalculation, 90)


    AdultUpsteinWorksheet['B225'] = round(bluegraphmedian, 4)
    AdultUpsteinWorksheet['C225'] = round( bluegraphtenpercentile, 4)
    AdultUpsteinWorksheet['D225'] = round(bluegraphpercentile, 4)

    AdultUpsteinWorksheet['A226'] = round(orangegraphmedian, 4)
    AdultUpsteinWorksheet['B226'] = round(orangegraphmedian, 4)
    AdultUpsteinWorksheet['C226'] = round(orangegraphtenpercentile, 4)
    AdultUpsteinWorksheet['D226'] = round(orangegraphpercentile, 4)

    AdultUpsteinWorksheet['A227'] = round(graygraphmedian, 4)
    AdultUpsteinWorksheet['B227'] = round(graygraphmedian, 4)
    AdultUpsteinWorksheet['C227'] = round(graygraphtenpercentile, 4)
    AdultUpsteinWorksheet['D227'] = round(graygraphpercentile, 4)



    x = datetime.now()
    AdultUpsteinNextWorksheet['N4'] = x.strftime("%B %d, %Y %I:%M %p")
    AdultUpsteinNextWorksheet['N5'] = f'{start_date_changed.strftime("%B %d, %Y")} to {end_date_changed.strftime("%B %d, %Y")}'

    siteregion = site
    if site == 'ALL':
        siteregion = ' ALL REGIONS'
    AdultUpsteinNextWorksheet['B4'] = "Region: " + siteregion


    print('#End of Last Chart - Intervals')

    AdultUstenworkbook.save("C:/Users/mcbarnettr/Desktop/Charts/AdultUsteinGraphs/CanRoc/Adult_Utstein_" + site + ".xlsx")


    AdultUstenworkbook.close()
    print('Antiarryth_total')
    print(Antiarryth_total)
    print(Antiarryth_Ntotal)




#
# sites = ['BC','SK','AL','PEI','TO','MTL',TT']
sites = ['BC','TO','OTT','PEI','SK','AL']
sites = ['SK']
answer = input('Do you want to print the National Charts or Regional charts? "Y" or "N"').upper()
if answer == 'Y':
    print('Going to do National charts')
    CreateSiteLevelCharts('ALL')
else:
    print('Going to do Regional charts')
    for site in sites:
        CreateSiteLevelCharts(site)


plt.close()
